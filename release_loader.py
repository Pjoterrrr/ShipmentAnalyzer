from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any

import pandas as pd
from openpyxl import load_workbook


LEGACY_REQUIRED_RAW_COLUMNS = [
    "PO Number",
    "PO Line #",
    "Release Version",
    "Release Date",
    "Part Number",
    "Part Description",
    "Ship Date",
    "Receipt Date",
    "Open Quantity",
    "Unit of Measure",
]

NORMALIZED_COLUMNS = [
    "source_file",
    "snapshot_date",
    "origin_doc",
    "item",
    "ship_to",
    "material",
    "description",
    "unrestr_qty",
    "unl_point",
    "customer_material",
    "po_number",
    "gi_date",
    "delivery_date",
    "open_qty",
    "cum_qty",
    "unit",
]

COMPARISON_SCHEMA_MAP = {
    "origin_doc": "Origin Doc",
    "item": "Item",
    "ship_to": "Ship To",
    "material": "Part Number",
    "description": "Part Description",
    "unrestr_qty": "Unrestricted Qty",
    "unl_point": "Unloading Point",
    "customer_material": "Customer Material",
    "po_number": "PO Number",
    "gi_date": "Ship Date",
    "delivery_date": "Receipt Date",
    "open_qty": "Open Quantity",
    "cum_qty": "CumQty",
    "unit": "Unit of Measure",
    "snapshot_date": "Snapshot Date",
    "source_file": "Source File",
}

COMPARISON_COLUMN_ORDER = [
    "Source File",
    "Snapshot Date",
    "PO Number",
    "Origin Doc",
    "Item",
    "Ship To",
    "Part Number",
    "Part Description",
    "Customer Material",
    "Unrestricted Qty",
    "Unloading Point",
    "Ship Date",
    "Receipt Date",
    "Open Quantity",
    "CumQty",
    "Unit of Measure",
    "Release Version",
    "Release Date",
]


def first_non_empty(series: pd.Series, default: str = "n/a") -> str:
    values = series.dropna().astype(str).str.strip()
    values = values[values.ne("") & values.ne("nan") & values.ne("NaT")]
    return values.iloc[0] if not values.empty else default


def _ensure_bytes(file: bytes | bytearray | io.BytesIO | Any) -> bytes:
    if isinstance(file, bytes):
        return file
    if isinstance(file, bytearray):
        return bytes(file)
    if hasattr(file, "getvalue"):
        return bytes(file.getvalue())
    if hasattr(file, "read"):
        current_position = file.tell() if hasattr(file, "tell") else None
        content = file.read()
        if current_position is not None and hasattr(file, "seek"):
            file.seek(current_position)
        return bytes(content)
    raise TypeError("Unsupported file object supplied to release loader.")


def _sanitize_columns(columns: list[Any]) -> list[str]:
    return [str(column).replace("? ", "").strip() for column in columns]


def _normalize_text(value: Any) -> str | pd.NA:
    if value is None:
        return pd.NA
    text = str(value).strip()
    if text == "" or text.lower() in {"nan", "nat", "none"}:
        return pd.NA
    return text


def _to_number(value: Any) -> float | pd.NA:
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(numeric):
        return pd.NA
    return float(numeric)


def _to_timestamp(value: Any) -> pd.Timestamp | pd.NaT:
    if value is None or value == "":
        return pd.NaT
    if isinstance(value, pd.Timestamp):
        return value
    if isinstance(value, datetime):
        return pd.Timestamp(value)
    if isinstance(value, date):
        return pd.Timestamp(value)
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=False)
    return parsed if not pd.isna(parsed) else pd.NaT


def _is_blank_row(values: tuple[Any, ...]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, float) and pd.isna(value):
            continue
        if str(value).strip() != "":
            return False
    return True


def _is_numeric_like(value: Any) -> bool:
    return pd.notna(pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0])


def _is_date_like(value: Any) -> bool:
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return True
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return False
    return not pd.isna(pd.to_datetime(value, errors="coerce", dayfirst=False))


def _extract_snapshot_date(file_name: str, fallback: Any = None) -> pd.Timestamp | pd.NaT:
    patterns = [
        (r"(\d{2})\.(\d{2})\.(\d{4})", "%d.%m.%Y"),
        (r"(\d{4})-(\d{2})-(\d{2})", "%Y-%m-%d"),
        (r"(\d{4})_(\d{2})_(\d{2})", "%Y_%m_%d"),
    ]
    for pattern, fmt in patterns:
        match = re.search(pattern, file_name)
        if match:
            try:
                return pd.Timestamp(datetime.strptime(match.group(0), fmt))
            except ValueError:
                pass
    parsed_fallback = _to_timestamp(fallback)
    return parsed_fallback if not pd.isna(parsed_fallback) else pd.NaT


def _read_excel_file(file_bytes: bytes) -> pd.ExcelFile:
    try:
        return pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
    except Exception as exc:
        raise ValueError(
            "Nie udało się odczytać pliku Excel. Upewnij się, że plik nie jest uszkodzony i ma poprawny format .xlsx."
        ) from exc


def _inspect_workbook(file: bytes | bytearray | io.BytesIO | Any) -> dict[str, Any]:
    file_bytes = _ensure_bytes(file)
    excel_file = _read_excel_file(file_bytes)
    sheet_names = list(excel_file.sheet_names)
    if not sheet_names:
        raise ValueError("Plik Excel nie zawiera żadnych arkuszy.")

    # Priorytetem jest pełna kompatybilność wsteczna: jeśli istnieje arkusz Raw,
    # zachowujemy dotychczasowe zachowanie. Gdy Raw nie ma, bierzemy pierwszy
    # dostępny arkusz, aby obsłużyć nowsze pliki VL10E i inne poprawne eksporty.
    selected_sheet = "Raw" if "Raw" in sheet_names else sheet_names[0]
    return {
        "file_bytes": file_bytes,
        "excel_file": excel_file,
        "sheet_names": sheet_names,
        "selected_sheet": selected_sheet,
        "has_raw_sheet": "Raw" in sheet_names,
    }


def _read_sheet(
    workbook_info: dict[str, Any],
    sheet_name: str,
    **kwargs: Any,
) -> pd.DataFrame:
    try:
        dataframe = pd.read_excel(
            workbook_info["excel_file"],
            sheet_name=sheet_name,
            **kwargs,
        )
    except ValueError as exc:
        raise ValueError(
            f"Nie udało się odczytać arkusza '{sheet_name}' z pliku Excel."
        ) from exc
    except Exception as exc:
        raise ValueError(
            f"Wystąpił błąd podczas odczytu arkusza '{sheet_name}' z pliku Excel."
        ) from exc

    if isinstance(dataframe, dict):
        raise ValueError(f"Nieoczekiwany wynik podczas odczytu arkusza '{sheet_name}'.")
    return dataframe


def _read_overview_sheet(workbook_info: dict[str, Any]) -> pd.DataFrame:
    try:
        overview_df = pd.read_excel(workbook_info["excel_file"], sheet_name=0, header=4)
    except Exception:
        return pd.DataFrame()

    if isinstance(overview_df, dict):
        return pd.DataFrame()
    overview_df.columns = _sanitize_columns(list(overview_df.columns))
    return overview_df


def _looks_like_vl10e_block(file_bytes: bytes, sheet_name: str | None = None) -> bool:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if not workbook.sheetnames:
        return False
    worksheet_name = sheet_name or workbook.sheetnames[0]
    worksheet = workbook[worksheet_name]
    master_rows = 0
    detail_rows = 0

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_index > 400:
            break
        if _is_blank_row(row):
            continue

        cell_b = row[1] if len(row) > 1 else None
        cell_c = row[2] if len(row) > 2 else None
        cell_e = row[4] if len(row) > 4 else None
        cell_f = row[5] if len(row) > 5 else None

        material_value = _normalize_text(cell_f)
        if _is_numeric_like(cell_b) and not _is_date_like(cell_b) and not pd.isna(material_value):
            master_rows += 1
        elif _is_date_like(cell_b) and _is_date_like(cell_c) and _is_numeric_like(cell_e):
            detail_rows += 1

        if master_rows > 0 and detail_rows > 0:
            return True

    return False


def detect_file_type(
    file: bytes | bytearray | io.BytesIO | Any,
    workbook_info: dict[str, Any] | None = None,
) -> str:
    workbook_info = workbook_info or _inspect_workbook(file)
    file_bytes = workbook_info["file_bytes"]
    selected_sheet = workbook_info["selected_sheet"]

    try:
        raw_preview = _read_sheet(workbook_info, selected_sheet, nrows=5)
        raw_preview.columns = _sanitize_columns(list(raw_preview.columns))
        if all(column in raw_preview.columns for column in LEGACY_REQUIRED_RAW_COLUMNS):
            return "legacy_wide"
    except Exception:
        pass

    try:
        if _looks_like_vl10e_block(file_bytes, sheet_name=selected_sheet):
            return "vl10e_block"
    except Exception:
        pass

    available_sheets = ", ".join(workbook_info["sheet_names"])
    raise ValueError(
        "Nie rozpoznano struktury pliku Excel. "
        f"Sprawdzono arkusz '{selected_sheet}' "
        f"(dostępne arkusze: {available_sheets}). "
        "Plik nie pasuje ani do starego formatu legacy, ani do nowego formatu VL10E."
    )


def parse_legacy_wide(
    file: bytes | bytearray | io.BytesIO | Any,
    sheet_name: str | None = None,
    workbook_info: dict[str, Any] | None = None,
) -> dict[str, pd.DataFrame]:
    workbook_info = workbook_info or _inspect_workbook(file)
    selected_sheet = sheet_name or workbook_info["selected_sheet"]
    raw_df = _read_sheet(workbook_info, selected_sheet)
    raw_df.columns = _sanitize_columns(list(raw_df.columns))

    missing_columns = [column for column in LEGACY_REQUIRED_RAW_COLUMNS if column not in raw_df.columns]
    if missing_columns:
        raise ValueError(
            f"Arkusz '{selected_sheet}' nie zawiera wymaganych kolumn starego formatu: "
            + ", ".join(missing_columns)
        )

    overview_df = _read_overview_sheet(workbook_info)
    return {"raw": raw_df, "overview": overview_df, "raw_sheet_name": selected_sheet}


def parse_vl10e_block(
    file: bytes | bytearray | io.BytesIO | Any,
    sheet_name: str | None = None,
    workbook_info: dict[str, Any] | None = None,
) -> pd.DataFrame:
    workbook_info = workbook_info or _inspect_workbook(file)
    file_bytes = workbook_info["file_bytes"]
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if not workbook.sheetnames:
        raise ValueError("Plik Excel nie zawiera żadnych arkuszy.")
    selected_sheet = sheet_name or workbook_info["selected_sheet"]
    worksheet = workbook[selected_sheet]

    current_master: dict[str, Any] | None = None
    rows: list[dict[str, Any]] = []

    for row in worksheet.iter_rows(values_only=True):
        if _is_blank_row(row):
            continue

        cell_b = row[1] if len(row) > 1 else None
        cell_c = row[2] if len(row) > 2 else None
        cell_d = row[3] if len(row) > 3 else None
        cell_e = row[4] if len(row) > 4 else None
        cell_f = row[5] if len(row) > 5 else None
        cell_g = row[6] if len(row) > 6 else None
        cell_h = row[7] if len(row) > 7 else None
        cell_i = row[8] if len(row) > 8 else None
        cell_j = row[9] if len(row) > 9 else None
        cell_k = row[10] if len(row) > 10 else None
        cell_l = row[11] if len(row) > 11 else None
        cell_m = row[12] if len(row) > 12 else None

        material_value = _normalize_text(cell_f)
        unit_g = _normalize_text(cell_g)
        unit_i = _normalize_text(cell_i)
        is_master_row = (
            _is_numeric_like(cell_b)
            and not _is_date_like(cell_b)
            and not pd.isna(material_value)
        )
        is_detail_row = _is_date_like(cell_b) and _is_date_like(cell_c) and _is_numeric_like(cell_e)

        if is_master_row:
            current_master = {
                "origin_doc": _normalize_text(cell_b),
                "item": _normalize_text(cell_c),
                "ship_to": _normalize_text(cell_d),
                "material": _normalize_text(cell_f),
                "description": _normalize_text(cell_h),
                "unrestr_qty": _to_number(cell_j),
                "unl_point": _normalize_text(cell_k),
                "customer_material": _normalize_text(cell_l),
                "po_number": _normalize_text(cell_m),
            }
            continue

        if is_detail_row and current_master is not None:
            rows.append(
                {
                    **current_master,
                    "gi_date": _to_timestamp(cell_b),
                    "delivery_date": _to_timestamp(cell_c),
                    "open_qty": _to_number(cell_e),
                    "cum_qty": _to_number(cell_h),
                    "unit": unit_g if not pd.isna(unit_g) else unit_i,
                }
            )

    if not rows:
        raise ValueError("No VL10E detail rows were detected in the workbook.")

    return pd.DataFrame(rows)


def normalize_data(
    parsed_data: pd.DataFrame | dict[str, pd.DataFrame],
    file_type: str,
    source_file: str,
    snapshot_date: Any = None,
) -> pd.DataFrame:
    normalized_snapshot = _to_timestamp(snapshot_date)

    if file_type == "legacy_wide":
        raw_df = parsed_data["raw"].copy()
        raw_df["PO Number"] = raw_df["PO Number"].astype(str).str.strip()
        raw_df["PO Line #"] = raw_df["PO Line #"].astype(str).str.strip()
        raw_df["Part Number"] = raw_df["Part Number"].astype(str).str.strip()
        raw_df["Part Description"] = raw_df["Part Description"].astype(str).str.strip()
        raw_df["Unit of Measure"] = raw_df["Unit of Measure"].astype(str).str.strip()
        raw_df["Release Date"] = pd.to_datetime(raw_df["Release Date"], errors="coerce")
        raw_df["Ship Date"] = pd.to_datetime(raw_df["Ship Date"], errors="coerce")
        raw_df["Receipt Date"] = pd.to_datetime(raw_df["Receipt Date"], errors="coerce")
        raw_df["Open Quantity"] = pd.to_numeric(raw_df["Open Quantity"], errors="coerce")

        if pd.isna(normalized_snapshot):
            normalized_snapshot = raw_df["Release Date"].dropna().min()

        normalized = pd.DataFrame(
            {
                "source_file": source_file,
                "snapshot_date": normalized_snapshot,
                "origin_doc": pd.NA,
                "item": raw_df["PO Line #"].replace("", pd.NA),
                "ship_to": pd.NA,
                "material": raw_df["Part Number"].replace("", pd.NA),
                "description": raw_df["Part Description"].replace("", pd.NA),
                "unrestr_qty": pd.NA,
                "unl_point": pd.NA,
                "customer_material": pd.NA,
                "po_number": raw_df["PO Number"].replace("", pd.NA),
                "gi_date": raw_df["Ship Date"],
                "delivery_date": raw_df["Receipt Date"],
                "open_qty": raw_df["Open Quantity"].fillna(0),
                "cum_qty": pd.NA,
                "unit": raw_df["Unit of Measure"].replace("", pd.NA),
            }
        )
    elif file_type == "vl10e_block":
        vl10e_df = parsed_data.copy()
        for column in [
            "origin_doc",
            "item",
            "ship_to",
            "material",
            "description",
            "unl_point",
            "customer_material",
            "po_number",
            "unit",
        ]:
            if column in vl10e_df.columns:
                vl10e_df[column] = vl10e_df[column].map(_normalize_text)

        for column in ["unrestr_qty", "open_qty", "cum_qty"]:
            if column in vl10e_df.columns:
                vl10e_df[column] = pd.to_numeric(vl10e_df[column], errors="coerce")

        for column in ["gi_date", "delivery_date"]:
            if column in vl10e_df.columns:
                vl10e_df[column] = pd.to_datetime(vl10e_df[column], errors="coerce")

        if pd.isna(normalized_snapshot):
            delivery_fallback = vl10e_df["delivery_date"].dropna().min() if "delivery_date" in vl10e_df else pd.NaT
            gi_fallback = vl10e_df["gi_date"].dropna().min() if "gi_date" in vl10e_df else pd.NaT
            normalized_snapshot = delivery_fallback if not pd.isna(delivery_fallback) else gi_fallback

        normalized = vl10e_df.reindex(columns=[column for column in NORMALIZED_COLUMNS if column in vl10e_df.columns]).copy()
        normalized["source_file"] = source_file
        normalized["snapshot_date"] = normalized_snapshot
        if "description" in normalized.columns:
            normalized["description"] = normalized["description"].fillna(normalized.get("material"))
    else:
        raise ValueError(f"Unsupported file type: {file_type}")

    normalized = normalized.reindex(columns=NORMALIZED_COLUMNS)
    normalized["snapshot_date"] = pd.to_datetime(normalized["snapshot_date"], errors="coerce")
    normalized["gi_date"] = pd.to_datetime(normalized["gi_date"], errors="coerce")
    normalized["delivery_date"] = pd.to_datetime(normalized["delivery_date"], errors="coerce")
    normalized["open_qty"] = pd.to_numeric(normalized["open_qty"], errors="coerce").fillna(0)
    normalized["unrestr_qty"] = pd.to_numeric(normalized["unrestr_qty"], errors="coerce")
    normalized["cum_qty"] = pd.to_numeric(normalized["cum_qty"], errors="coerce")

    normalized = normalized.dropna(subset=["material", "gi_date", "delivery_date"]).copy()
    normalized["description"] = normalized["description"].fillna(normalized["material"])
    return normalized.reset_index(drop=True)


def _build_release_version(
    file_type: str,
    file_name: str,
    parsed_data: pd.DataFrame | dict[str, pd.DataFrame],
    snapshot_date: pd.Timestamp | pd.NaT,
) -> str:
    if file_type == "legacy_wide":
        raw_df = parsed_data["raw"]
        return first_non_empty(raw_df["Release Version"])

    if not pd.isna(snapshot_date):
        return pd.Timestamp(snapshot_date).strftime("%Y-%m-%d")
    return file_name


def load_release(file_bytes: bytes, file_name: str) -> tuple[pd.DataFrame, dict[str, Any]]:
    workbook_info = _inspect_workbook(file_bytes)
    file_type = detect_file_type(file_bytes, workbook_info=workbook_info)
    selected_sheet = workbook_info["selected_sheet"]
    parsed_data = (
        parse_legacy_wide(file_bytes, sheet_name=selected_sheet, workbook_info=workbook_info)
        if file_type == "legacy_wide"
        else parse_vl10e_block(file_bytes, sheet_name=selected_sheet, workbook_info=workbook_info)
    )
    snapshot_date = _extract_snapshot_date(file_name)
    normalized_df = normalize_data(parsed_data, file_type=file_type, source_file=file_name, snapshot_date=snapshot_date)
    normalized_snapshot = normalized_df["snapshot_date"].dropna().min()

    comparison_df = normalized_df.rename(columns=COMPARISON_SCHEMA_MAP).copy()
    comparison_df["Release Date"] = normalized_snapshot
    comparison_df["Release Version"] = _build_release_version(
        file_type, file_name, parsed_data, normalized_snapshot
    )
    comparison_df = comparison_df.reindex(columns=COMPARISON_COLUMN_ORDER)

    overview_df = parsed_data["overview"] if file_type == "legacy_wide" else pd.DataFrame()
    po_value = (
        first_non_empty(comparison_df["PO Number"])
        if "PO Number" in comparison_df
        else "n/a"
    )
    if po_value == "n/a" and "Origin Doc" in comparison_df:
        po_value = first_non_empty(comparison_df["Origin Doc"])

    metadata = {
        "file_name": file_name,
        "file_type": file_type,
        "sheet_name": selected_sheet,
        "sheet_names": workbook_info["sheet_names"],
        "po_number": po_value,
        "release_version": first_non_empty(comparison_df["Release Version"]),
        "release_date": normalized_snapshot,
        "planner_name": (
            first_non_empty(overview_df["Planner Name"])
            if "Planner Name" in overview_df.columns
            else "n/a"
        ),
        "planner_email": (
            first_non_empty(overview_df["Planner Email"])
            if "Planner Email" in overview_df.columns
            else "n/a"
        ),
        "products": comparison_df["Part Number"].nunique(),
        "rows": len(comparison_df),
    }
    return comparison_df, metadata


def _has_meaningful_values(*series_collection: pd.Series) -> bool:
    combined = pd.concat(series_collection, ignore_index=True)
    non_empty = combined.dropna().astype(str).str.strip()
    non_empty = non_empty[non_empty.ne("") & non_empty.ne("n/a") & non_empty.ne("nan") & non_empty.ne("NaT")]
    return not non_empty.empty


def _comparison_keys(prev_df: pd.DataFrame, curr_df: pd.DataFrame) -> list[str]:
    candidate_keys = [
        "PO Number",
        "Origin Doc",
        "Item",
        "Ship To",
        "Part Number",
        "Part Description",
        "Customer Material",
        "Unloading Point",
        "Ship Date",
        "Receipt Date",
    ]
    keys: list[str] = []
    for column in candidate_keys:
        if column in prev_df.columns and column in curr_df.columns:
            if _has_meaningful_values(prev_df[column], curr_df[column]):
                keys.append(column)
    if "Part Number" not in keys:
        keys.append("Part Number")
    if "Part Description" not in keys and "Part Description" in prev_df.columns and "Part Description" in curr_df.columns:
        keys.append("Part Description")
    if "Ship Date" not in keys:
        keys.append("Ship Date")
    if "Receipt Date" not in keys:
        keys.append("Receipt Date")
    return keys


def _safe_percent_change(current_value: float, previous_value: float) -> float:
    if previous_value == 0:
        return 100.0 if current_value > 0 else 0.0
    return round(((current_value - previous_value) / previous_value) * 100, 2)


def _demand_status(previous_qty: float, current_qty: float) -> str:
    if previous_qty == 0 and current_qty > 0:
        return "NEW DEMAND"
    if previous_qty > 0 and current_qty == 0:
        return "REMOVED DEMAND"
    return ""


def compare_releases(prev_df: pd.DataFrame, curr_df: pd.DataFrame, threshold: float = 15) -> pd.DataFrame:
    keys = _comparison_keys(prev_df, curr_df)

    prev_summary = prev_df.groupby(keys, as_index=False, dropna=False).agg(
        Quantity_Prev=("Open Quantity", "sum"),
        UoM_Prev=("Unit of Measure", "first"),
        CumQty_Prev=("CumQty", "first"),
        UnrestrictedQty_Prev=("Unrestricted Qty", "first"),
        SnapshotDate_Prev=("Snapshot Date", "first"),
        SourceFile_Prev=("Source File", "first"),
    )
    curr_summary = curr_df.groupby(keys, as_index=False, dropna=False).agg(
        Quantity_Curr=("Open Quantity", "sum"),
        UoM_Curr=("Unit of Measure", "first"),
        CumQty_Curr=("CumQty", "first"),
        UnrestrictedQty_Curr=("Unrestricted Qty", "first"),
        SnapshotDate_Curr=("Snapshot Date", "first"),
        SourceFile_Curr=("Source File", "first"),
    )

    merged = prev_summary.merge(curr_summary, on=keys, how="outer")
    merged["Quantity_Prev"] = merged["Quantity_Prev"].fillna(0)
    merged["Quantity_Curr"] = merged["Quantity_Curr"].fillna(0)
    merged["Unit of Measure"] = merged["UoM_Prev"].combine_first(merged["UoM_Curr"])
    merged["CumQty"] = merged["CumQty_Curr"].combine_first(merged["CumQty_Prev"])
    merged["Unrestricted Qty"] = merged["UnrestrictedQty_Curr"].combine_first(
        merged["UnrestrictedQty_Prev"]
    )
    merged["Snapshot Date Previous"] = merged["SnapshotDate_Prev"]
    merged["Snapshot Date Current"] = merged["SnapshotDate_Curr"]
    merged["Source File Previous"] = merged["SourceFile_Prev"]
    merged["Source File Current"] = merged["SourceFile_Curr"]
    merged["Delta"] = merged["Quantity_Curr"] - merged["Quantity_Prev"]
    merged["Abs Delta"] = merged["Delta"].abs()
    merged["Percent Change"] = merged.apply(
        lambda row: _safe_percent_change(row["Quantity_Curr"], row["Quantity_Prev"]),
        axis=1,
    )
    merged["Alert"] = merged["Percent Change"].abs() >= float(threshold)
    merged["Change Direction"] = merged["Delta"].apply(
        lambda value: "Increase" if value > 0 else ("Decrease" if value < 0 else "No Change")
    )
    merged["Demand Status"] = merged.apply(
        lambda row: _demand_status(row["Quantity_Prev"], row["Quantity_Curr"]),
        axis=1,
    )
    merged["Product Label"] = (
        merged["Part Number"].fillna("").astype(str)
        + " | "
        + merged["Part Description"].fillna(merged["Part Number"]).astype(str)
    )
    merged = merged.drop(
        columns=[
            "UoM_Prev",
            "UoM_Curr",
            "CumQty_Prev",
            "CumQty_Curr",
            "UnrestrictedQty_Prev",
            "UnrestrictedQty_Curr",
            "SnapshotDate_Prev",
            "SnapshotDate_Curr",
            "SourceFile_Prev",
            "SourceFile_Curr",
        ]
    )
    return merged.sort_values(["Receipt Date", "Ship Date", "Part Description"]).reset_index(drop=True)
