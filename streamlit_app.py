import base64
import binascii
import hashlib
import html
import io
import json
from pathlib import Path
import sys
from types import SimpleNamespace
from xml.etree import ElementTree as ET
import altair as alt
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl.chart import BarChart, LineChart, Reference
from analytics_calendar import (
    build_calendar_frame,
    build_weekly_summary,
    classify_polish_day,
    get_last_completed_reference_week,
)
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from modules.admin import render as render_admin_module
from modules.context import ModuleDataContext
from modules.dashboard import render as render_dashboard_module
from modules.details import render as render_details_module
from modules.planner import render as render_planner_module
from modules.reports import render as render_reports_module
from modules.ui_shell import APP_TITLE, PRIMARY_TILES
from modules import ui_shell
from planner_helpers import (
    prepare_planner_source,
)
from release_loader import compare_releases as compare_release_frames
from release_loader import load_release as load_release_file


THRESHOLD = 15
MAX_MATRIX_STYLE_CELLS = 50000
BRAND_NAME = "Pjoter Development"
PRIMARY_VIEW_KEYS = {"dashboard", "reports"}
MAIN_VIEW_OPTIONS = ("dashboard", "reports")
FILE_VIEW_OPTIONS = {
    "overview": "Workspace",
    "planner": "Planner",
    "details": "Eksport i dane",
    "admin": "Admin",
}
PLOTLY_THEME = {
    "layout": {
        "paper_bgcolor": "rgba(0,0,0,0)",
        "plot_bgcolor": "rgba(0,0,0,0)",
        "font": {"family": "Inter, system-ui, sans-serif", "color": "#f0f6fc", "size": 12},
        "margin": {"l": 24, "r": 20, "t": 28, "b": 24},
        "hoverlabel": {
            "bgcolor": "#161b22",
            "bordercolor": "rgba(255,255,255,0.10)",
            "font": {"color": "#f0f6fc"},
        },
        "legend": {
            "orientation": "h",
            "yanchor": "bottom",
            "y": 1.02,
            "xanchor": "left",
            "x": 0,
            "font": {"color": "#8b949e"},
        },
        "xaxis": {
            "showgrid": True,
            "gridcolor": "rgba(255,255,255,0.06)",
            "zeroline": False,
            "linecolor": "rgba(255,255,255,0.08)",
            "tickfont": {"color": "#8b949e"},
            "title": {"font": {"color": "#8b949e"}},
        },
        "yaxis": {
            "showgrid": True,
            "gridcolor": "rgba(255,255,255,0.06)",
            "zeroline": False,
            "linecolor": "rgba(255,255,255,0.08)",
            "tickfont": {"color": "#8b949e"},
            "title": {"font": {"color": "#8b949e"}},
        },
    }
}
PLOTLY_CONFIG = {
    "displaylogo": False,
    "responsive": True,
    "scrollZoom": False,
    "modeBarButtonsToRemove": [
        "lasso2d",
        "select2d",
        "autoScale2d",
        "resetScale2d",
        "toggleSpikelines",
    ],
}
UPLOAD_STATE_KEYS = {
    "previous": "uploaded_previous_release",
    "current": "uploaded_current_release",
}
UPLOAD_NONCE_KEYS = {
    "previous": "uploaded_previous_release_nonce",
    "current": "uploaded_current_release_nonce",
}
BASE_DIR = Path(__file__).resolve().parent
RUNTIME_ROOT = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else BASE_DIR


def resolve_runtime_path(relative_path):
    try:
        external_path = RUNTIME_ROOT / relative_path
        internal_path = BASE_DIR / relative_path
        return external_path if external_path.exists() else internal_path
    except Exception:
        return BASE_DIR / relative_path


BRAND_LOGO_OVERRIDE_PATHS = [
    Path("assets") / "Nowe logo.png",
    Path("assets") / "aplikacja_analityczna_logo.png",
    Path("assets") / "aplikacja-analityczna-logo.png",
    Path("assets") / "brand_logo.png",
    Path("assets") / "branding_logo.png",
    Path("assets") / "logo_app.png",
]


def resolve_brand_logo_path():
    for relative_path in BRAND_LOGO_OVERRIDE_PATHS:
        candidate = resolve_runtime_path(relative_path)
        if candidate.exists():
            return candidate
    return resolve_runtime_path(Path("assets") / "logo.png")


LOGO_PATH = resolve_brand_logo_path()
REQUESTED_BRAND_LOGO_PRESENT = any(
    resolve_runtime_path(relative_path).exists() for relative_path in BRAND_LOGO_OVERRIDE_PATHS
)
AUTH_USERS_PATH = resolve_runtime_path(Path("config") / "users.json")
DATE_OPTIONS = ["Receipt Date", "Ship Date"]
DATE_LABELS = {
    "Receipt Date": "Data odbioru",
    "Ship Date": "Data wysyłki",
}
CHANGE_DIRECTION_LABELS = {
    "Increase": "Wzrost",
    "Decrease": "Spadek",
    "No Change": "Bez zmian",
}
MATRIX_METRIC_LABELS = {
    "Current Quantity": "Aktualna ilość",
    "Previous Quantity": "Poprzednia ilość",
    "Delta": "Zmiana ilości",
    "Percent Change": "Zmiana procentowa",
}
VIEW_MODE_LABELS = {
    "chart": "Wykres",
    "table": "Dane",
}
MODULE_OPTIONS = ["dashboard", "planner", "reports", "details", "admin"]
MODULE_LABELS = {
    "dashboard": "Dashboard",
    "planner": "Planner",
    "reports": "Reports",
    "details": "Details",
    "admin": "Admin",
}
ROLE_MODULE_PERMISSIONS = {
    "Admin": {
        "dashboard": "edit",
        "planner": "edit",
        "reports": "edit",
        "details": "edit",
        "admin": "edit",
    },
    "Planner": {
        "dashboard": "edit",
        "planner": "edit",
        "reports": "edit",
        "details": "edit",
    },
    "Viewer": {
        "dashboard": "read",
        "reports": "read",
    },
    "Warehouse": {
        "dashboard": "read",
        "planner": "read",
        "reports": "read",
    },
}


st.set_page_config(
    page_title="Pjoter Development | Analiza zamówień i wysyłek",
    layout="wide",
    initial_sidebar_state="collapsed",
)
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Manrope:wght@500;600;700;800&family=IBM+Plex+Sans:wght@400;500;600;700&display=swap');

    :root {
        --ink: #172033;
        --navy: #0f2742;
        --slate: #5c6678;
        --muted: #7b8798;
        --line: rgba(23, 32, 51, 0.09);
        --line-strong: rgba(23, 32, 51, 0.16);
        --canvas: #f4f7fb;
        --canvas-soft: #eef4fa;
        --panel: #f7faff;
        --panel-strong: #ffffff;
        --card: #ffffff;
        --card-soft: #fbfdff;
        --interactive: #f1f6fb;
        --interactive-strong: #e6f0fb;
        --mint: #1f8f64;
        --rose: #d25a5a;
        --steel: #1e88e5;
        --steel-soft: rgba(30, 136, 229, 0.12);
        --amber: #cf8a2e;
        --card-radius: 22px;
        --card-padding: 1.2rem 1.25rem;
        --panel-radius: 26px;
        --input-radius: 14px;
        --surface-shadow: 0 18px 40px rgba(18, 38, 63, 0.08);
        --surface-shadow-hover: 0 24px 52px rgba(18, 38, 63, 0.11);
        --surface-shadow-soft: 0 10px 24px rgba(18, 38, 63, 0.06);
        --focus-ring: 0 0 0 3px rgba(30, 136, 229, 0.16);
    }

    html, body, [class*="css"] {
        font-family: "IBM Plex Sans", "Segoe UI", sans-serif !important;
        color: var(--ink);
    }

    h1, h2, h3, h4, h5, h6 {
        font-family: "Manrope", "Segoe UI", sans-serif !important;
        color: var(--ink) !important;
        letter-spacing: -0.03em;
        line-height: 1.08;
        text-wrap: balance;
    }

    .stApp {
        background:
            radial-gradient(circle at top left, rgba(30, 136, 229, 0.10), transparent 24%),
            radial-gradient(circle at top right, rgba(34, 95, 155, 0.08), transparent 20%),
            linear-gradient(180deg, #fbfdff 0%, var(--canvas) 52%, var(--canvas-soft) 100%) !important;
        color: var(--ink) !important;
    }

    .block-container {
        max-width: 1540px !important;
        padding-top: 1.15rem !important;
        padding-bottom: 2.4rem !important;
    }

    p, label, span, div {
        text-wrap: pretty;
    }

    .stMarkdown,
    .stCaption,
    [data-testid="stMarkdownContainer"] p,
    [data-testid="stCaptionContainer"] p {
        color: var(--slate) !important;
    }

    .stMarkdown a {
        color: var(--steel) !important;
    }

    [data-testid="stToolbar"] {
        display: none !important;
    }

    [data-testid="stExpander"] {
        border: 1px solid var(--line);
        border-radius: var(--panel-radius);
        background: linear-gradient(180deg, var(--panel) 0%, var(--card-soft) 100%);
        box-shadow: var(--surface-shadow-soft);
        overflow: hidden;
        margin: 0 0 1.35rem 0;
    }

    [data-testid="stExpander"] details {
        border-radius: var(--panel-radius);
    }

    [data-testid="stExpander"] summary {
        background: linear-gradient(180deg, #ffffff, #f7fbff);
        padding: 0.95rem 1.05rem;
    }

    [data-testid="stExpander"] summary p {
        font-family: "Manrope", "Segoe UI", sans-serif;
        font-size: 0.98rem;
        font-weight: 700;
        color: var(--ink) !important;
    }

    .hero-card,
    .upload-card,
    .quick-card,
    .finding-card,
    .meta-card {
        border: 1px solid var(--line) !important;
        border-radius: var(--card-radius) !important;
        padding: var(--card-padding) !important;
        background: linear-gradient(180deg, var(--card) 0%, var(--card-soft) 100%) !important;
        box-shadow: var(--surface-shadow-soft) !important;
    }

    .filter-panel-shell,
    .compact-header,
    .compact-brand-box,
    .sidebar-user-card,
    .report-meta-card,
    .kpi-card,
    .insight-card,
    .upload-status-card,
    .app-header,
    .empty-state-shell,
    .login-brand-card,
    .login-form-card,
    div[data-testid="stForm"] {
        border: 1px solid var(--line) !important;
        border-radius: var(--card-radius) !important;
        background: linear-gradient(180deg, var(--panel) 0%, var(--card) 100%) !important;
        box-shadow: var(--surface-shadow) !important;
    }

    .hero-card {
        background:
            radial-gradient(circle at top right, rgba(30, 136, 229, 0.08), transparent 34%),
            linear-gradient(180deg, var(--card) 0%, var(--card-soft) 100%) !important;
    }

    .filter-panel-shell {
        position: sticky;
        top: 0.85rem;
        padding: 1rem 1rem 1.05rem 1rem !important;
        backdrop-filter: blur(10px);
    }

    .side-panel-divider {
        border: 0;
        border-top: 1px solid var(--line);
        margin: 0.9rem 0;
    }

    .filter-panel-kicker,
    .upload-step,
    .section-kicker,
    .meta-label,
    .report-meta-label,
    .upload-status-label,
    .sidebar-user-label,
    .login-kicker,
    .aa-panel-intro__kicker,
    .aa-shell__eyebrow,
    .empty-state-kicker,
    .app-header__eyebrow,
    .finding-label,
    .hero-kicker,
    .compact-header-kicker {
        color: var(--steel) !important;
        font-size: 0.75rem;
        text-transform: uppercase;
        letter-spacing: 0.14em;
        font-weight: 800;
    }

    .filter-panel-title,
    .upload-title,
    .quick-title,
    .finding-title,
    .section-title,
    .report-meta-value,
    .hero-title,
    .compact-header-title,
    .login-title,
    .login-form-heading,
    .sidebar-user-name,
    .aa-panel-intro__title,
    .aa-shell__title,
    .empty-state-title,
    .app-header__title,
    .meta-value {
        color: var(--ink) !important;
    }

    .filter-panel-copy,
    .upload-copy,
    .quick-copy,
    .finding-copy,
    .section-copy,
    .sidebar-user-role,
    .login-copy,
    .login-form-copy,
    .aa-panel-intro__copy,
    .aa-shell__copy,
    .empty-state-subtitle,
    .app-header__subtitle,
    .upload-status-meta,
    .upload-status-caption,
    .compact-header-copy,
    .compact-brand-copy,
    .hero-copy,
    .meta-value,
    .app-header-caption {
        color: var(--slate) !important;
    }

    .filter-panel-shell .stRadio > label,
    .filter-panel-shell .stMultiSelect label,
    .filter-panel-shell .stTextInput label,
    .filter-panel-shell .stDateInput label,
    .filter-panel-shell .stCheckbox > label,
    .filter-panel-shell .stSelectbox label {
        color: var(--ink) !important;
        font-weight: 700 !important;
        letter-spacing: 0.01em;
    }

    .stSelectbox label,
    .stMultiSelect label,
    .stTextInput label,
    .stDateInput label,
    .stCheckbox > label,
    .stRadio > label {
        color: var(--ink) !important;
        font-weight: 700 !important;
    }

    .stCheckbox label p,
    .stRadio label p {
        color: var(--slate) !important;
    }

    div[data-baseweb="input"] > div,
    div[data-baseweb="base-input"] > div,
    div[data-baseweb="select"] > div,
    .stDateInput > div > div,
    .stMultiSelect [data-baseweb="tag"],
    .stTextInput > div > div > input,
    .stDateInput input,
    .stNumberInput input,
    .stTextArea textarea {
        background: var(--card) !important;
        color: var(--ink) !important;
        border-color: var(--line-strong) !important;
        border-radius: var(--input-radius) !important;
        transition: all 0.18s ease !important;
        box-shadow: none !important;
    }

    .stTextInput input,
    .stDateInput input,
    .stNumberInput input,
    .stTextArea textarea {
        color: var(--ink) !important;
    }

    div[data-baseweb="input"]:focus-within > div,
    div[data-baseweb="base-input"]:focus-within > div,
    div[data-baseweb="select"]:focus-within > div,
    .stDateInput > div:focus-within > div,
    .stMultiSelect div[data-baseweb="select"]:focus-within > div,
    .stTextInput > div > div:focus-within input,
    .stNumberInput > div > div:focus-within input,
    .stTextArea > div > div:focus-within textarea {
        border-color: var(--steel) !important;
        box-shadow: var(--focus-ring) !important;
        background: #ffffff !important;
    }

    .stButton > button,
    .stDownloadButton > button,
    .stFormSubmitButton > button,
    button[kind="primary"],
    button[kind="secondary"] {
        border-radius: var(--input-radius) !important;
        border: 1px solid var(--line-strong) !important;
        background: linear-gradient(180deg, #ffffff 0%, var(--interactive) 100%) !important;
        color: var(--ink) !important;
        min-height: 2.8rem;
        font-weight: 700 !important;
        box-shadow: var(--surface-shadow-soft) !important;
        transition: all 0.18s ease !important;
    }

    .stButton > button:hover,
    .stDownloadButton > button:hover,
    .stFormSubmitButton > button:hover,
    button[kind="primary"]:hover,
    button[kind="secondary"]:hover {
        border-color: rgba(30, 136, 229, 0.28) !important;
        box-shadow: var(--surface-shadow) !important;
        transform: translateY(-1px);
    }

    .stButton > button:focus,
    .stDownloadButton > button:focus,
    .stFormSubmitButton > button:focus,
    button[kind="primary"]:focus,
    button[kind="secondary"]:focus {
        outline: none !important;
        border-color: var(--steel) !important;
        box-shadow: var(--focus-ring) !important;
    }

    .stDownloadButton button {
        width: 100%;
    }

    div[class*="st-key-sidebar_logout_button"] button,
    div[class*="st-key-workspace_logout_button"] button,
    div[class*="st-key-legacy_sidebar_logout_button"] button {
        border-color: rgba(210, 90, 90, 0.34) !important;
        background: linear-gradient(180deg, #fff9f8 0%, #fff1ef 100%) !important;
        color: var(--rose) !important;
    }

    div[class*="st-key-sidebar_logout_button"] button:hover,
    div[class*="st-key-workspace_logout_button"] button:hover,
    div[class*="st-key-legacy_sidebar_logout_button"] button:hover {
        border-color: rgba(210, 90, 90, 0.52) !important;
        box-shadow: 0 16px 30px rgba(210, 90, 90, 0.12) !important;
    }

    div[class*="st-key-home_tile_"] {
        background: linear-gradient(180deg, var(--card) 0%, var(--card-soft) 100%);
        border: 1px solid var(--line);
        border-radius: var(--card-radius);
        box-shadow: var(--surface-shadow-soft);
        padding: 1.2rem 1.2rem 1rem 1.2rem;
        min-height: 100%;
        transition: all 0.18s ease;
    }

    div[class*="st-key-home_tile_"]:hover {
        transform: translateY(-3px);
        box-shadow: var(--surface-shadow-hover);
        border-color: rgba(30, 136, 229, 0.18);
    }

    div[class*="st-key-home_tile_"] .stButton > button {
        width: 100%;
        margin-top: 0.55rem;
    }

    .aa-hero {
        max-width: 820px;
        margin: 0 auto 2rem auto;
        text-align: center;
        padding: 2rem 1rem 0.35rem 1rem;
    }

    .aa-hero__logo {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        min-height: 148px;
        margin-bottom: 1rem;
    }

    .aa-hero__logo img {
        max-width: min(460px, 76vw);
        width: auto;
        max-height: 138px;
        object-fit: contain;
        filter: drop-shadow(0 18px 32px rgba(17, 38, 64, 0.10));
    }

    .aa-hero__fallback {
        width: 110px;
        height: 110px;
        border-radius: 28px;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        background: linear-gradient(135deg, #1e88e5, #2aa7d8);
        color: #ffffff;
        font-family: "Manrope", "Segoe UI", sans-serif;
        font-size: 2.1rem;
        font-weight: 800;
        letter-spacing: -0.04em;
        box-shadow: var(--surface-shadow);
    }

    .aa-hero__title,
    .hero-title {
        font-family: "Manrope", "Segoe UI", sans-serif;
        font-size: clamp(2.1rem, 4.6vw, 3.45rem);
        font-weight: 800;
        color: var(--ink);
        letter-spacing: -0.04em;
        margin: 0;
    }

    .aa-hero__copy {
        margin: 0.9rem auto 0 auto;
        max-width: 660px;
        color: var(--slate);
        font-size: 1rem;
        line-height: 1.65;
    }

    .hero-logo {
        width: 168px;
        max-width: 100%;
        height: auto;
        display: block;
        margin-bottom: 1rem;
        filter: drop-shadow(0 16px 28px rgba(17, 38, 64, 0.10));
    }

    .hero-stat-grid {
        display: grid;
        grid-template-columns: repeat(2, minmax(0, 1fr));
        gap: 0.8rem;
        margin-top: 1rem;
    }

    .hero-stat {
        border: 1px solid var(--line);
        border-radius: 18px;
        padding: 0.9rem 1rem;
        background: var(--interactive);
    }

    .hero-stat-label {
        font-size: 0.74rem;
        text-transform: uppercase;
        letter-spacing: 0.12em;
        color: var(--muted);
        margin-bottom: 0.35rem;
        font-weight: 700;
    }

    .hero-stat-value {
        color: var(--ink);
        font-size: 1.25rem;
        font-weight: 800;
        line-height: 1.1;
    }

    .aa-shell {
        display: flex;
        justify-content: space-between;
        gap: 1.25rem;
        padding: 1.05rem 1.15rem;
        border-radius: var(--panel-radius);
        border: 1px solid var(--line);
        background: linear-gradient(180deg, var(--panel) 0%, var(--card) 100%);
        box-shadow: var(--surface-shadow-soft);
        margin-bottom: 1.35rem;
        align-items: center;
        flex-wrap: wrap;
    }

    .aa-shell__brand {
        display: flex;
        align-items: center;
        gap: 1rem;
        min-width: 0;
        flex: 1 1 540px;
    }

    .aa-shell__logo {
        width: 68px;
        height: 68px;
        border-radius: 20px;
        background: linear-gradient(180deg, #ffffff, #f4f8fc);
        border: 1px solid var(--line);
        display: flex;
        align-items: center;
        justify-content: center;
        overflow: hidden;
        box-shadow: inset 0 1px 0 rgba(255,255,255,0.7);
    }

    .aa-shell__logo img {
        width: 100%;
        height: 100%;
        object-fit: contain;
    }

    .aa-context-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(168px, 1fr));
        gap: 0.9rem;
        margin: 0 0 1.35rem 0;
    }

    .aa-context-card {
        border-radius: 18px;
        border: 1px solid var(--line);
        background: var(--card);
        padding: 0.95rem 1rem;
        box-shadow: var(--surface-shadow-soft);
    }

    .aa-context-card__label {
        font-size: 0.76rem;
        text-transform: uppercase;
        letter-spacing: 0.14em;
        font-weight: 700;
        color: var(--muted);
        margin-bottom: 0.35rem;
    }

    .aa-context-card__value {
        font-size: 0.98rem;
        font-weight: 700;
        color: var(--ink);
        line-height: 1.45;
        word-break: break-word;
    }

    .aa-tile__icon {
        width: 58px;
        height: 58px;
        border-radius: 18px;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        margin-bottom: 1rem;
        background: linear-gradient(180deg, #eef5ff, #f7fbff);
        border: 1px solid rgba(30, 136, 229, 0.10);
        color: var(--navy);
    }

    .aa-tile__title {
        font-family: "Manrope", "Segoe UI", sans-serif;
        font-size: 1.22rem;
        font-weight: 800;
        color: var(--ink);
        margin-bottom: 0.5rem;
        letter-spacing: -0.02em;
    }

    .aa-tile__copy {
        color: var(--slate);
        line-height: 1.65;
        min-height: 4.9rem;
    }

    .aa-panel-intro {
        border-radius: var(--card-radius);
        border: 1px solid var(--line);
        background: linear-gradient(180deg, var(--panel) 0%, var(--card) 100%);
        padding: 1.2rem 1.25rem;
        box-shadow: var(--surface-shadow-soft);
        margin-bottom: 1rem;
    }

    .compact-header,
    .app-header,
    .empty-state-shell {
        padding: 1.15rem 1.2rem;
        margin-bottom: 1rem;
    }

    .compact-header {
        display: grid;
        grid-template-columns: minmax(0, 1.45fr) auto;
        gap: 1rem;
        align-items: center;
    }

    .compact-pill-row,
    .context-chip-row,
    .pill-row {
        display: flex;
        flex-wrap: wrap;
        gap: 0.55rem;
    }

    .pill-row {
        margin: 0.5rem 0 0.25rem 0;
    }

    .compact-pill,
    .context-chip,
    .pill,
    .brand-badge,
    .app-shell-chip {
        display: inline-flex;
        align-items: center;
        gap: 0.35rem;
        border-radius: 999px;
        padding: 0.42rem 0.78rem;
        border: 1px solid rgba(30, 136, 229, 0.12);
        background: var(--interactive);
        color: var(--ink);
        font-size: 0.79rem;
        font-weight: 700;
        line-height: 1;
    }

    .pill-positive,
    .kpi-card--positive,
    .upload-status-card--ready,
    .insight-card--positive {
        border-color: rgba(31, 143, 100, 0.20) !important;
    }

    .pill-positive {
        color: var(--mint) !important;
        background: rgba(31, 143, 100, 0.10) !important;
    }

    .pill-negative,
    .kpi-card--negative,
    .insight-card--critical {
        border-color: rgba(210, 90, 90, 0.20) !important;
    }

    .pill-negative {
        color: var(--rose) !important;
        background: rgba(210, 90, 90, 0.10) !important;
    }

    .pill-neutral {
        color: var(--steel) !important;
        background: var(--steel-soft) !important;
        border-color: rgba(30, 136, 229, 0.18) !important;
    }

    .brand-wordmark {
        display: inline-block;
        margin-bottom: 1rem;
        color: var(--ink);
        font-family: "Manrope", "Segoe UI", sans-serif;
        font-size: clamp(2rem, 1.35rem + 1.8vw, 3.4rem);
        font-weight: 800;
        line-height: 0.95;
        letter-spacing: -0.05em;
        text-wrap: balance;
    }

    .brand-wordmark--soft {
        color: var(--ink);
    }

    .login-shell {
        max-width: 1080px;
        margin: 0 auto 1rem auto;
        padding: 0.25rem 0 1rem 0;
    }

    .login-grid {
        display: grid;
        grid-template-columns: 1.15fr 0.95fr;
        gap: 1rem;
        align-items: stretch;
    }

    .login-brand-card,
    div[data-testid="stForm"] {
        padding: 1.55rem 1.6rem !important;
        min-height: 460px;
    }

    div[data-testid="stForm"] {
        display: flex;
        flex-direction: column;
        justify-content: center;
    }

    div[data-testid="stForm"] > form {
        display: flex;
        flex-direction: column;
        gap: 0.8rem;
        height: 100%;
    }

    .login-points {
        display: grid;
        gap: 0.8rem;
        margin-top: 1rem;
    }

    .login-point {
        border: 1px solid var(--line);
        border-radius: 18px;
        padding: 0.9rem 1rem;
        background: var(--interactive);
    }

    .login-point-title,
    .upload-status-name,
    .insight-title,
    .kpi-value {
        color: var(--ink);
        font-weight: 800;
    }

    .login-point-copy,
    .kpi-copy,
    .insight-copy {
        color: var(--slate);
    }

    .sidebar-user-card {
        padding: 0.95rem 1rem;
        margin-bottom: 0.8rem;
    }

    .meta-card,
    .report-meta-card,
    .upload-status-card,
    .kpi-card,
    .insight-card {
        padding: 1rem 1.05rem !important;
    }

    .report-metadata-grid {
        display: grid;
        grid-template-columns: repeat(5, minmax(0, 1fr));
        gap: 0.8rem;
        margin-bottom: 1rem;
    }

    .upload-status-grid {
        display: grid;
        gap: 0.75rem;
        margin-top: 0.85rem;
        margin-bottom: 0.85rem;
    }

    .kpi-card {
        min-height: 138px;
    }

    .insight-card {
        min-height: 190px;
    }

    .finding-card {
        min-height: 196px;
    }

    .upload-card {
        min-height: auto;
        padding-bottom: 1rem !important;
        margin-bottom: 0.55rem;
    }

    .quick-card {
        min-height: 156px;
    }

    .file-type-banner {
        width: 100%;
        border-radius: 20px;
        border: 1px solid var(--line);
        display: flex;
        align-items: center;
        justify-content: center;
        text-align: center;
        overflow: hidden;
        position: relative;
        box-shadow: var(--surface-shadow-soft);
        background: linear-gradient(180deg, #f7fbff 0%, #edf4fb 100%);
    }

    .file-type-banner::before {
        content: "";
        position: absolute;
        inset: 0;
        background:
            linear-gradient(90deg, rgba(30, 136, 229, 0.08), transparent 28%, transparent 72%, rgba(30, 136, 229, 0.06)),
            radial-gradient(circle at top left, rgba(30, 136, 229, 0.12), transparent 34%);
        pointer-events: none;
        opacity: 0.8;
    }

    .file-type-banner--sidebar {
        min-height: 108px;
        margin: 0 0 0.95rem 0;
        padding: 0.9rem 1rem;
    }

    .file-type-banner--header {
        min-height: 118px;
        padding: 1rem 1.15rem;
    }

    .file-type-banner--tesla {
        background: linear-gradient(180deg, #f6f8fb 0%, #ebeff4 100%);
    }

    .file-type-banner--mercedes {
        background: linear-gradient(180deg, #f4f7fa 0%, #e9eef4 100%);
    }

    .file-type-banner--audi {
        background: linear-gradient(180deg, #f6f9fc 0%, #eaf1f8 100%);
    }

    .file-type-banner--default {
        background: linear-gradient(180deg, #f7fbff 0%, #edf4fb 100%);
    }

    .file-type-banner__text {
        position: relative;
        z-index: 1;
        color: var(--navy);
        font-family: "Manrope", "Segoe UI", sans-serif;
        font-size: clamp(1.05rem, 0.96rem + 0.55vw, 1.55rem);
        font-weight: 800;
        letter-spacing: 0.16em;
        line-height: 1.1;
        text-transform: uppercase;
        text-wrap: balance;
    }

    section[data-testid="stFileUploader"] {
        border: 1px solid var(--line) !important;
        border-radius: 18px !important;
        background: linear-gradient(180deg, var(--card) 0%, var(--card-soft) 100%) !important;
        box-shadow: var(--surface-shadow-soft) !important;
        padding: 0.35rem 0.5rem 0.55rem 0.5rem !important;
    }

    div[data-testid="stFileUploaderDropzone"] {
        border: 1.5px dashed rgba(23, 32, 51, 0.18) !important;
        border-radius: 16px !important;
        background: var(--interactive) !important;
        padding: 1rem 0.95rem !important;
        transition: all 0.18s ease !important;
    }

    section[data-testid="stFileUploader"]:hover div[data-testid="stFileUploaderDropzone"],
    div[data-testid="stFileUploaderDropzone"]:hover {
        border-color: var(--steel) !important;
        background: #f8fbff !important;
    }

    div[data-testid="stFileUploaderDropzoneInstructions"] span {
        color: var(--slate) !important;
        font-weight: 600;
    }

    div[data-testid="stMetric"] {
        border: 1px solid var(--line) !important;
        border-radius: 20px !important;
        background: linear-gradient(180deg, var(--card) 0%, var(--card-soft) 100%) !important;
        box-shadow: var(--surface-shadow-soft) !important;
        min-height: 132px;
        padding: 0.95rem 1rem !important;
    }

    div[data-testid="stMetric"] label,
    div[data-testid="stMetric"] [data-testid="stMetricLabel"] {
        color: var(--muted) !important;
    }

    div[data-testid="stMetricValue"] {
        color: var(--ink) !important;
    }

    div[data-testid="stMetricDelta"] {
        color: var(--steel) !important;
        background: var(--steel-soft) !important;
        border-radius: 999px !important;
        padding: 0.18rem 0.58rem !important;
        font-weight: 700 !important;
    }

    div[data-testid="stMetricDelta"]:has([data-testid="stMetricDeltaIcon-Up"]) {
        color: var(--mint) !important;
        background: rgba(31, 143, 100, 0.12) !important;
    }

    div[data-testid="stMetricDelta"]:has([data-testid="stMetricDeltaIcon-Down"]) {
        color: var(--rose) !important;
        background: rgba(210, 90, 90, 0.12) !important;
    }

    div[data-testid="stMetricDelta"] *,
    div[data-testid="stMetricDelta"] svg {
        color: inherit !important;
        fill: currentColor !important;
    }

    div[data-testid="stAlert"] {
        border-radius: 18px !important;
        border: 1px solid var(--line) !important;
        background: linear-gradient(180deg, var(--card) 0%, var(--card-soft) 100%) !important;
        box-shadow: var(--surface-shadow-soft) !important;
    }

    div[data-testid="stElementContainer"]:has(> div[data-testid="stVegaLiteChart"]),
    div[data-testid="stElementContainer"]:has(> div[data-testid="stDataFrame"]) {
        border-radius: var(--card-radius);
        overflow: hidden;
        border: 1px solid var(--line);
        background: linear-gradient(180deg, var(--card) 0%, var(--card-soft) 100%);
        box-shadow: var(--surface-shadow-soft);
    }

    div[data-testid="stVegaLiteChart"],
    div[data-testid="stDataFrame"] {
        border-radius: inherit !important;
        overflow: hidden !important;
        border: 0 !important;
        box-shadow: none !important;
        background: transparent !important;
    }

    .vega-embed details,
    .vega-embed .vega-actions,
    .stVegaLiteChart details,
    .stVegaLiteChart summary {
        display: none !important;
    }

    div[data-testid="stFullScreenFrame"]:has(div[data-testid="stVegaLiteChart"])
        [data-testid="stElementToolbar"] {
        display: none !important;
    }

    [data-testid="stButtonGroup"] {
        width: 100%;
        margin-bottom: 0.2rem;
    }

    [data-testid="stButtonGroup"] > div {
        width: 100%;
    }

    [data-testid="stButtonGroup"] [data-baseweb="button-group"] {
        width: 100%;
        padding: 0.22rem;
        border: 1px solid var(--line);
        border-radius: 999px;
        background: var(--interactive);
    }

    [data-testid="stButtonGroup"] button {
        border-radius: 999px !important;
        border: 1px solid transparent !important;
        background: transparent !important;
        color: var(--slate) !important;
        box-shadow: none !important;
        transition: all 0.18s ease !important;
    }

    [data-testid="stButtonGroup"] button:hover {
        color: var(--ink) !important;
        background: rgba(255, 255, 255, 0.66) !important;
    }

    [data-testid="stButtonGroup"] button[kind*="Active"] {
        background: #ffffff !important;
        border-color: rgba(30, 136, 229, 0.18) !important;
        color: var(--ink) !important;
        font-weight: 800 !important;
        box-shadow: var(--surface-shadow-soft) !important;
    }

    [data-testid="stRadio"] [role="radiogroup"] {
        display: flex;
        flex-wrap: wrap;
        gap: 0.35rem;
        padding: 0.24rem;
        border-radius: 999px;
        border: 1px solid var(--line);
        background: var(--interactive);
        width: fit-content;
    }

    [data-testid="stRadio"] [role="radiogroup"] label {
        margin: 0 !important;
        border-radius: 999px;
        padding: 0.35rem 0.7rem;
        border: 1px solid transparent;
        background: transparent;
        transition: all 0.18s ease;
    }

    [data-testid="stRadio"] [role="radiogroup"] label:hover {
        background: rgba(255, 255, 255, 0.72);
        border-color: rgba(30, 136, 229, 0.12);
    }

    div[data-baseweb="tab-list"] {
        gap: 0.45rem !important;
        margin-bottom: 0.8rem;
    }

    button[data-baseweb="tab"] {
        border-radius: 999px !important;
        background: var(--interactive) !important;
        border: 1px solid var(--line) !important;
        color: var(--slate) !important;
        padding: 0.5rem 0.95rem !important;
        box-shadow: none !important;
        transition: all 0.18s ease !important;
    }

    button[data-baseweb="tab"]:hover {
        color: var(--ink) !important;
        border-color: rgba(30, 136, 229, 0.18) !important;
        background: #f8fbff !important;
    }

    button[data-baseweb="tab"][aria-selected="true"] {
        background: #ffffff !important;
        color: var(--ink) !important;
        border-color: rgba(30, 136, 229, 0.22) !important;
        font-weight: 800 !important;
        box-shadow: var(--surface-shadow-soft) !important;
    }

    .section-banner,
    .section-head {
        margin: 0.2rem 0 0.9rem 0;
    }

    .app-header {
        display: grid;
        grid-template-columns: minmax(0, 1.55fr) minmax(220px, 0.75fr);
        gap: 1rem;
        align-items: start;
    }

    .app-header__banner {
        display: grid;
        gap: 0.5rem;
    }

    .empty-state-shell {
        display: grid;
        grid-template-columns: minmax(0, 1.25fr) minmax(220px, 280px);
        gap: 1rem;
        align-items: stretch;
    }

    .empty-state-copy {
        display: flex;
        flex-direction: column;
        gap: 0.72rem;
        justify-content: center;
        min-width: 0;
    }

    .empty-state-banner {
        display: flex;
        align-items: center;
    }

    .empty-state-banner .file-type-banner {
        width: 100%;
        min-height: 120px;
    }

    @media (max-width: 1200px) {
        .report-metadata-grid {
            grid-template-columns: repeat(3, minmax(0, 1fr));
        }

        .app-header {
            grid-template-columns: 1fr;
        }
    }

    @media (max-width: 920px) {
        .hero-title {
            font-size: 2rem;
        }

        .hero-stat-grid,
        .login-grid,
        .compact-header,
        .empty-state-shell,
        .app-header {
            grid-template-columns: 1fr;
        }

        .compact-brand-box {
            justify-items: start;
            text-align: left;
        }

        .login-brand-card,
        .login-form-card,
        div[data-testid="stForm"] {
            min-height: auto !important;
        }
    }

    @media (max-width: 768px) {
        .finding-card {
            min-height: auto !important;
        }

        .report-metadata-grid {
            grid-template-columns: repeat(2, minmax(0, 1fr));
        }
    }

    @media (max-width: 640px) {
        .hero-title {
            font-size: 1.6rem;
        }

        .hero-stat-grid {
            grid-template-columns: 1fr;
        }

        .pill-row {
            gap: 0.35rem;
        }

        .report-metadata-grid {
            grid-template-columns: 1fr;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    :root {
      --font: 'Inter', system-ui, sans-serif;
      --sp-1: 4px; --sp-2: 8px; --sp-3: 12px; --sp-4: 16px; --sp-5: 20px; --sp-6: 24px; --sp-8: 32px; --sp-10: 40px;
      --bg-app: #0d1117;
      --bg-panel: #161b22;
      --bg-card: #1c2230;
      --bg-hover: #212840;
      --bg-input: #131929;
      --border: rgba(255,255,255,0.07);
      --border-strong: rgba(255,255,255,0.13);
      --border-focus: #3b82f6;
      --text-primary: #f0f6fc;
      --text-secondary: #8b949e;
      --text-muted: #484f58;
      --accent-blue: #2d81ff;
      --accent-teal: #00c4b4;
      --accent-green: #3fb950;
      --accent-amber: #d29922;
      --accent-red: #f85149;
      --accent-purple: #8957e5;
      --shadow-sm: 0 1px 3px rgba(0,0,0,0.4);
      --shadow-md: 0 4px 16px rgba(0,0,0,0.5);
      --shadow-lg: 0 8px 32px rgba(0,0,0,0.6);
      --radius-sm: 6px;
      --radius-md: 10px;
      --radius-lg: 16px;
      --radius-xl: 22px;
      --ease: cubic-bezier(0.4, 0, 0.2, 1);
      --duration-fast: 120ms;
      --duration-base: 200ms;
    }

    html, body, [class*="css"] {
      font-family: var(--font) !important;
      -webkit-font-smoothing: antialiased;
    }

    .stApp {
      background:
        radial-gradient(circle at top left, rgba(45, 129, 255, 0.10), transparent 20%),
        radial-gradient(circle at top right, rgba(0, 196, 180, 0.07), transparent 16%),
        linear-gradient(180deg, #0d1117 0%, #0f141d 42%, #0b1017 100%) !important;
      color: var(--text-primary) !important;
    }

    .block-container {
      padding-top: var(--sp-6) !important;
      padding-bottom: var(--sp-8) !important;
      max-width: 1600px !important;
    }

    h1, h2, h3, h4, h5, h6,
    p strong,
    .stMarkdown p strong {
      color: var(--text-primary) !important;
    }

    p, label, span, div, .stCaption {
      color: var(--text-secondary);
    }

    .stMarkdown a {
      color: var(--accent-blue) !important;
    }

    .layout-shell {
      display: grid;
      grid-template-columns: minmax(300px, 340px) minmax(0, 1fr);
      gap: var(--sp-6);
      align-items: start;
    }

    .layout-sidebar {
      position: sticky;
      top: var(--sp-5);
      display: grid;
      gap: var(--sp-4);
      padding: var(--sp-5);
      background: linear-gradient(180deg, rgba(22,27,34,0.96), rgba(19,24,33,0.94));
      border: 1px solid var(--border);
      border-radius: var(--radius-xl);
      box-shadow: var(--shadow-lg);
    }

    .layout-main {
      display: grid;
      gap: var(--sp-5);
      min-width: 0;
    }

    .nav-shell,
    .hero-card,
    .upload-card,
    .quick-card,
    .finding-card,
    .meta-card,
    .sidebar-user-card,
    .report-meta-card,
    .upload-status-card,
    .compact-header,
    .section-card,
    .chart-card,
    .app-header,
    .empty-state-shell,
    .filter-panel-shell,
    .login-brand-card,
    div[data-testid="stForm"],
    div[data-testid="stMetric"] {
      background: linear-gradient(180deg, rgba(28,34,48,0.96), rgba(22,27,34,0.96)) !important;
      border: 1px solid var(--border) !important;
      border-radius: var(--radius-lg) !important;
      box-shadow: var(--shadow-md) !important;
      color: var(--text-primary) !important;
    }

    .section-card,
    .chart-card {
      padding: var(--sp-5);
      animation: fadeSlideUp 360ms var(--ease);
    }

    .hero-card,
    .upload-card,
    .quick-card,
    .finding-card,
    .meta-card,
    .report-meta-card,
    .upload-status-card,
    .chart-card,
    .section-card,
    .sidebar-user-card,
    div[data-testid="stMetric"] {
      transition: transform var(--duration-base) var(--ease), box-shadow var(--duration-base) var(--ease), border-color var(--duration-base) var(--ease);
      animation: fadeSlideUp 360ms var(--ease);
    }

    .hero-card:hover,
    .upload-card:hover,
    .quick-card:hover,
    .finding-card:hover,
    .meta-card:hover,
    .report-meta-card:hover,
    .upload-status-card:hover,
    .chart-card:hover,
    .section-card:hover,
    div[data-testid="stMetric"]:hover {
      transform: translateY(-2px);
      box-shadow: var(--shadow-lg) !important;
      border-color: var(--border-strong) !important;
    }

    .nav-shell {
      display: flex;
      justify-content: space-between;
      gap: var(--sp-4);
      align-items: center;
      padding: var(--sp-5);
    }

    .nav-shell__label,
    .section-kicker,
    .meta-label,
    .report-meta-label,
    .upload-step,
    .sidebar-user-label,
    .hero-kicker,
    .filter-panel-kicker {
      color: var(--text-secondary) !important;
      font-size: 11px !important;
      font-weight: 700 !important;
      text-transform: uppercase;
      letter-spacing: 0.08em;
    }

    .nav-shell__title {
      color: var(--text-primary);
      font-size: clamp(1.4rem, 1.2rem + 0.8vw, 2.1rem);
      font-weight: 800;
      letter-spacing: -0.04em;
      margin: 4px 0 6px 0;
    }

    .nav-shell__copy,
    .section-copy,
    .hero-copy,
    .upload-copy,
    .quick-copy,
    .finding-copy,
    .sidebar-user-role,
    .compact-header-copy,
    .app-header__subtitle,
    .empty-state-subtitle,
    .login-copy,
    .login-form-copy,
    .upload-status-caption,
    .upload-status-meta {
      color: var(--text-secondary) !important;
      line-height: 1.6;
    }

    .section-head {
      margin: 0 0 var(--sp-4) 0;
    }

    .section-title,
    .hero-title,
    .upload-title,
    .quick-title,
    .finding-title,
    .meta-value,
    .sidebar-user-name,
    .report-meta-value,
    .compact-header-title,
    .app-header__title,
    .empty-state-title,
    .login-title,
    .login-form-heading {
      color: var(--text-primary) !important;
    }

    [data-testid="stButtonGroup"],
    [data-testid="stButtonGroup"] > div {
      width: 100%;
    }

    [data-testid="stButtonGroup"] [data-baseweb="button-group"] {
      width: 100%;
      border-radius: 999px;
      background: var(--bg-panel);
      border: 1px solid var(--border);
      padding: 4px;
    }

    [data-testid="stButtonGroup"] button,
    button[data-baseweb="tab"],
    [data-testid="stRadio"] [role="radiogroup"] label {
      border-radius: 999px !important;
      border: 1px solid transparent !important;
      background: transparent !important;
      color: var(--text-secondary) !important;
      transition: all var(--duration-base) var(--ease) !important;
      min-height: 42px;
      font-weight: 600 !important;
      box-shadow: none !important;
    }

    [data-testid="stButtonGroup"] button:hover,
    button[data-baseweb="tab"]:hover,
    [data-testid="stRadio"] [role="radiogroup"] label:hover {
      background: var(--bg-hover) !important;
      color: var(--text-primary) !important;
      transform: translateY(-1px);
    }

    [data-testid="stButtonGroup"] button[kind*="Active"],
    button[data-baseweb="tab"][aria-selected="true"] {
      background: linear-gradient(180deg, rgba(45,129,255,0.24), rgba(45,129,255,0.16)) !important;
      border-color: rgba(59,130,246,0.45) !important;
      color: var(--text-primary) !important;
      font-weight: 800 !important;
      box-shadow: 0 0 0 1px rgba(59,130,246,0.18), 0 8px 24px rgba(0,0,0,0.35) !important;
    }

    [data-testid="stRadio"] [role="radiogroup"] {
      display: inline-flex;
      flex-wrap: wrap;
      gap: 4px;
      padding: 4px;
      border-radius: 999px;
      background: var(--bg-panel);
      border: 1px solid var(--border);
    }

    div[data-baseweb="tab-list"] {
      gap: var(--sp-2) !important;
      margin-bottom: var(--sp-4) !important;
    }

    .stButton > button,
    .stDownloadButton > button,
    .stFormSubmitButton > button,
    button[kind="primary"],
    button[kind="secondary"] {
      border-radius: var(--radius-md) !important;
      border: 1px solid var(--border) !important;
      background: linear-gradient(180deg, rgba(28,34,48,0.96), rgba(19,25,41,0.96)) !important;
      color: var(--text-primary) !important;
      box-shadow: var(--shadow-sm) !important;
      transition: all var(--duration-base) var(--ease) !important;
      min-height: 42px;
      font-weight: 700 !important;
    }

    .stButton > button:hover,
    .stDownloadButton > button:hover,
    .stFormSubmitButton > button:hover,
    button[kind="primary"]:hover,
    button[kind="secondary"]:hover {
      transform: translateY(-2px);
      box-shadow: var(--shadow-md) !important;
      border-color: var(--border-strong) !important;
    }

    .stDownloadButton > button:focus,
    .stButton > button:focus,
    .stFormSubmitButton > button:focus {
      border-color: var(--border-focus) !important;
      box-shadow: 0 0 0 3px rgba(59,130,246,0.18) !important;
    }

    .download-cta .stDownloadButton > button:active {
      animation: pulseGlow 420ms var(--ease);
    }

    div[class*="st-key-sidebar_logout_button"] button,
    div[class*="st-key-workspace_logout_button"] button,
    div[class*="st-key-legacy_sidebar_logout_button"] button {
      border-color: rgba(248,81,73,0.34) !important;
      color: #ffb3ad !important;
      background: linear-gradient(180deg, rgba(83,27,27,0.96), rgba(48,18,18,0.96)) !important;
    }

    div[data-baseweb="input"] > div,
    div[data-baseweb="base-input"] > div,
    div[data-baseweb="select"] > div,
    .stDateInput > div > div,
    .stMultiSelect [data-baseweb="tag"],
    .stTextInput > div > div > input,
    .stDateInput input,
    .stNumberInput input,
    .stTextArea textarea {
      background: var(--bg-input) !important;
      border: 1px solid var(--border) !important;
      color: var(--text-primary) !important;
      border-radius: var(--radius-md) !important;
      transition: all var(--duration-base) var(--ease) !important;
    }

    div[data-baseweb="input"]:focus-within > div,
    div[data-baseweb="base-input"]:focus-within > div,
    div[data-baseweb="select"]:focus-within > div,
    .stDateInput > div:focus-within > div,
    .stMultiSelect div[data-baseweb="select"]:focus-within > div {
      border-color: var(--border-focus) !important;
      box-shadow: 0 0 0 3px rgba(59,130,246,0.16) !important;
    }

    section[data-testid="stSidebar"] {
      display: block !important;
      background: var(--bg-panel);
      border-right: 1px solid var(--border);
    }

    section[data-testid="stSidebar"] > div {
      background: var(--bg-panel);
    }

    section[data-testid="stSidebar"] label {
      color: var(--text-secondary) !important;
      font-size: 11px !important;
      font-weight: 600 !important;
      text-transform: uppercase;
      letter-spacing: 0.08em;
    }

    .filter-panel-shell,
    .upload-card {
      padding: var(--sp-4) !important;
    }

    section[data-testid="stFileUploader"] {
      border: 1px solid var(--border) !important;
      border-radius: var(--radius-lg) !important;
      background: var(--bg-panel) !important;
      box-shadow: none !important;
      padding: var(--sp-2) !important;
    }

    div[data-testid="stFileUploaderDropzone"] {
      border: 1px dashed rgba(255,255,255,0.12) !important;
      border-radius: var(--radius-md) !important;
      background: linear-gradient(180deg, rgba(19,25,41,0.96), rgba(14,19,29,0.96)) !important;
      transition: all var(--duration-base) var(--ease) !important;
    }

    section[data-testid="stFileUploader"]:hover div[data-testid="stFileUploaderDropzone"],
    div[data-testid="stFileUploaderDropzone"]:hover {
      border-color: rgba(45,129,255,0.72) !important;
      box-shadow: 0 0 0 1px rgba(45,129,255,0.32), 0 0 18px rgba(45,129,255,0.18) !important;
    }

    .kpi-grid {
      display: grid;
      grid-template-columns: repeat(5, minmax(0, 1fr));
      gap: var(--sp-4);
    }

    .kpi-card {
      position: relative;
      overflow: hidden;
      padding: var(--sp-5);
      border-radius: var(--radius-lg);
      border: 1px solid var(--border);
      background: linear-gradient(180deg, rgba(28,34,48,0.98), rgba(18,23,34,0.98));
      box-shadow: var(--shadow-md);
      animation: fadeSlideUp 360ms var(--ease);
      transition: transform var(--duration-base) var(--ease), box-shadow var(--duration-base) var(--ease), border-color var(--duration-base) var(--ease);
    }

    .kpi-card:hover {
      transform: translateY(-2px);
      box-shadow: var(--shadow-lg);
      border-color: var(--border-strong);
    }

    .kpi-card::after {
      content: "";
      position: absolute;
      inset: auto 0 0 0;
      height: 3px;
      background: var(--kpi-accent, var(--accent-blue));
      opacity: 0.95;
    }

    .kpi-label {
      color: var(--text-secondary) !important;
      font-size: 11px !important;
      font-weight: 700 !important;
      text-transform: uppercase;
      letter-spacing: 0.08em;
      margin-bottom: var(--sp-3);
    }

    .kpi-value {
      color: var(--text-primary) !important;
      font-size: clamp(1.75rem, 1.45rem + 0.8vw, 2.4rem);
      line-height: 1;
      font-weight: 800;
      margin-bottom: var(--sp-3);
      animation: metricPop 360ms var(--ease);
    }

    .kpi-delta {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: var(--sp-3);
      margin-bottom: var(--sp-3);
      color: var(--text-secondary);
      font-size: 0.84rem;
      font-weight: 600;
    }

    .kpi-delta-value {
      color: var(--kpi-accent, var(--accent-blue));
      font-weight: 700;
    }

    .kpi-progress {
      height: 6px;
      border-radius: 999px;
      background: rgba(255,255,255,0.07);
      overflow: hidden;
      margin-bottom: var(--sp-3);
    }

    .kpi-progress > span {
      display: block;
      height: 100%;
      width: var(--delta-width, 42%);
      background: linear-gradient(90deg, var(--kpi-accent, var(--accent-blue)), rgba(255,255,255,0.28));
    }

    .kpi-copy {
      color: var(--text-secondary) !important;
      font-size: 0.84rem;
      line-height: 1.5;
      margin-bottom: var(--sp-3);
    }

    .kpi-sparkline {
      display: flex;
      align-items: end;
      gap: 4px;
      width: 100%;
      height: 34px;
      margin-top: var(--sp-2);
      opacity: 0.92;
    }

    .kpi-sparkline__bar {
      flex: 1 1 0;
      min-width: 0;
      border-radius: 999px 999px 3px 3px;
      background: linear-gradient(180deg, color-mix(in srgb, var(--kpi-accent, var(--accent-blue)) 92%, #ffffff 8%), color-mix(in srgb, var(--kpi-accent, var(--accent-blue)) 28%, transparent 72%));
      opacity: 0.95;
    }

    div[data-testid="stElementContainer"]:has(> div[data-testid="stPlotlyChart"]),
    div[data-testid="stElementContainer"]:has(> div[data-testid="stDataFrame"]) {
      border-radius: var(--radius-lg);
      overflow: hidden;
    }

    div[data-testid="stPlotlyChart"],
    div[data-testid="stDataFrame"] {
      background: transparent !important;
      border: 0 !important;
      box-shadow: none !important;
    }

    .stDataFrame thead tr th {
      background: #151b25 !important;
      color: var(--text-secondary) !important;
    }

    .stDataFrame tbody tr:hover {
      background: rgba(255,255,255,0.03) !important;
    }

    .app-header,
    .compact-header {
      padding: var(--sp-5);
    }

    .app-header__eyebrow,
    .compact-header-kicker {
      color: var(--accent-blue) !important;
    }

    .app-header__banner .file-type-banner,
    .compact-brand-box .file-type-banner,
    .empty-state-banner .file-type-banner {
      background: linear-gradient(180deg, rgba(33,40,64,0.95), rgba(22,27,34,0.98)) !important;
      border-color: var(--border) !important;
    }

    .file-type-banner__text {
      color: var(--text-primary) !important;
    }

    .login-shell {
      max-width: 1120px;
      margin: 0 auto;
    }

    .login-grid {
      display: grid;
      grid-template-columns: 1.1fr 0.9fr;
      gap: var(--sp-5);
      align-items: stretch;
    }

    .login-brand-card,
    div[data-testid="stForm"] {
      min-height: 460px;
      padding: var(--sp-6) !important;
    }

    div[data-testid="stForm"] > form {
      display: flex;
      flex-direction: column;
      gap: var(--sp-4);
      justify-content: center;
      height: 100%;
    }

    @keyframes fadeSlideUp {
      from { opacity: 0; transform: translateY(10px); }
      to { opacity: 1; transform: translateY(0); }
    }

    @keyframes metricPop {
      from { opacity: 0; transform: translateY(6px); }
      to { opacity: 1; transform: translateY(0); }
    }

    @keyframes pulseGlow {
      0% { box-shadow: 0 0 0 rgba(45,129,255,0); }
      50% { box-shadow: 0 0 0 5px rgba(45,129,255,0.16); }
      100% { box-shadow: 0 0 0 rgba(45,129,255,0); }
    }

    @media (max-width: 1280px) {
      .kpi-grid { grid-template-columns: repeat(3, minmax(0, 1fr)); }
    }

    @media (max-width: 1100px) {
      .layout-shell { grid-template-columns: 1fr; }
      .layout-sidebar { position: static; }
    }

    @media (max-width: 860px) {
      .kpi-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .login-grid { grid-template-columns: 1fr; }
      .login-brand-card, div[data-testid="stForm"] { min-height: auto; }
    }

    @media (max-width: 640px) {
      .kpi-grid { grid-template-columns: 1fr; }
      .nav-shell { flex-direction: column; align-items: stretch; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

def first_non_empty(series):
    values = series.dropna().astype(str).str.strip()
    values = values[values.ne("")]
    return values.iloc[0] if not values.empty else "n/a"


def format_date(value):
    if pd.isna(value):
        return "n/a"
    return pd.Timestamp(value).strftime("%Y-%m-%d")


def format_signed_int(value):
    return f"{float(value):+,.0f}"


def format_signed_pct(value):
    return f"{float(value):+,.1f}%"


def get_date_label(value):
    return DATE_LABELS.get(value, value)


def get_change_label(value):
    return CHANGE_DIRECTION_LABELS.get(value, value)


def get_metric_label(value):
    return MATRIX_METRIC_LABELS.get(value, value)


def get_view_mode_label(value):
    return VIEW_MODE_LABELS.get(value, value)


def format_release_label(meta):
    release_version = str(meta.get("release_version", "")).strip()
    release_date = meta.get("release_date")
    if release_version and release_version.lower() != "n/a":
        return f"v{release_version}"
    if not pd.isna(release_date):
        return f"Snapshot {format_date(release_date)}"
    return "n/a"


def format_release_summary(meta):
    release_label = format_release_label(meta)
    release_date = meta.get("release_date")
    if release_label == "n/a":
        return "n/a"
    if pd.isna(release_date):
        return release_label
    return f"{release_label} / {format_date(release_date)}"


def available_detail_columns(dataframe):
    preferred_columns = [
        "PO Number",
        "Origin Doc",
        "Item",
        "Ship To",
        "Part Number",
        "Part Description",
        "Customer Material",
        "Unrestricted Qty",
        "Unloading Point",
        "Ship Date",
        "Receipt Date",
        "Unit of Measure",
        "CumQty",
        "Quantity_Prev",
        "Quantity_Curr",
        "Delta",
        "Percent Change",
        "Demand Status",
        "Change Direction",
        "Alert",
    ]
    return [column for column in preferred_columns if column in dataframe.columns]


def format_chart_source_table(dataframe):
    source_table = dataframe.copy()
    for column in source_table.columns:
        if pd.api.types.is_datetime64_any_dtype(source_table[column]):
            source_table[column] = source_table[column].dt.strftime("%Y-%m-%d")
        elif pd.api.types.is_bool_dtype(source_table[column]):
            source_table[column] = source_table[column].map(
                lambda value: "Tak" if value else "Nie"
            )
    return source_table


def _is_plotly_figure(chart):
    return isinstance(chart, go.Figure)


def apply_plotly_theme(chart):
    if chart is None or not _is_plotly_figure(chart):
        return chart
    chart.update_layout(**PLOTLY_THEME["layout"])
    return chart


def render_chart_table_switch(
    key,
    chart,
    source_df,
    *,
    chart_empty_message="Brak danych do wykresu.",
    table_empty_message="Brak danych źródłowych.",
    table_height=320,
    chart_export_config=None,
):
    export_state = get_chart_export_state(key, chart_export_config) if chart_export_config else {}
    export_dataset = (
        get_chart_export_dataset(key, source_df, export_state) if chart_export_config else pd.DataFrame()
    )
    export_bytes = (
        build_excel_chart_workbook(key, export_dataset, export_state) if chart_export_config else None
    )

    state_key = f"{key}_view_mode"
    st.session_state.setdefault(state_key, "chart")
    controls_left, controls_right = st.columns([0.68, 0.32], gap="small")
    with controls_left:
        selected_view = st.segmented_control(
            "Widok sekcji",
            options=["chart", "table"],
            selection_mode="single",
            default=st.session_state[state_key],
            required=True,
            key=state_key,
            label_visibility="visible",
            format_func=get_view_mode_label,
            width="stretch",
        )
    with controls_right:
        if chart_export_config:
            export_title = str(export_state.get("title", "chart_export")).strip() or "chart_export"
            st.download_button(
                "Eksportuj wykres do Excel",
                data=export_bytes or b"",
                file_name=f"{slugify_filename(export_title)}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"{key}_excel_chart_export",
                use_container_width=True,
                disabled=not export_bytes,
            )
    selected_view = selected_view or st.session_state[state_key]

    if selected_view == "chart":
        if chart is None:
            st.info(chart_empty_message)
        elif _is_plotly_figure(chart):
            st.plotly_chart(
                apply_plotly_theme(chart),
                use_container_width=True,
                config=PLOTLY_CONFIG,
                key=f"{key}_plotly_chart",
            )
        else:
            st.altair_chart(chart, use_container_width=True)
        return

    source_table = format_chart_source_table(source_df)
    if source_table.empty:
        st.info(table_empty_message)
    else:
        st.dataframe(source_table, use_container_width=True, height=table_height)


def slugify_filename(value):
    normalized = "".join(
        character.lower() if str(character).isalnum() else "_"
        for character in str(value or "").strip()
    )
    normalized = "_".join(part for part in normalized.split("_") if part)
    return normalized or "chart_export"


def normalize_excel_export_table(dataframe):
    if dataframe is None or dataframe.empty:
        return pd.DataFrame()

    export = dataframe.copy()
    for column in export.columns:
        if pd.api.types.is_datetime64_any_dtype(export[column]):
            export[column] = pd.to_datetime(export[column], errors="coerce").dt.strftime("%Y-%m-%d")
        elif pd.api.types.is_bool_dtype(export[column]):
            export[column] = export[column].map(lambda value: "Tak" if value else "Nie")
    return export


def build_active_filter_export_summary():
    filter_state = st.session_state.get("active_filters")
    if not isinstance(filter_state, dict):
        return "Pelny aktywny zakres danych"

    summary_parts = [
        f"Data: {format_workspace_date_range(filter_state)}",
        f"Tygodnie: {format_workspace_week_range(filter_state)}",
        f"Oś daty: {get_date_label(filter_state.get('date_basis', DATE_OPTIONS[0]))}",
    ]
    selected_products = filter_state.get("selected_products", [])
    if selected_products:
        summary_parts.append(f"Produkty: {len(selected_products)}")
    search_term = str(filter_state.get("search_term", "")).strip()
    if search_term:
        summary_parts.append(f"Szukaj: {search_term}")
    if bool(filter_state.get("only_alerts", False)):
        summary_parts.append("Tylko alerty")
    change_directions = filter_state.get("selected_change_directions", [])
    if change_directions and len(change_directions) < len(CHANGE_DIRECTION_LABELS):
        labels = ", ".join(get_change_label(direction) for direction in change_directions)
        summary_parts.append(f"Kierunki: {labels}")
    return " | ".join(summary_parts)


def get_chart_export_state(chart_id, chart_export_config=None):
    export_state = dict(chart_export_config or {})
    export_state.setdefault("chart_id", chart_id)
    export_state.setdefault("title", chart_id.replace("_", " ").title())
    export_state.setdefault("filter_summary", build_active_filter_export_summary())
    filtered_data = st.session_state.get("filtered_data")
    filtered_record_count = (
        len(filtered_data)
        if isinstance(filtered_data, pd.DataFrame)
        else int(export_state.get("filtered_record_count", 0) or 0)
    )
    export_state.setdefault("filtered_record_count", filtered_record_count)
    export_state.setdefault(
        "generated_at",
        pd.Timestamp.now(tz="Europe/Warsaw").strftime("%Y-%m-%d %H:%M:%S %Z"),
    )
    return export_state


def get_chart_export_dataset(chart_id, filtered_data, state):
    dataset = state.get("dataset", filtered_data)
    if callable(dataset):
        dataset = dataset()
    if dataset is None:
        return pd.DataFrame()
    if isinstance(dataset, pd.DataFrame):
        return dataset.copy()
    return pd.DataFrame(dataset)


def build_matrix_chart_export_table(source_df, value_label, label_columns=None):
    if source_df is None or source_df.empty:
        return pd.DataFrame(columns=["Date", value_label])

    label_columns = tuple(label_columns or ("Part Number", "Part Description"))
    value_columns = [column for column in source_df.columns if column not in label_columns]
    totals = [
        {
            "Date": str(column),
            value_label: float(pd.to_numeric(source_df[column], errors="coerce").fillna(0).sum()),
        }
        for column in value_columns
    ]
    return pd.DataFrame(totals)


def prepare_chart_export_plot_table(source_df, state):
    if source_df is None or source_df.empty:
        return pd.DataFrame()

    builder = str(state.get("builder", "identity")).strip().lower()
    if builder == "matrix_totals":
        value_label = (state.get("series_columns") or ["Value"])[0]
        return build_matrix_chart_export_table(
            source_df,
            value_label=value_label,
            label_columns=state.get("label_columns"),
        )

    category_column = state.get("category_column")
    series_columns = list(state.get("series_columns") or [])
    selected_columns = [column for column in [category_column, *series_columns] if column in source_df.columns]
    if not selected_columns:
        return source_df.copy()
    return source_df[selected_columns].copy()


def add_excel_chart_export_data(data_sheet, raw_export_df, chart_export_df):
    if raw_export_df.empty:
        return None

    helper_start_col = raw_export_df.shape[1] + 3
    for column_offset, column_name in enumerate(chart_export_df.columns, start=0):
        data_sheet.cell(row=1, column=helper_start_col + column_offset, value=column_name)
    for row_offset, row in enumerate(chart_export_df.itertuples(index=False), start=2):
        for column_offset, value in enumerate(row, start=0):
            data_sheet.cell(row=row_offset, column=helper_start_col + column_offset, value=value)

    for column_offset in range(chart_export_df.shape[1]):
        data_sheet.column_dimensions[get_column_letter(helper_start_col + column_offset)].hidden = True

    return helper_start_col


def build_excel_native_chart(data_sheet, helper_start_col, chart_export_df, state):
    if chart_export_df is None or chart_export_df.empty or helper_start_col is None:
        return None

    chart_type = str(state.get("chart_type", "line")).strip().lower()
    category_label = chart_export_df.columns[0]
    series_columns = list(chart_export_df.columns[1:])
    if not series_columns:
        return None

    if chart_type == "line":
        chart = LineChart()
    else:
        chart = BarChart()
        chart.type = "bar" if chart_type == "bar" else "col"
        if state.get("stacked"):
            chart.grouping = "stacked"

    chart.style = 10
    chart.title = state.get("title", "Chart export")
    chart.height = float(state.get("height", 8.5))
    chart.width = float(state.get("width", 18))
    chart.x_axis.title = state.get("x_axis_title", category_label)
    chart.y_axis.title = state.get("y_axis_title", series_columns[0])

    data_reference = Reference(
        data_sheet,
        min_col=helper_start_col + 1,
        max_col=helper_start_col + len(series_columns),
        min_row=1,
        max_row=1 + len(chart_export_df),
    )
    category_reference = Reference(
        data_sheet,
        min_col=helper_start_col,
        min_row=2,
        max_row=1 + len(chart_export_df),
    )
    chart.add_data(data_reference, titles_from_data=True)
    chart.set_categories(category_reference)
    chart.legend.position = "r"
    return chart


def build_chart_export_metadata(chart_id, chart_title, chart_dataset, plot_dataset, state):
    return [
        ("Nazwa wykresu", chart_title),
        ("Chart ID", chart_id),
        ("Aktywne filtry", state.get("filter_summary", "n/a")),
        ("Wiersze po filtracji", state.get("filtered_record_count", 0)),
        ("Wiersze datasetu wykresu", len(chart_dataset)),
        ("Wiersze danych wykresu Excel", len(plot_dataset)),
        ("Wygenerowano", state.get("generated_at", "")),
    ]


def write_chart_export_metadata(worksheet, metadata_rows, title):
    worksheet.merge_cells("A1:F1")
    worksheet["A1"] = title
    worksheet["A1"].font = Font(size=16, bold=True, color="0F172A")
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

    worksheet["A3"] = "Metadane eksportu"
    worksheet["A3"].font = Font(size=13, bold=True, color="0F172A")
    for row_offset, (label, value) in enumerate(metadata_rows, start=4):
        worksheet.cell(row=row_offset, column=1, value=label)
        worksheet.cell(row=row_offset, column=2, value=value)
        worksheet.cell(row=row_offset, column=1).font = Font(bold=True, color="334155")
        worksheet.cell(row=row_offset, column=1).fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
        worksheet.cell(row=row_offset, column=1).alignment = Alignment(horizontal="left", vertical="center")
        worksheet.cell(row=row_offset, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 88


def build_excel_chart_workbook(chart_id, filtered_data, state):
    chart_dataset = get_chart_export_dataset(chart_id, filtered_data, state)
    if chart_dataset is None or chart_dataset.empty:
        return None

    chart_export_df = prepare_chart_export_plot_table(chart_dataset, state)
    if chart_export_df.empty or chart_export_df.shape[1] < 2:
        return None

    raw_export_df = normalize_excel_export_table(chart_dataset)
    chart_export_df = normalize_excel_export_table(chart_export_df)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        raw_export_df.to_excel(writer, sheet_name="Data", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Chart", index=False)

        data_sheet = writer.book["Data"]
        style_excel_header(data_sheet, 1)
        style_table_region(data_sheet, 1)
        data_sheet.freeze_panes = "A2"
        data_sheet.auto_filter.ref = data_sheet.dimensions
        autosize_worksheet(data_sheet)
        ensure_numeric_cells_black(data_sheet, start_row=2)

        helper_start_col = add_excel_chart_export_data(data_sheet, raw_export_df, chart_export_df)
        chart_sheet = writer.book["Chart"]
        metadata_rows = build_chart_export_metadata(
            chart_id,
            state.get("title", chart_id),
            raw_export_df,
            chart_export_df,
            state,
        )
        write_chart_export_metadata(chart_sheet, metadata_rows, state.get("title", chart_id))
        excel_chart = build_excel_native_chart(data_sheet, helper_start_col, chart_export_df, state)
        if excel_chart is not None:
            chart_sheet.add_chart(excel_chart, "A12")

    return output.getvalue()


def render_meta_card(title, body_lines):
    body_html = "<br>".join(body_lines)
    st.markdown(
        f"""
        <div class="meta-card">
            <div class="meta-label">{title}</div>
            <div class="meta-value">{body_html}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_status_pills(total_delta, alert_count, products_changed):
    delta_class = "pill-positive" if total_delta > 0 else "pill-negative" if total_delta < 0 else "pill-neutral"
    st.markdown(
        f"""
        <div class="pill-row">
            <div class="pill {delta_class}">Bilans zmian {format_signed_int(total_delta)}</div>
            <div class="pill pill-neutral">Alerty {alert_count:,}</div>
            <div class="pill pill-neutral">Zmienne produkty {products_changed:,}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_finding_card(label, title, copy):
    st.markdown(
        f"""
        <div class="finding-card">
            <div class="finding-label">{html.escape(str(label))}</div>
            <div class="finding-title">{html.escape(str(title))}</div>
            <div class="finding-copy">{html.escape(str(copy))}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_upload_card(step, title, copy):
    st.markdown(
        f"""
        <div class="upload-card">
            <div class="upload-step">{step}</div>
            <div class="upload-title">{title}</div>
            <div class="upload-copy">{copy}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_quick_card(title, copy):
    st.markdown(
        f"""
        <div class="quick-card">
            <div class="quick-title">{title}</div>
            <div class="quick-copy">{copy}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def format_file_type_label(file_type):
    mapping = {
        "legacy_wide": "Tesla / ReleaseData",
        "vl10e_block": "Mercedes / VL10E",
        "cw_weekly_pivot": "Audi Q7/Q9 weekly",
    }
    return mapping.get(str(file_type or "").strip().lower(), "Nieznany format")


def guess_file_type_label(file_name):
    normalized = str(file_name or "").strip().lower()
    if any(keyword in normalized for keyword in ["vl10e", "mercedes", "merc"]):
        return "Mercedes / VL10E"
    if any(keyword in normalized for keyword in ["audi", "q7", "q9", "megatech", "cw17"]):
        return "Audi Q7/Q9 weekly"
    if any(keyword in normalized for keyword in ["tesla", "releasedata", "raw"]):
        return "Tesla / ReleaseData"
    return "Oczekuje na rozpoznanie parsera"


def count_status_matches(dataframe, *keywords):
    if dataframe.empty or "Demand Status" not in dataframe.columns:
        return 0
    normalized = dataframe["Demand Status"].fillna("").astype(str).str.lower()
    return int(
        normalized.apply(lambda value: any(keyword in value for keyword in keywords)).sum()
    )


def render_section_header(kicker, title, copy=None):
    copy_html = (
        f'<div class="section-copy">{html.escape(str(copy))}</div>'
        if copy
        else ""
    )
    markup = (
        '<div class="section-head">'
        f'<div class="section-kicker">{html.escape(str(kicker))}</div>'
        f'<div class="section-title">{html.escape(str(title))}</div>'
        f"{copy_html}"
        "</div>"
    )
    st.markdown(markup, unsafe_allow_html=True)


def section_header(kicker, title, copy=None):
    render_section_header(kicker, title, copy)


def render_app_header(brand_context, title, subtitle, meta_items=None, file_caption=""):
    meta_items = meta_items or []
    chips_html = "".join(
        f'<div class="context-chip">{html.escape(str(item))}</div>'
        for item in meta_items
        if item
    )
    chips_html = (
        f'<div class="context-chip-row">{chips_html}</div>' if chips_html else ""
    )
    banner_html = build_file_type_banner_markup(brand_context, variant="header")
    caption_html = (
        f'<div class="app-header-caption">{html.escape(str(file_caption))}</div>'
        if file_caption
        else ""
    )
    header_markup = (
        '<div class="app-header">'
        '<div class="app-header__copy">'
        '<div class="app-header__eyebrow">Pjoter Development Analytics</div>'
        f'<div class="app-header__title">{html.escape(str(title))}</div>'
        f'<div class="app-header__subtitle">{html.escape(str(subtitle))}</div>'
        f"{chips_html}"
        "</div>"
        '<div class="app-header__banner">'
        f"{banner_html}"
        f"{caption_html}"
        "</div>"
        "</div>"
    )
    st.markdown(header_markup, unsafe_allow_html=True)


def render_empty_state_header(brand_context, title, subtitle, meta_items=None):
    meta_items = meta_items or []
    chips_html = "".join(
        f'<div class="context-chip">{html.escape(str(item))}</div>'
        for item in meta_items
        if item
    )
    chips_html = (
        f'<div class="context-chip-row">{chips_html}</div>' if chips_html else ""
    )
    markup = (
        '<div class="empty-state-shell">'
        '<div class="empty-state-copy">'
        '<div class="empty-state-kicker">Workspace status</div>'
        f'<div class="empty-state-title">{html.escape(str(title))}</div>'
        f'<div class="empty-state-subtitle">{html.escape(str(subtitle))}</div>'
        f"{chips_html}"
        "</div>"
        '<div class="empty-state-banner">'
        f'{build_file_type_banner_markup(brand_context, variant="header")}'
        "</div>"
        "</div>"
    )
    st.markdown(markup, unsafe_allow_html=True)


def render_report_metadata(items):
    cards_html = "".join(
        (
            '<div class="report-meta-card">'
            f'<div class="report-meta-label">{html.escape(str(item.get("label", "")))}</div>'
            f'<div class="report-meta-value">{html.escape(str(item.get("value", "n/a")))}</div>'
            "</div>"
        )
        for item in items
    )
    st.markdown(
        f'<div class="report-metadata-grid">{cards_html}</div>',
        unsafe_allow_html=True,
    )


def _build_delta_width(value, reference=None):
    magnitude = abs(float(value or 0))
    baseline = abs(float(reference or 0))
    if baseline <= 0:
        baseline = magnitude if magnitude > 0 else 1.0
    return max(12.0, min(100.0, (magnitude / baseline) * 100.0))


def _sparkline_svg(values, stroke):
    series = [float(value) for value in values if pd.notna(value)]
    if len(series) < 2:
        return ""
    min_value = min(series)
    max_value = max(series)
    span = max(max_value - min_value, 1.0)
    bars = []
    for value in series:
        height_pct = 18 + (((value - min_value) / span) * 82)
        bars.append(
            f'<span class="kpi-sparkline__bar" style="height:{height_pct:.1f}%"></span>'
        )
    return f'<div class="kpi-sparkline" aria-hidden="true">{"".join(bars)}</div>'


def build_kpi_card_markup(metric):
    accent = html.escape(str(metric.get("accent", "var(--accent-blue)")))
    delta_text = html.escape(str(metric.get("delta", metric.get("copy", ""))))
    delta_label = html.escape(str(metric.get("delta_label", "Delta")))
    delta_width = float(metric.get("delta_width", 42.0))
    sparkline = _sparkline_svg(metric.get("sparkline", []), accent)
    parts = [
        f'<div class="kpi-card" style="--kpi-accent: {accent}; --delta-width: {delta_width:.1f}%;">',
        f'<div class="kpi-label">{html.escape(str(metric.get("label", "")))}</div>',
        f'<div class="kpi-value">{html.escape(str(metric.get("value", "0")))}</div>',
        '<div class="kpi-delta">',
        f"<span>{delta_label}</span>",
        f'<span class="kpi-delta-value">{delta_text}</span>',
        "</div>",
        '<div class="kpi-progress"><span></span></div>',
        f'<div class="kpi-copy">{html.escape(str(metric.get("copy", "")))}</div>',
    ]
    if sparkline:
        parts.append(sparkline)
    parts.append("</div>")
    return "".join(parts)


def build_kpi_grid_markup(metrics):
    cards_html = "".join(build_kpi_card_markup(metric) for metric in metrics)
    markup = f'<div class="kpi-grid">{cards_html}</div>'
    ET.fromstring(f"<root>{markup}</root>")
    return markup


def render_kpi_cards(metrics):
    st.markdown(build_kpi_grid_markup(metrics), unsafe_allow_html=True)


def render_kpi_row(metrics):
    render_kpi_cards(metrics)


def build_dashboard_kpi_metrics(filtered_df, product_summary, date_summary):
    previous_qty = float(filtered_df["Quantity_Prev"].sum())
    current_qty = float(filtered_df["Quantity_Curr"].sum())
    balance_delta = float(filtered_df["Delta"].sum())
    alert_count = int(filtered_df["Alert"].sum()) if "Alert" in filtered_df.columns else 0
    changed_products = int((product_summary["Delta"] != 0).sum()) if not product_summary.empty else 0

    prev_series = date_summary.sort_values("Analysis Date")["Quantity_Prev"].tail(10).tolist()
    curr_series = date_summary.sort_values("Analysis Date")["Quantity_Curr"].tail(10).tolist()
    delta_series = date_summary.sort_values("Analysis Date")["Delta"].tail(10).tolist()
    alert_series = date_summary.sort_values("Analysis Date")["Alerts"].tail(10).tolist() if "Alerts" in date_summary.columns else []
    changed_series = (
        product_summary.assign(Abs_Delta=product_summary["Delta"].abs())
        .sort_values("Abs_Delta", ascending=False)["Abs_Delta"]
        .head(10)
        .tolist()
        if not product_summary.empty
        else []
    )

    delta_pct = (balance_delta / previous_qty * 100.0) if previous_qty else 0.0
    alert_ratio = (alert_count / max(len(filtered_df), 1)) * 100.0
    changed_ratio = (changed_products / max(product_summary["Part Number"].nunique(), 1)) * 100.0 if not product_summary.empty else 0.0

    return [
        {
            "label": "Poprzednia ilość",
            "value": f"{previous_qty:,.0f}",
            "delta_label": "Release",
            "delta": "Baseline",
            "copy": "Suma wolumenu poprzedniego release'u w aktywnym zakresie.",
            "accent": "#8957e5",
            "delta_width": 100,
            "sparkline": prev_series,
        },
        {
            "label": "Aktualna ilość",
            "value": f"{current_qty:,.0f}",
            "delta_label": "Zmiana",
            "delta": f"{delta_pct:+.1f}%",
            "copy": "Aktualny wolumen po zastosowaniu filtrów.",
            "accent": "#2d81ff",
            "delta_width": _build_delta_width(current_qty, previous_qty),
            "sparkline": curr_series,
        },
        {
            "label": "Bilans zmian",
            "value": format_signed_int(balance_delta),
            "delta_label": "Delta %",
            "delta": f"{delta_pct:+.1f}%",
            "copy": "Bilans aktualnego release'u względem poprzedniego.",
            "accent": "#3fb950" if balance_delta >= 0 else "#f85149",
            "delta_width": _build_delta_width(balance_delta, previous_qty),
            "sparkline": delta_series,
        },
        {
            "label": "Alerty",
            "value": f"{alert_count:,}",
            "delta_label": "Udział",
            "delta": f"{alert_ratio:.1f}%",
            "copy": f"Wiersze przekraczające próg {THRESHOLD}% w bieżącym widoku.",
            "accent": "#d29922" if alert_count == 0 else "#f85149",
            "delta_width": _build_delta_width(alert_ratio, 100),
            "sparkline": alert_series,
        },
        {
            "label": "Zmienne produkty",
            "value": f"{changed_products:,}",
            "delta_label": "Coverage",
            "delta": f"{changed_ratio:.1f}%",
            "copy": "Produkty ze zmianą wolumenu w analizowanym oknie.",
            "accent": "#00c4b4",
            "delta_width": _build_delta_width(changed_ratio, 100),
            "sparkline": changed_series,
        },
    ]


def build_kpi_metrics(filtered_df, product_summary):
    increase_total = filtered_df.loc[filtered_df["Delta"] > 0, "Delta"].sum()
    decrease_total = abs(filtered_df.loc[filtered_df["Delta"] < 0, "Delta"].sum())
    changed_rows = int((filtered_df["Delta"] != 0).sum())
    new_positions = count_status_matches(filtered_df, "new")
    removed_positions = count_status_matches(filtered_df, "removed", "delete")
    percent_series = pd.to_numeric(filtered_df["Percent Change"], errors="coerce")
    finite_percent = percent_series[pd.notna(percent_series) & percent_series.ne(float("inf")) & percent_series.ne(float("-inf"))]
    largest_percent = (
        format_signed_pct(finite_percent.loc[finite_percent.abs().idxmax()])
        if not finite_percent.empty
        else "n/a"
    )

    return [
        {
            "label": "Liczba zmian",
            "value": f"{changed_rows:,}",
            "copy": "Wiersze z inną ilością niż w poprzednim release.",
            "tone": "neutral",
        },
        {
            "label": "Łączny wzrost",
            "value": f"{increase_total:,.0f}",
            "copy": "Suma dodatnich zmian w aktualnym widoku.",
            "tone": "positive",
        },
        {
            "label": "Łączny spadek",
            "value": f"{decrease_total:,.0f}",
            "copy": "Suma spadków wymagających weryfikacji.",
            "tone": "negative",
        },
        {
            "label": "Nowe pozycje",
            "value": f"{new_positions:,}",
            "copy": "Wiersze oznaczone jako nowy demand.",
            "tone": "neutral",
        },
        {
            "label": "Usunięte pozycje",
            "value": f"{removed_positions:,}",
            "copy": "Wiersze oznaczone jako removed demand.",
            "tone": "negative",
        },
        {
            "label": "Największa zmiana %",
            "value": largest_percent,
            "copy": f"Największy ruch procentowy w {product_summary['Part Number'].nunique():,} produktach.",
            "tone": "neutral",
        },
    ]


def build_alert_items(filtered_df, key_findings):
    alert_items = []
    alert_count = int(filtered_df["Alert"].sum())
    if alert_count:
        alert_items.append(
            {
                "badge": "Wysoki priorytet",
                "title": f"{alert_count:,} pozycji przekracza próg {THRESHOLD}%",
                "copy": "Te wiersze mają największy potencjał wpływu na plan i warto je sprawdzić w pierwszej kolejności.",
                "tone": "critical",
            }
        )

    new_positions = count_status_matches(filtered_df, "new")
    if new_positions:
        alert_items.append(
            {
                "badge": "Nowy demand",
                "title": f"{new_positions:,} nowych pozycji w aktualnym zakresie",
                "copy": "Pojawiły się nowe linie zapotrzebowania, które nie występowały w poprzednim release.",
                "tone": "positive",
            }
        )

    removed_positions = count_status_matches(filtered_df, "removed", "delete")
    if removed_positions:
        alert_items.append(
            {
                "badge": "Removed demand",
                "title": f"{removed_positions:,} pozycji zostało usuniętych",
                "copy": "Warto potwierdzić, czy zniknięcie tych pozycji jest oczekiwane biznesowo.",
                "tone": "negative",
            }
        )

    for finding in key_findings:
        alert_items.append(
            {
                "badge": finding["label"],
                "title": finding["title"],
                "copy": finding["copy"],
                "tone": "neutral",
            }
        )

    return alert_items[:4]


def render_alerts(alert_items):
    if not alert_items:
        st.markdown(
            """
            <div class="insight-card insight-card--neutral">
                <div class="insight-badge">Stabilny zakres</div>
                <div class="insight-title">Brak istotnych alertów w aktywnym widoku</div>
                <div class="insight-copy">Po zastosowanych filtrach nie ma sygnałów, które przekraczałyby próg ostrzegawczy.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        return

    insight_cols = st.columns(len(alert_items), gap="medium")
    for index, item in enumerate(alert_items):
        tone = html.escape(str(item.get("tone", "neutral")))
        with insight_cols[index]:
            st.markdown(
                f"""
                <div class="insight-card insight-card--{tone}">
                    <div class="insight-badge">{html.escape(str(item.get('badge', 'Insight')))}</div>
                    <div class="insight-title">{html.escape(str(item.get('title', '')))}</div>
                    <div class="insight-copy">{html.escape(str(item.get('copy', '')))}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )


def build_file_slot_payload(slot_label, file_obj=None, meta=None):
    if meta:
        return {
            "slot": slot_label,
            "status": "Załadowany",
            "name": meta.get("file_name", "n/a"),
            "detail": format_file_type_label(meta.get("file_type")),
            "caption": f"Release {format_release_summary(meta)}",
            "tone": "ready",
        }
    if file_obj is not None:
        file_name = (
            file_obj.get("name")
            if isinstance(file_obj, dict)
            else getattr(file_obj, "name", None)
        )
        return {
            "slot": slot_label,
            "status": "Plik dodany",
            "name": file_name or "n/a",
            "detail": guess_file_type_label(file_name or ""),
            "caption": "Plik czeka na wspólne uruchomienie analizy.",
            "tone": "pending",
        }
    return {
        "slot": slot_label,
        "status": "Oczekiwanie",
        "name": "Brak pliku",
        "detail": "Dodaj plik wejściowy",
        "caption": "Sekcja uzupełni się po dodaniu pliku.",
        "tone": "empty",
    }


def render_file_slot_cards(prev_file=None, current_file=None, prev_meta=None, curr_meta=None):
    slots = [
        build_file_slot_payload("Poprzedni plik", prev_file, prev_meta),
        build_file_slot_payload("Aktualny plik", current_file, curr_meta),
    ]
    markup = "".join(
        (
            f'<div class="upload-status-card upload-status-card--{html.escape(str(slot["tone"]))}">'
            f'<div class="upload-status-label">{html.escape(str(slot["slot"]))}</div>'
            f'<div class="upload-status-name">{html.escape(str(slot["name"]))}</div>'
            f'<div class="upload-status-meta">{html.escape(str(slot["status"]))} | {html.escape(str(slot["detail"]))}</div>'
            f'<div class="upload-status-caption">{html.escape(str(slot["caption"]))}</div>'
            "</div>"
        )
        for slot in slots
    )
    st.markdown(
        f'<div class="upload-status-grid">{markup}</div>',
        unsafe_allow_html=True,
    )


def render_upload_section():
    render_section_header(
        "Workspace",
        "Pliki wejściowe",
        "Dodaj poprzedni i aktualny release. Panel po lewej utrzymuje cały kontekst analizy w jednym miejscu.",
    )
    render_upload_card(
        "Poprzedni",
        "Baseline release",
        "Plik referencyjny, do którego porównywany będzie aktualny stan planu i wysyłek.",
    )
    prev_file = st.file_uploader(
        "Upload Previous Release",
        type=["xlsx"],
        key="previous_release_upload",
        label_visibility="visible",
    )
    render_upload_card(
        "Aktualny",
        "Current release",
        "Nowy plik wejściowy, z którego aplikacja policzy zmiany, alerty i bilans wolumenu.",
    )
    current_file = st.file_uploader(
        "Upload Current Release",
        type=["xlsx"],
        key="current_release_upload",
        label_visibility="visible",
    )
    return prev_file, current_file


def render_preload_state(logo_markup=None):
    render_section_header(
        "Start",
        "Dodaj pliki do analizy",
        "Wgraj Previous Release i Current Release. Po dodaniu obu plikow aplikacja uruchomi analize i pokaze dashboard.",
    )
    if logo_markup:
        st.markdown(logo_markup, unsafe_allow_html=True)
    previous_release, current_release = render_workspace_upload_panel()
    render_file_slot_cards(prev_file=previous_release, current_file=current_release)
    if st.session_state.get("error_message"):
        st.error(st.session_state["error_message"])
    elif uploads_ready():
        st.success("Oba pliki sa zapisane. Analiza uruchomi sie automatycznie po odswiezeniu pipeline.")
    elif st.session_state.get("success_status"):
        st.success(st.session_state["success_status"])
    if previous_release is None and current_release is None:
        st.info(
            "Zacznij od dodania dwoch plikow Excel. Po zaladowaniu obu release'ow dashboard uruchomi pelna analize porownawcza."
        )
    elif previous_release is None or current_release is None:
        missing_label = "Previous Release" if previous_release is None else "Current Release"
        st.info(
            f"Brakuje jeszcze pliku: {missing_label}. Po dodaniu drugiego pliku analiza uruchomi sie automatycznie."
        )


def render_export_actions(csv_bytes, excel_bytes, professional_excel_bytes=None):
    render_section_header(
        "Eksport",
        "Pobierz wyniki",
        "Pobierz przefiltrowane dane albo pełny raport Excel bez opuszczania panelu roboczego.",
    )
    st.download_button(
        "Pobierz filtrowane dane CSV",
        data=csv_bytes,
        file_name="pjoter_development_release_change_filtered.csv",
        mime="text/csv",
        use_container_width=True,
    )
    st.download_button(
        "Pobierz raport Excel",
        data=excel_bytes,
        file_name="pjoter_development_release_change_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


def render_extended_export_actions(csv_bytes, excel_bytes, professional_excel_bytes=None):
    render_section_header(
        "Eksport",
        "Pobierz wyniki",
        "Pobierz dane CSV, uproszczony pakiet raportowy Excel albo specjalistyczny raport Weekly by Part.",
    )
    download_left, download_center, download_right = st.columns(3, gap="small")
    with download_left:
        st.download_button(
            "Pobierz filtrowane dane CSV",
            data=csv_bytes,
            file_name="pjoter_development_release_change_filtered.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with download_center:
        st.download_button(
            "Pobierz Reports Pack Excel",
            data=excel_bytes,
            file_name="pjoter_development_reports_pack.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with download_right:
        st.download_button(
            "Pobierz Weekly by Part Excel",
            data=professional_excel_bytes or b"",
            file_name="weekly_by_part_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            disabled=not professional_excel_bytes,
        )


def render_module_navigation(auth_user=None):
    allowed_modules = get_allowed_modules(auth_user=auth_user)
    if not allowed_modules:
        allowed_modules = ["dashboard"]
    current_module = st.session_state.get("active_module")
    if current_module not in allowed_modules:
        st.session_state["active_module"] = allowed_modules[0]

    render_section_header(
        "Nawigacja",
        "Moduły aplikacji",
        "Wybierz aktywny moduł. Upload, filtry i stan sesji pozostają wspólne dla całej aplikacji.",
    )
    selected_module = st.radio(
        "Aktywny moduł",
        options=allowed_modules,
        index=allowed_modules.index(st.session_state.get("active_module", allowed_modules[0])),
        key="active_module",
        label_visibility="visible",
        format_func=lambda value: MODULE_LABELS.get(value, value),
    )
    return selected_module or allowed_modules[0]


def build_planner_scope_source(dataframe, filter_state):
    planner_df = apply_date_week_filters(dataframe, filter_state, date_basis_override="Ship Date")

    selected_products = filter_state.get("selected_products", [])
    if selected_products:
        planner_df = planner_df[planner_df["Product Label"].isin(selected_products)]
    else:
        planner_df = planner_df.iloc[0:0]

    search_term = str(filter_state.get("search_term", "")).strip()
    if search_term:
        query = search_term.lower()
        planner_df = planner_df[
            planner_df["Part Number"].str.lower().str.contains(query, na=False)
            | planner_df["Part Description"].str.lower().str.contains(query, na=False)
        ]

    return prepare_planner_source(planner_df)


def get_planner_storage_key(curr_meta):
    file_name = str(curr_meta.get("file_name", "planner")).strip().lower()
    return f"planner_inputs::{file_name}"


def render_planner_tab(planner_source, curr_meta):
    render_section_header(
        "Planner",
        "Planowanie produkcji względem Ship Date",
        "Part Number i Part Description są pobierane automatycznie z release'u. Operator wpisuje tylko Stock oraz opcjonalny Safety Stock.",
    )

    if planner_source.empty:
        st.info(
            "Brak dodatniego demandu w aktualnym zakresie Ship Date. Poszerz zakres dat albo wybór produktów, aby uruchomić Planner."
        )
        return

    storage_key = get_planner_storage_key(curr_meta)
    stored_inputs = st.session_state.get(storage_key, {})
    planner_input_df = build_planner_input_frame(planner_source, stored_inputs)
    editor_key = f"{storage_key}::editor"

    st.caption(
        "Planner liczy wyłącznie na podstawie Ship Date oraz Quantity_Curr. Filtry zakresu dat, produktów i wyszukiwarka pozostają aktywne."
    )
    edited_inputs = st.data_editor(
        planner_input_df,
        key=editor_key,
        hide_index=True,
        use_container_width=True,
        num_rows="fixed",
        disabled=["Part Number", "Part Description"],
        column_config={
            "Part Number": st.column_config.TextColumn("Part Number", width="medium"),
            "Part Description": st.column_config.TextColumn("Part Description", width="large"),
            "Stock": st.column_config.NumberColumn("Stock", min_value=0.0, step=1.0, format="%.0f"),
            "Safety Stock": st.column_config.NumberColumn("Safety Stock", min_value=0.0, step=1.0, format="%.0f"),
        },
    )
    edited_inputs["Stock"] = pd.to_numeric(edited_inputs["Stock"], errors="coerce").fillna(0.0)
    edited_inputs["Safety Stock"] = pd.to_numeric(edited_inputs["Safety Stock"], errors="coerce").fillna(0.0)
    st.session_state[storage_key] = planner_inputs_to_state(edited_inputs)

    planner_results, planner_daily = calculate_planner_outputs(planner_source, edited_inputs)
    planner_results_table = build_planner_display_table(planner_results)

    planner_kpis = build_planner_kpis(planner_results)
    planner_priority_chart = build_planner_priority_chart(planner_results)
    planner_coverage_chart = build_planner_coverage_chart(planner_results)
    render_kpi_cards(
        [
            {
                "label": "Produkty w plannerze",
                "value": f"{planner_kpis['products']:,}",
                "copy": "Materiały z dodatnim demandem w aktualnym zakresie Ship Date.",
                "tone": "neutral",
            },
            {
                "label": "Pozycje krytyczne",
                "value": f"{planner_kpis['critical']:,}",
                "copy": "Status Krytyczne lub Wysokie ryzyko.",
                "tone": "negative",
            },
            {
                "label": "Qty To Produce Now",
                "value": f"{planner_kpis['to_produce']:,.0f}",
                "copy": "Łączna ilość brakująca do zabezpieczenia popytu i safety stock.",
                "tone": "positive" if planner_kpis["to_produce"] <= 0 else "negative",
            },
            {
                "label": "Średni Coverage %",
                "value": f"{planner_kpis['avg_coverage']:.1f}%",
                "copy": f"Pokryte produkty: {planner_kpis['covered_share']:.1f}%",
                "tone": "neutral",
            },
        ]
    )

    planner_chart_left, planner_chart_right = st.columns(2, gap="large")
    with planner_chart_left:
        render_chart_table_switch(
            "planner_priority",
            apply_chart_theme(planner_priority_chart) if planner_priority_chart is not None else None,
            planner_results_table,
            chart_empty_message="Brak danych do rankingu Planner.",
            table_height=360,
        )
    with planner_chart_right:
        render_chart_table_switch(
            "planner_coverage",
            apply_chart_theme(planner_coverage_chart) if planner_coverage_chart is not None else None,
            planner_results_table,
            chart_empty_message="Brak danych do wykresu coverage.",
            table_height=360,
        )

    st.subheader("Wyniki Planner")
    st.dataframe(planner_results_table, use_container_width=True, height=420)

    selected_planner_part = st.selectbox(
        "Szczegół produktu dzień po dniu",
        options=planner_results["Part Number"].tolist(),
        format_func=lambda value: (
            f"{value} | {planner_results.loc[planner_results['Part Number'] == value, 'Part Description'].iloc[0]}"
        ),
    )
    planner_daily_detail = build_planner_daily_display(planner_daily, selected_planner_part)
    st.dataframe(planner_daily_detail, use_container_width=True, height=360)

    planner_csv_bytes = planner_results_table.to_csv(index=False).encode("utf-8")
    planner_excel_bytes = build_planner_excel_bytes(edited_inputs, planner_results, planner_daily)
    download_left, download_right = st.columns(2)
    with download_left:
        st.download_button(
            "Pobierz Planner CSV",
            data=planner_csv_bytes,
            file_name="planner_summary.csv",
            mime="text/csv",
        )
    with download_right:
        st.download_button(
            "Pobierz Planner Excel",
            data=planner_excel_bytes,
            file_name="planner_summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def build_weekly_focus_table(
    weekly_summary,
    reference_week_label,
    previous_week_label,
    reference_release_delta,
    reference_release_pct,
    reference_wow_delta,
    reference_wow_pct,
):
    reference_row, previous_week_row = get_reference_week_rows(weekly_summary)
    return pd.DataFrame(
        [
            {
                "Widok": "Referencyjny tydzień",
                "Tydzień ISO": reference_week_label,
                "Aktualny release": (
                    f"{float(reference_row['Quantity_Curr']):,.0f}" if reference_row is not None else "0"
                ),
                "Poprzedni release": (
                    f"{float(reference_row['Quantity_Prev']):,.0f}" if reference_row is not None else "0"
                ),
                "Delta release": reference_release_delta,
                "Zmiana release %": reference_release_pct,
                "Delta WoW": reference_wow_delta,
                "Zmiana WoW %": reference_wow_pct,
            },
            {
                "Widok": "Poprzedni tydzień",
                "Tydzień ISO": previous_week_label,
                "Aktualny release": (
                    f"{float(previous_week_row['Quantity_Curr']):,.0f}" if previous_week_row is not None else "0"
                ),
                "Poprzedni release": (
                    f"{float(previous_week_row['Quantity_Prev']):,.0f}" if previous_week_row is not None else "0"
                ),
                "Delta release": (
                    format_signed_int(previous_week_row["Delta"]) if previous_week_row is not None else "+0"
                ),
                "Zmiana release %": (
                    format_percent_display(previous_week_row["Release Percent Label"])
                    if previous_week_row is not None
                    else "n/a"
                ),
                "Delta WoW": (
                    format_signed_int(previous_week_row["WoW Delta"]) if previous_week_row is not None else "+0"
                ),
                "Zmiana WoW %": (
                    format_percent_display(previous_week_row["WoW Percent Label"])
                    if previous_week_row is not None
                    else "n/a"
                ),
            },
        ]
    )


def build_product_detail_table(product_detail):
    product_table = product_detail[available_detail_columns(product_detail)].copy()
    product_table["Ship Date"] = product_table["Ship Date"].dt.strftime("%Y-%m-%d")
    product_table["Receipt Date"] = product_table["Receipt Date"].dt.strftime("%Y-%m-%d")
    product_table["Change Direction"] = product_table["Change Direction"].map(
        get_change_label
    )
    product_table["Alert"] = product_table["Alert"].map(
        lambda value: "Tak" if value else "Nie"
    )
    return product_table.rename(
        columns={
            "Part Number": "Numer czesci",
            "Part Description": "Opis produktu",
            "Origin Doc": "Origin Doc",
            "Item": "Pozycja",
            "Ship To": "Ship-to",
            "Customer Material": "Material klienta",
            "Unrestricted Qty": "Ilosc unrestr.",
            "Unloading Point": "Punkt rozladunku",
            "Ship Date": "Data wysylki",
            "Receipt Date": "Data odbioru",
            "Unit of Measure": "JM",
            "CumQty": "CumQty",
            "Quantity_Prev": "Poprzednia ilosc",
            "Quantity_Curr": "Aktualna ilosc",
            "Delta": "Zmiana ilosci",
            "Percent Change": "Zmiana %",
            "Demand Status": "Status popytu",
            "Change Direction": "Kierunek zmiany",
            "Alert": "Alert",
        }
    )


def render_welcome_state(prev_file, current_file):
    brand_context = detect_brand_context(
        *(meta for meta in [
            {"file_name": prev_file.name} if prev_file is not None else None,
            {"file_name": current_file.name} if current_file is not None else None,
        ] if meta)
    )
    has_any_file = prev_file is not None or current_file is not None
    title = (
        "Dodaj dwa pliki, aby uruchomić porównanie release'ów"
        if not has_any_file
        else "Dodaj drugi plik, aby dokończyć analizę"
    )
    subtitle = (
        "Lewa kolumna służy do uploadu, filtrów i eksportu. Po dodaniu kompletu plików "
        "dashboard automatycznie pokaże KPI, alerty, wykresy i tabele szczegółowe."
    )
    meta_items = [
        "Upload po lewej stronie",
        "Porównanie daily i weekly",
        "Eksport CSV / Excel",
        "1 / 2 plików gotowe" if has_any_file else "0 / 2 plików gotowe",
    ]
    render_empty_state_header(brand_context, title, subtitle, meta_items)

    render_section_header(
        "Po uruchomieniu analizy",
        "Co zobaczysz w raporcie",
        "Po lewej stronie zostaje sterowanie analizą, a główna sekcja skupia się na wynikach i najważniejszych sygnałach.",
    )
    quick_cols = st.columns(3, gap="medium")
    with quick_cols[0]:
        render_quick_card(
            "Szybkie KPI",
            "Najważniejsze liczby i sygnały będą zawsze na górze raportu, gotowe do szybkiego odczytu.",
        )
    with quick_cols[1]:
        render_quick_card(
            "Alerty i insighty",
            "Sekcja alertów porządkuje anomalie, nowe pozycje i zmiany przekraczające ustalony próg.",
        )
    with quick_cols[2]:
        render_quick_card(
            "Stały kontekst pracy",
            "Filtry, upload i eksport pozostają w jednym miejscu, dzięki czemu dashboard nie gubi kontekstu pracy.",
        )

    if prev_file is None and current_file is None:
        st.info("Dodaj dwa pliki Excel w panelu po lewej, aby uruchomić porównanie release'ów.")
    else:
        missing_label = "poprzedni" if prev_file is None else "aktualny"
        st.info(
            f"Jeden plik jest już gotowy. Dodaj jeszcze plik {missing_label}, aby uruchomić pełną analizę."
        )
    return


def build_detail_export_table(dataframe):
    detail_table = dataframe[available_detail_columns(dataframe)].copy()
    if "Ship Date" in detail_table.columns:
        detail_table["Ship Date"] = detail_table["Ship Date"].dt.strftime("%Y-%m-%d")
    if "Receipt Date" in detail_table.columns:
        detail_table["Receipt Date"] = detail_table["Receipt Date"].dt.strftime("%Y-%m-%d")
    if "Change Direction" in detail_table.columns:
        detail_table["Change Direction"] = detail_table["Change Direction"].map(
            get_change_label
        )
    if "Alert" in detail_table.columns:
        detail_table["Alert"] = detail_table["Alert"].map(
            lambda value: "Tak" if value else "Nie"
        )
    return detail_table.rename(
        columns={
            "PO Number": "Numer PO",
            "Origin Doc": "Origin Doc",
            "Item": "Pozycja",
            "Ship To": "Ship-to",
            "Part Number": "Numer części",
            "Part Description": "Opis produktu",
            "Customer Material": "Materiał klienta",
            "Unrestricted Qty": "Ilość unrestr.",
            "Unloading Point": "Punkt rozładunku",
            "Ship Date": "Data wysyłki",
            "Receipt Date": "Data odbioru",
            "Unit of Measure": "JM",
            "CumQty": "CumQty",
            "Quantity_Prev": "Poprzednia ilość",
            "Quantity_Curr": "Aktualna ilość",
            "Delta": "Zmiana ilości",
            "Percent Change": "Zmiana %",
            "Demand Status": "Status popytu",
            "Change Direction": "Kierunek zmiany",
            "Alert": "Alert",
        }
    )


def load_auth_config():
    if not AUTH_USERS_PATH.exists():
        # Default user for Streamlit Cloud deployment
        return [
            {
                "username": "admin",
                "display_name": "Administrator",
                "role": "Admin",
                "active": True,
                "salt": "c6b02c39a66d2460b6a3a3885b467ad0",
                "password_hash": "f951c24eead1d41496fc80c791f5ac802af477002998494b058dde362f1e2dda"
            }
        ]
    try:
        with AUTH_USERS_PATH.open("r", encoding="utf-8") as file:
            payload = json.load(file)
        return payload.get("users", [])
    except Exception:
        # Return default user if JSON is corrupted
        return [
            {
                "username": "admin",
                "display_name": "Administrator",
                "role": "Admin",
                "active": True,
                "salt": "c6b02c39a66d2460b6a3a3885b467ad0",
                "password_hash": "f951c24eead1d41496fc80c791f5ac802af477002998494b058dde362f1e2dda"
            }
        ]


def verify_password(password, salt_hex, password_hash_hex):
    salt = binascii.unhexlify(salt_hex.encode("utf-8"))
    computed_hash = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), salt, 120000
    )
    return binascii.hexlify(computed_hash).decode("utf-8") == password_hash_hex


def init_auth_state():
    st.session_state.setdefault("authenticated", False)
    st.session_state.setdefault("auth_user", None)


def attempt_login(username, password):
    users = load_auth_config()
    normalized_username = username.strip().lower()
    for user in users:
        if not user.get("active", True):
            continue
        if str(user.get("username", "")).strip().lower() != normalized_username:
            continue
        if verify_password(
            password,
            user.get("salt", ""),
            user.get("password_hash", ""),
        ):
            st.session_state["authenticated"] = True
            st.session_state["auth_user"] = {
                "username": user.get("username", ""),
                "display_name": user.get("display_name", user.get("username", "")),
                "role": user.get("role", "User"),
            }
            return True
    return False


def logout_user():
    st.session_state["authenticated"] = False
    st.session_state["auth_user"] = None


def get_auth_user():
    return st.session_state.get("auth_user") or {}


def get_user_role(auth_user=None):
    auth_user = auth_user or get_auth_user()
    role = str(auth_user.get("role", "Viewer")).strip()
    return role if role in ROLE_MODULE_PERMISSIONS else "Viewer"


def get_role_module_permissions(role=None, auth_user=None):
    resolved_role = role or get_user_role(auth_user=auth_user)
    return ROLE_MODULE_PERMISSIONS.get(resolved_role, ROLE_MODULE_PERMISSIONS["Viewer"])


def get_allowed_modules(auth_user=None):
    permissions = get_role_module_permissions(auth_user=auth_user)
    return [module for module in MODULE_OPTIONS if module in permissions]


def get_module_access_level(module_name, auth_user=None):
    permissions = get_role_module_permissions(auth_user=auth_user)
    return permissions.get(module_name, "none")


def can_access_module(module_name, auth_user=None):
    return get_module_access_level(module_name, auth_user=auth_user) != "none"


def _legacy_render_sidebar_user():
    auth_user = st.session_state.get("auth_user") or {}
    st.sidebar.markdown(
        f"""
        <div class="sidebar-user-card">
            <div class="sidebar-user-label">Aktywna sesja</div>
            <div class="sidebar-user-name">{auth_user.get('display_name', 'User')}</div>
            <div class="sidebar-user-role">{auth_user.get('role', 'User')} · {auth_user.get('username', '')}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if st.sidebar.button("Wyloguj", key="legacy_sidebar_logout_button", use_container_width=True):
        logout_user()
        st.rerun()


def render_filter_panel_shell():
    st.markdown(
        """
        <div class="filter-panel-shell">
            <div class="filter-panel-kicker">Panel Nawigacji</div>
            <div class="filter-panel-title">Filtry i kontekst analizy</div>
            <div class="filter-panel-copy">
                Ten panel pozostaje stale widoczny, aby filtrowanie, kalendarz i kontrola zakresu
                były zawsze pod ręką podczas pracy z dashboardem.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_login_screen():
    st.markdown('<div class="login-shell">', unsafe_allow_html=True)
    left_col, right_col = st.columns([1.18, 0.92], gap="large")
    with left_col:
        logo_html = (
            f'<img class="hero-logo" src="{logo_data_uri()}" alt="{APP_TITLE} logo" />'
            if logo_available()
            else '<div class="brand-wordmark brand-wordmark--soft">Aplikacja Analityczna</div>'
        )
        st.markdown(
            f"""
            <div class="login-brand-card">
                {logo_html}
                <div class="login-kicker">Secure Access</div>
                <div class="login-title">Analiza zamówień i wysyłek</div>
                <div class="login-copy">
                    Zaloguj się, aby otworzyć dashboard, porównywać dwa pliki Excel i generować raporty
                    dla planowania, logistyki oraz przeglądów managerskich.
                </div>
                <div class="login-points">
                    <div class="login-point">
                        <div class="login-point-title">Premium dashboard analytics</div>
                        <div class="login-point-copy">Czytelne KPI, alerty, wykresy i macierz zmian z wyraźną hierarchią informacji.</div>
                    </div>
                    <div class="login-point">
                        <div class="login-point-title">Raport gotowy do eksportu</div>
                        <div class="login-point-copy">Filtrowane dane CSV i biznesowy raport Excel przygotowany do dalszej pracy.</div>
                    </div>
                    <div class="login-point">
                        <div class="login-point-title">Dostęp na wielu komputerach</div>
                        <div class="login-point-copy">Ta sama aplikacja może działać lokalnie, w sieci LAN lub jako uruchamiany launcher EXE.</div>
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with right_col:
        with st.form("login_form", clear_on_submit=False):
            st.markdown(
                """
                <div class="login-form-heading">Zaloguj się do aplikacji</div>
                <div class="login-form-copy">
                    Użyj swojego loginu i hasła, aby otworzyć panel analityczny. Dane dostępowe są trzymane
                    lokalnie w konfiguracji aplikacji.
                </div>
                """,
                unsafe_allow_html=True,
            )
            username = st.text_input("Login")
            password = st.text_input("Hasło", type="password")
            submitted = st.form_submit_button("Zaloguj", use_container_width=True)
        if submitted:
            if attempt_login(username, password):
                st.success("Logowanie zakończone powodzeniem.")
                st.rerun()
            st.error("Nieprawidłowy login lub hasło.")
        st.info(
            "Domyślny login jest zapisany w pliku config/users.json. Po wdrożeniu zmień hasło administratora."
        )
    st.markdown("</div>", unsafe_allow_html=True)


def logo_available():
    return LOGO_PATH.exists()


def asset_data_uri(path):
    try:
        if not path.exists():
            return ""
        extension = path.suffix.lower()
        if extension == ".svg":
            mime_type = "image/svg+xml"
        elif extension in {".jpg", ".jpeg"}:
            mime_type = "image/jpeg"
        elif extension == ".webp":
            mime_type = "image/webp"
        else:
            mime_type = "image/png"
        encoded = base64.b64encode(path.read_bytes()).decode("utf-8")
        return f"data:{mime_type};base64,{encoded}"
    except Exception:
        return ""


def logo_data_uri():
    return asset_data_uri(LOGO_PATH)


def init_ui_state():
    st.session_state.setdefault("active_view", "dashboard")
    st.session_state.setdefault("filters_expanded", False)
    st.session_state.setdefault("file_view", "overview")
    st.session_state.setdefault("active_filters", {})
    st.session_state.setdefault("date_from", None)
    st.session_state.setdefault("date_to", None)
    st.session_state.setdefault("week_from", None)
    st.session_state.setdefault("week_to", None)
    st.session_state.setdefault("filtered_data", pd.DataFrame())
    st.session_state.setdefault("reset_filters_trigger", False)
    for nonce_key in UPLOAD_NONCE_KEYS.values():
        st.session_state.setdefault(nonce_key, 0)
    st.session_state.setdefault("uploaded_files", {})
    st.session_state.setdefault("file_type", {})
    st.session_state.setdefault("parsed_data", {})
    st.session_state.setdefault("transformed_data", {})
    st.session_state.setdefault("analysis_results", None)
    st.session_state.setdefault("analysis_signature", None)
    st.session_state.setdefault("error_message", "")
    st.session_state.setdefault("success_status", "")
    st.session_state.setdefault("debug_status", "")
    previous_upload = get_stored_upload("previous")
    current_upload = get_stored_upload("current")
    st.session_state["prev_release_bytes"] = previous_upload["bytes"] if previous_upload else None
    st.session_state["curr_release_bytes"] = current_upload["bytes"] if current_upload else None
    st.session_state["uploaded_files"] = {
        slot_name: {
            "name": upload_payload.get("name"),
            "size": upload_payload.get("size"),
            "sha1": upload_payload.get("sha1"),
        }
        for slot_name, upload_payload in (
            ("previous", previous_upload),
            ("current", current_upload),
        )
        if upload_payload is not None
    }


def set_active_view(view_name, *, close_filters=True):
    if view_name in PRIMARY_VIEW_KEYS:
        st.session_state["active_view"] = view_name
    if close_filters:
        st.session_state["filters_expanded"] = False


def open_filters_panel():
    st.session_state["filters_expanded"] = True


def close_filters_panel():
    st.session_state["filters_expanded"] = False


def get_upload_widget_key(slot_name):
    return f"{slot_name}_release_upload_{st.session_state.get(UPLOAD_NONCE_KEYS[slot_name], 0)}"


def get_stored_upload(slot_name):
    return st.session_state.get(UPLOAD_STATE_KEYS[slot_name])


def build_uploaded_file_payload(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    return {
        "name": uploaded_file.name,
        "bytes": file_bytes,
        "size": len(file_bytes),
        "sha1": hashlib.sha1(file_bytes).hexdigest(),
    }


def set_pipeline_debug(message):
    st.session_state["debug_status"] = str(message)
    print(f"[pipeline] {message}")


def clear_analysis_state():
    st.session_state["file_type"] = {}
    st.session_state["parsed_data"] = {}
    st.session_state["transformed_data"] = {}
    st.session_state["analysis_results"] = None
    st.session_state["analysis_signature"] = None
    st.session_state["error_message"] = ""
    st.session_state["success_status"] = ""


def upload_signature(upload_payload):
    if not isinstance(upload_payload, dict):
        return ""
    return f"{upload_payload.get('name', '')}:{upload_payload.get('size', 0)}:{upload_payload.get('sha1', '')}"


def current_upload_pair_signature():
    return "|".join(
        [
            upload_signature(get_stored_upload("previous")),
            upload_signature(get_stored_upload("current")),
        ]
    )


def store_uploaded_release(slot_name, uploaded_file):
    if uploaded_file is None:
        return get_stored_upload(slot_name)
    payload = build_uploaded_file_payload(uploaded_file)
    existing_payload = get_stored_upload(slot_name)
    st.session_state[UPLOAD_STATE_KEYS[slot_name]] = payload
    if slot_name == "previous":
        st.session_state["prev_release_bytes"] = payload["bytes"]
    else:
        st.session_state["curr_release_bytes"] = payload["bytes"]
    uploaded_files = dict(st.session_state.get("uploaded_files", {}))
    uploaded_files[slot_name] = {
        "name": payload["name"],
        "size": payload["size"],
        "sha1": payload["sha1"],
    }
    st.session_state["uploaded_files"] = uploaded_files
    if upload_signature(existing_payload) != upload_signature(payload):
        clear_analysis_state()
        st.session_state["success_status"] = f"Plik {payload['name']} zostal zapisany."
        set_pipeline_debug(f"stored {slot_name} upload: {payload['name']} ({payload['size']} bytes)")
    return payload


def clear_uploaded_release(slot_name):
    st.session_state.pop(UPLOAD_STATE_KEYS[slot_name], None)
    st.session_state[UPLOAD_NONCE_KEYS[slot_name]] = st.session_state.get(UPLOAD_NONCE_KEYS[slot_name], 0) + 1
    if slot_name == "previous":
        st.session_state["prev_release_bytes"] = None
    else:
        st.session_state["curr_release_bytes"] = None
    uploaded_files = dict(st.session_state.get("uploaded_files", {}))
    uploaded_files.pop(slot_name, None)
    st.session_state["uploaded_files"] = uploaded_files
    clear_analysis_state()
    set_pipeline_debug(f"cleared {slot_name} upload")


def clear_workspace_uploads():
    clear_uploaded_release("previous")
    clear_uploaded_release("current")
    close_filters_panel()
    st.session_state["file_view"] = "overview"


def workspace_has_uploads():
    return get_stored_upload("previous") is not None or get_stored_upload("current") is not None


def uploads_ready():
    return bool(
        st.session_state.get("prev_release_bytes")
        and st.session_state.get("curr_release_bytes")
    )


def workspace_is_ready():
    return uploads_ready()


def sync_uploaded_files_from_widgets():
    changed = False
    for slot_name in ("previous", "current"):
        widget_value = st.session_state.get(get_upload_widget_key(slot_name))
        if widget_value is None:
            continue
        existing_payload = get_stored_upload(slot_name)
        widget_payload = build_uploaded_file_payload(widget_value)
        if upload_signature(existing_payload) == upload_signature(widget_payload):
            continue
        store_uploaded_release(slot_name, widget_value)
        changed = True
    if changed:
        set_pipeline_debug(
            "synchronized uploader widgets before analysis gate"
        )
    return changed


def trigger_analysis_refresh():
    st.session_state["analysis_signature"] = None
    st.session_state["analysis_results"] = None
    st.session_state["error_message"] = ""
    st.session_state["success_status"] = ""
    set_pipeline_debug("manual analysis refresh requested")


def render_sidebar_user(target=st.sidebar):
    auth_user = st.session_state.get("auth_user") or {}
    target.markdown(
        f"""
        <div class="sidebar-user-card">
            <div class="sidebar-user-label">Aktywna sesja</div>
            <div class="sidebar-user-name">{html.escape(str(auth_user.get('display_name', 'User')))}</div>
            <div class="sidebar-user-role">{html.escape(str(auth_user.get('role', 'User')))} &middot; {html.escape(str(auth_user.get('username', '')))}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if target.button("Wyloguj", key="sidebar_logout_button", use_container_width=True):
        logout_user()
        st.rerun()


def detect_brand_context(*metas):
    file_names = [
        str(meta.get("file_name", "")).lower()
        for meta in metas
        if isinstance(meta, dict)
    ]
    file_types = {
        str(meta.get("file_type", "")).strip().lower()
        for meta in metas
        if isinstance(meta, dict)
    }
    joined_names = " ".join(file_names)

    if "cw_weekly_pivot" in file_types or any(
        keyword in joined_names for keyword in ["audi", "q7", "q9", "megatech"]
    ):
        return {
            "brand_key": "audi",
            "label": "Audi Q7/Q9",
            "status": "Klient: Audi Q7/Q9",
            "format_copy": "Rozpoznano tygodniowy format CW / Audi Q7/Q9.",
            "banner_text": "AUDI Q7/Q9",
            "banner_theme": "audi",
        }

    if any(keyword in joined_names for keyword in ["mercedes", "merc", "vl10e"]) or "vl10e_block" in file_types:
        return {
            "brand_key": "mercedes",
            "label": "Mercedes-Benz",
            "status": "Klient: Mercedes-Benz",
            "format_copy": "Rozpoznano plik VL10E / Mercedes.",
            "banner_text": "MERCEDES-BENZ",
            "banner_theme": "mercedes",
        }

    if any(keyword in joined_names for keyword in ["releasedata", "tesla"]) or (
        "legacy_wide" in file_types and "releasedata" in joined_names
    ):
        return {
            "brand_key": "tesla",
            "label": "Tesla",
            "status": "Klient: Tesla / ReleaseData",
            "format_copy": "Rozpoznano plik ReleaseData / legacy wide.",
            "banner_text": "TESLA",
            "banner_theme": "tesla",
        }

    return {
        "brand_key": "default",
        "label": "Analytics Dashboard",
        "status": "Klient: neutralny",
        "format_copy": "Brak dedykowanej marki dla załadowanego pliku.",
        "banner_text": "ANALYTICS DASHBOARD",
        "banner_theme": "default",
    }


def describe_format_context(*metas):
    file_types = [
        str(meta.get("file_type", "")).strip().lower()
        for meta in metas
        if isinstance(meta, dict) and meta.get("file_type")
    ]
    if not file_types:
        return "Oczekiwanie na dwa pliki wejściowe."
    unique_types = sorted(set(file_types))
    if unique_types == ["cw_weekly_pivot"]:
        return "Format: Weekly pivot"
    if unique_types == ["vl10e_block"]:
        return "Format: VL10E block"
    if unique_types == ["legacy_wide"]:
        return "Format: Legacy wide"
    if set(unique_types) == {"cw_weekly_pivot", "legacy_wide"}:
        return "Format: daily + weekly"
    return "Format: mixed / mixed release sources"


def build_file_type_banner_markup(brand_context, variant="sidebar"):
    banner_text = html.escape(str(brand_context.get("banner_text", "ANALYTICS DASHBOARD")))
    banner_theme = html.escape(str(brand_context.get("banner_theme", "default")))
    banner_variant = "header" if variant == "header" else "sidebar"
    return (
        f'<div class="file-type-banner file-type-banner--{banner_variant} '
        f'file-type-banner--{banner_theme}">'
        f'<div class="file-type-banner__text">{banner_text}</div>'
        "</div>"
    )


def render_file_type_banner(brand_context, target=st, variant="sidebar"):
    target.markdown(
        build_file_type_banner_markup(brand_context, variant=variant),
        unsafe_allow_html=True,
    )


def render_side_panel_brand(brand_context):
    render_file_type_banner(brand_context, variant="sidebar")


def render_compact_header(brand_context, prev_meta, curr_meta, date_basis, selected_start_date, selected_end_date):
    format_context = describe_format_context(prev_meta, curr_meta)
    render_app_header(
        brand_context,
        f"Raport zmian dla PO {curr_meta.get('po_number', 'n/a')}",
        (
            f"{brand_context.get('status', 'Klient: neutralny')}. "
            f"{brand_context.get('format_copy', '')} "
            f"Zakres analizy: {selected_start_date:%Y-%m-%d} do {selected_end_date:%Y-%m-%d} "
            f"na osi {get_date_label(date_basis)}."
        ),
        [
            format_context,
            f"Poprzedni: {format_release_label(prev_meta)}",
            f"Aktualny: {format_release_label(curr_meta)}",
        ],
        curr_meta.get("file_name", ""),
    )


def apply_chart_theme(chart):
    return (
        chart.configure_view(strokeOpacity=0)
        .configure(background="transparent")
        .configure_axis(
            grid=False,
            domainColor="#c8d4e3",
            tickColor="#94a3b8",
            labelColor="#5b667a",
            titleColor="#172033",
            labelFontSize=12,
            titleFontSize=13,
            tickSize=6,
            labelPadding=10,
            titlePadding=12,
        )
        .configure_axisX(
            grid=True,
            gridColor="rgba(148, 163, 184, 0.18)",
            gridDash=[2, 6],
            domain=False,
            tickColor="#94a3b8",
            labelColor="#5b667a",
        )
        .configure_axisY(
            grid=True,
            gridColor="rgba(148, 163, 184, 0.18)",
            gridDash=[2, 6],
            domain=False,
            tickColor="#94a3b8",
            labelColor="#5b667a",
        )
        .configure_legend(
            labelColor="#4a5568",
            titleColor="#172033",
            labelFontSize=12,
            titleFontSize=13,
            symbolType="circle",
        )
        .configure_title(color="#172033", fontSize=16, fontWeight="bold", anchor="start")
    )


def normalize_date_selection(selection, default_start, default_end):
    if isinstance(selection, tuple):
        values = list(selection)
    elif isinstance(selection, list):
        values = selection
    else:
        values = [selection]

    if len(values) == 0:
        return default_start, default_end
    if len(values) == 1:
        return values[0], values[0]
    return values[0], values[1]


def _coerce_date_value(value, fallback):
    try:
        if value is None:
            return fallback
        return pd.Timestamp(value).date()
    except Exception:
        return fallback


def parse_week_label_sort_key(week_label):
    normalized = str(week_label or "").strip().upper()
    if "-W" not in normalized:
        return None
    year_part, week_part = normalized.split("-W", 1)
    try:
        return int(year_part) * 100 + int(week_part)
    except ValueError:
        return None


def build_week_filter_options(dataframe, date_basis):
    if dataframe is None or dataframe.empty or date_basis not in dataframe.columns:
        return []

    date_series = pd.to_datetime(dataframe[date_basis], errors="coerce")
    valid_mask = date_series.notna()
    if not valid_mask.any():
        return []

    iso = date_series[valid_mask].dt.isocalendar()
    weeks = pd.DataFrame(
        {
            "iso_year": iso["year"].astype(int),
            "iso_week": iso["week"].astype(int),
        }
    )
    weeks["label"] = weeks["iso_year"].astype(str) + "-W" + weeks["iso_week"].astype(str).str.zfill(2)
    weeks["sort_key"] = weeks["iso_year"] * 100 + weeks["iso_week"]
    weeks = weeks.drop_duplicates(subset=["sort_key"]).sort_values("sort_key").reset_index(drop=True)
    return weeks.to_dict("records")


def ensure_filter_defaults(result, date_basis, full_product_summary):
    if result is None or result.empty or date_basis not in result.columns:
        return None

    available_dates = pd.to_datetime(result[date_basis], errors="coerce").dropna().sort_values()
    if available_dates.empty:
        return None

    min_date = available_dates.min().date()
    max_date = available_dates.max().date()
    week_options = build_week_filter_options(result, date_basis)
    week_labels = [option["label"] for option in week_options]
    all_products = full_product_summary["Product Label"].tolist()

    if st.session_state.get("reset_filters_trigger"):
        st.session_state["date_from"] = min_date
        st.session_state["date_to"] = max_date
        st.session_state["week_from"] = week_labels[0] if week_labels else None
        st.session_state["week_to"] = week_labels[-1] if week_labels else None
        st.session_state["selected_products_filter"] = all_products
        st.session_state["analysis_search_term"] = ""
        st.session_state["analysis_change_direction"] = ["Increase", "Decrease", "No Change"]
        st.session_state["analysis_only_alerts"] = False
        st.session_state["reset_filters_trigger"] = False

    st.session_state["date_from"] = _coerce_date_value(st.session_state.get("date_from"), min_date)
    st.session_state["date_to"] = _coerce_date_value(st.session_state.get("date_to"), max_date)

    if st.session_state["date_from"] < min_date or st.session_state["date_from"] > max_date:
        st.session_state["date_from"] = min_date
    if st.session_state["date_to"] < min_date or st.session_state["date_to"] > max_date:
        st.session_state["date_to"] = max_date

    if week_labels:
        if st.session_state.get("week_from") not in week_labels:
            st.session_state["week_from"] = week_labels[0]
        if st.session_state.get("week_to") not in week_labels:
            st.session_state["week_to"] = week_labels[-1]
    else:
        st.session_state["week_from"] = None
        st.session_state["week_to"] = None

    selected_products = st.session_state.get("selected_products_filter")
    if not isinstance(selected_products, list):
        selected_products = list(all_products)
    else:
        selected_products = [product for product in selected_products if product in all_products]
    st.session_state["selected_products_filter"] = selected_products
    st.session_state.setdefault("analysis_search_term", "")
    st.session_state.setdefault("analysis_change_direction", ["Increase", "Decrease", "No Change"])
    st.session_state.setdefault("analysis_only_alerts", False)
    return {
        "min_date": min_date,
        "max_date": max_date,
        "week_options": week_options,
        "all_products": all_products,
    }


def apply_date_week_filters(dataframe, filter_state, *, date_basis_override=None):
    if dataframe is None:
        return pd.DataFrame()

    filtered = dataframe.copy()
    date_basis = date_basis_override or filter_state.get("date_basis", DATE_OPTIONS[0])
    if date_basis not in filtered.columns:
        return filtered

    date_series = pd.to_datetime(filtered[date_basis], errors="coerce")
    mask = date_series.notna()

    selected_start_date = filter_state.get("selected_start_date")
    selected_end_date = filter_state.get("selected_end_date")
    if selected_start_date is not None and selected_end_date is not None:
        mask &= date_series.dt.date.between(selected_start_date, selected_end_date)

    week_from_sort = parse_week_label_sort_key(filter_state.get("week_from"))
    week_to_sort = parse_week_label_sort_key(filter_state.get("week_to"))
    if week_from_sort is not None or week_to_sort is not None:
        iso = date_series.dt.isocalendar()
        week_sort = iso["year"].astype("Int64") * 100 + iso["week"].astype("Int64")
        if week_from_sort is not None:
            mask &= week_sort >= week_from_sort
        if week_to_sort is not None:
            mask &= week_sort <= week_to_sort

    return filtered.loc[mask].copy()


def apply_analysis_filters(dataframe, filter_state):
    filtered = apply_date_week_filters(dataframe, filter_state)

    selected_products = filter_state.get("selected_products", [])
    if selected_products:
        filtered = filtered[filtered["Product Label"].isin(selected_products)]
    else:
        filtered = filtered.iloc[0:0]

    search_term = str(filter_state.get("search_term", "")).strip()
    if search_term:
        query = search_term.lower()
        filtered = filtered[
            filtered["Part Number"].str.lower().str.contains(query, na=False)
            | filtered["Part Description"].str.lower().str.contains(query, na=False)
        ]

    selected_change_directions = filter_state.get(
        "selected_change_directions",
        ["Increase", "Decrease", "No Change"],
    )
    filtered = filtered[filtered["Change Direction"].isin(selected_change_directions)]

    if filter_state.get("only_alerts"):
        filtered = filtered[filtered["Alert"]]

    return filtered.copy()


def format_active_filter_summary(filter_state):
    date_label = format_workspace_date_range(filter_state)
    week_from = filter_state.get("week_from") or "n/a"
    week_to = filter_state.get("week_to") or "n/a"
    return f"Daty: {date_label} | Tygodnie: {week_from} -> {week_to}"


def build_filter_state(
    date_basis,
    selected_start_date,
    selected_end_date,
    selected_products,
    search_term,
    selected_change_directions,
    only_alerts,
    full_product_summary,
    week_from,
    week_to,
):
    filter_state = {
        "date_basis": date_basis,
        "selected_start_date": selected_start_date,
        "selected_end_date": selected_end_date,
        "selected_products": selected_products,
        "search_term": search_term,
        "selected_change_directions": selected_change_directions,
        "only_alerts": only_alerts,
        "full_product_summary": full_product_summary,
        "week_from": week_from,
        "week_to": week_to,
    }
    st.session_state["active_filters"] = {
        "date_basis": date_basis,
        "date_from": selected_start_date,
        "date_to": selected_end_date,
        "week_from": week_from,
        "week_to": week_to,
        "selected_products": list(selected_products),
        "search_term": search_term,
        "selected_change_directions": list(selected_change_directions),
        "only_alerts": bool(only_alerts),
    }
    return filter_state


def render_filter_panel_shell(
    kicker="Panel Nawigacji",
    title="Filtry i kontekst analizy",
    copy=(
        "Ten panel pozostaje stale widoczny, aby filtrowanie, kalendarz i kontrola zakresu "
        "byly zawsze pod reka podczas pracy z dashboardem."
    ),
):
    st.markdown(
        f"""
        <div class="filter-panel-shell">
            <div class="filter-panel-kicker">{html.escape(str(kicker))}</div>
            <div class="filter-panel-title">{html.escape(str(title))}</div>
            <div class="filter-panel-copy">
                {html.escape(str(copy))}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_filter_controls(result):
    return render_filters_panel(result)


def render_date_week_filters(result, date_basis):
    full_product_summary = summarize_products(result)
    defaults = ensure_filter_defaults(result, date_basis, full_product_summary)
    if defaults is None:
        st.warning("Brak poprawnych wartosci dat dla aktywnej osi analizy.")
        return None

    min_date = defaults["min_date"]
    max_date = defaults["max_date"]
    week_options = defaults["week_options"]
    week_labels = [option["label"] for option in week_options]

    action_left, action_right = st.columns(2, gap="small")
    with action_right:
        if st.button("Reset filtrów", key="reset_analysis_filters", use_container_width=True):
            st.session_state["reset_filters_trigger"] = True
            st.rerun()

    st.markdown("###### Zakres dat")
    date_left, date_right = st.columns(2, gap="small")
    with date_left:
        selected_start_date = st.date_input(
            "Data od",
            min_value=min_date,
            max_value=max_date,
            key="date_from",
        )
    with date_right:
        selected_end_date = st.date_input(
            "Data do",
            min_value=min_date,
            max_value=max_date,
            key="date_to",
        )
    if selected_start_date > selected_end_date:
        selected_start_date, selected_end_date = selected_end_date, selected_start_date
        st.session_state["date_from"] = selected_start_date
        st.session_state["date_to"] = selected_end_date
        st.warning("Zamieniono kolejność dat, aby zachować poprawny zakres analizy.")

    st.markdown("###### Zakres tygodni")
    if week_labels:
        week_left, week_right = st.columns(2, gap="small")
        with week_left:
            week_from = st.selectbox("Tydzień od", options=week_labels, key="week_from")
        with week_right:
            week_to = st.selectbox("Tydzień do", options=week_labels, key="week_to")
        week_from_sort = parse_week_label_sort_key(week_from)
        week_to_sort = parse_week_label_sort_key(week_to)
        if week_from_sort is not None and week_to_sort is not None and week_from_sort > week_to_sort:
            week_from, week_to = week_to, week_from
            st.session_state["week_from"] = week_from
            st.session_state["week_to"] = week_to
            st.warning("Zamieniono kolejność tygodni, aby zachować poprawny zakres analizy.")
    else:
        week_from = None
        week_to = None
        st.info("Brak danych tygodniowych dla aktywnej osi dat.")

    st.caption(
        f"Aktywny zakres dat: {selected_start_date.strftime('%Y-%m-%d')} — {selected_end_date.strftime('%Y-%m-%d')}"
    )
    if week_from and week_to:
        st.caption(f"Aktywny zakres tygodni: {week_from} — {week_to}")

    return {
        "selected_start_date": selected_start_date,
        "selected_end_date": selected_end_date,
        "week_from": week_from,
        "week_to": week_to,
        "full_product_summary": full_product_summary,
        "all_products": defaults["all_products"],
    }


def render_filters_panel(result):
    render_section_header(
        "Filtry",
        "Zakres i oś analizy",
        "Utrzymuj ten sam kontekst pracy podczas przeglądania wszystkich zakładek raportu.",
    )
    date_basis = st.segmented_control(
        "Oś dat",
        DATE_OPTIONS,
        selection_mode="single",
        default=DATE_OPTIONS[0],
        required=True,
        key="analysis_date_basis",
        format_func=get_date_label,
        width="stretch",
    )
    date_basis = date_basis or DATE_OPTIONS[0]

    date_week_state = render_date_week_filters(result, date_basis)
    if date_week_state is None:
        return build_default_filter_state()

    selected_start_date = date_week_state["selected_start_date"]
    selected_end_date = date_week_state["selected_end_date"]
    week_from = date_week_state["week_from"]
    week_to = date_week_state["week_to"]
    full_product_summary = date_week_state["full_product_summary"]
    all_products = date_week_state["all_products"]

    st.markdown("###### Filtry produktowe")
    selected_products = st.multiselect(
        "Produkty",
        options=all_products,
        key="selected_products_filter",
    )
    search_term = st.text_input("Szukaj po numerze lub opisie", key="analysis_search_term")
    st.markdown("###### Kierunek zmiany")
    selected_change_directions = st.segmented_control(
        "Kierunek zmiany",
        options=["Increase", "Decrease", "No Change"],
        selection_mode="multi",
        default=["Increase", "Decrease", "No Change"],
        key="analysis_change_direction",
        format_func=get_change_label,
        width="stretch",
    )
    selected_change_directions = selected_change_directions or ["Increase", "Decrease", "No Change"]
    only_alerts = st.checkbox(f"Tylko alerty >= {THRESHOLD}%", key="analysis_only_alerts")

    filter_state = build_filter_state(
        date_basis,
        selected_start_date,
        selected_end_date,
        selected_products,
        search_term,
        selected_change_directions,
        only_alerts,
        full_product_summary,
        week_from,
        week_to,
    )
    preview_filtered = apply_analysis_filters(result, filter_state)
    st.caption(format_active_filter_summary(filter_state))
    st.caption(f"Liczba rekordów po filtracji: {len(preview_filtered):,}")
    return filter_state


def render_welcome_side_panel(prev_file, current_file):
    pending_meta = []
    if prev_file is not None:
        pending_meta.append({"file_name": prev_file.name})
    if current_file is not None:
        pending_meta.append({"file_name": current_file.name})
    brand_context = detect_brand_context(*pending_meta) if pending_meta else detect_brand_context()

    render_filter_panel_shell(
        kicker="Workspace",
        title="Panel aplikacji",
        copy=(
            "Upload, status plikow i pozniejszy panel filtrow pozostaja w jednej stalej kolumnie roboczej."
        ),
    )
    render_side_panel_brand(brand_context)
    render_file_slot_cards(prev_file=prev_file, current_file=current_file)
    st.info(
        "Po zaladowaniu obu plikow ten panel zostanie uzupelniony o filtry, status formatu i akcje eksportu."
    )


def render_analysis_side_panel(result, brand_context, prev_meta=None, curr_meta=None):
    render_filter_panel_shell(
        kicker="Analysis Controls",
        title="Filtry i status analizy",
        copy="Lewa kolumna utrzymuje upload, status plikow, filtry oraz eksport w jednym miejscu pracy.",
    )
    render_side_panel_brand(brand_context)
    st.caption(brand_context.get("format_copy", ""))
    render_file_slot_cards(prev_meta=prev_meta, curr_meta=curr_meta)
    return render_filter_controls(result)


def build_ui_helpers():
    return SimpleNamespace(
        apply_chart_theme=apply_chart_theme,
        apply_plotly_theme=apply_plotly_theme,
        available_detail_columns=available_detail_columns,
        build_alert_items=build_alert_items,
        build_change_mix_chart=build_change_mix_chart,
        build_change_mix_source=build_change_mix_source,
        build_dashboard_kpi_metrics=build_dashboard_kpi_metrics,
        build_delta_chart=build_delta_chart,
        build_detail_export_table=build_detail_export_table,
        build_kpi_metrics=build_kpi_metrics,
        build_matrix=build_matrix,
        build_product_bar_chart=build_product_bar_chart,
        build_product_bar_source=build_product_bar_source,
        build_product_waterfall_chart=build_product_waterfall_chart,
        build_product_detail_table=build_product_detail_table,
        build_quantity_chart=build_quantity_chart,
        build_weekly_delta_chart=build_weekly_delta_chart,
        build_weekly_focus_table=build_weekly_focus_table,
        build_weekly_quantity_chart=build_weekly_quantity_chart,
        build_weekly_summary=build_weekly_summary,
        format_signed_int=format_signed_int,
        get_change_label=get_change_label,
        get_date_label=get_date_label,
        get_metric_label=get_metric_label,
        max_matrix_style_cells=MAX_MATRIX_STYLE_CELLS,
        prepare_weekly_display_table=prepare_weekly_display_table,
        render_alerts=render_alerts,
        render_chart_table_switch=render_chart_table_switch,
        render_kpi_cards=render_kpi_cards,
        render_kpi_row=render_kpi_row,
        render_section_header=render_section_header,
        style_matrix=style_matrix,
        summarize_dates=summarize_dates,
        threshold=THRESHOLD,
        to_excel_bytes=to_excel_bytes,
    )


def build_reference_snapshot(weekly_summary, selected_end_date):
    reference_week = get_last_completed_reference_week(selected_end_date)
    reference_row, previous_week_row = get_reference_week_rows(weekly_summary)
    return {
        "reference_week_label": (
            reference_row["Week Label"]
            if reference_row is not None
            else reference_week.week_label
        ),
        "reference_range_label": (
            format_week_range(reference_row["Week Start"], reference_row["Week End"])
            if reference_row is not None
            else format_week_range(reference_week.week_start, reference_week.week_end)
        ),
        "reference_release_delta": (
            format_signed_int(reference_row["Delta"])
            if reference_row is not None
            else "+0"
        ),
        "reference_release_pct": (
            format_percent_display(reference_row["Release Percent Label"])
            if reference_row is not None
            else "n/a"
        ),
        "reference_wow_delta": (
            format_signed_int(reference_row["WoW Delta"])
            if reference_row is not None
            else "+0"
        ),
        "reference_wow_pct": (
            format_percent_display(reference_row["WoW Percent Label"])
            if reference_row is not None
            else "n/a"
        ),
        "reference_working_days": (
            int(reference_row["Working_Days_PL"])
            if reference_row is not None
            else 0
        ),
        "reference_per_day": (
            "n/a"
            if reference_row is None or pd.isna(reference_row["Avg Current / Working Day"])
            else f"{float(reference_row['Avg Current / Working Day']):,.2f} / dzien"
        ),
        "previous_week_label": (
            previous_week_row["Week Label"]
            if previous_week_row is not None
            else "brak"
        ),
        "reference_curr_qty": (
            f"{float(reference_row['Quantity_Curr']):,.0f}"
            if reference_row is not None
            else "0"
        ),
        "reference_prev_qty": (
            f"{float(reference_row['Quantity_Prev']):,.0f}"
            if reference_row is not None
            else "0"
        ),
    }


def build_module_context(
    filtered_df,
    planner_source,
    product_summary,
    date_summary,
    weekly_summary,
    key_findings,
    prev_meta,
    curr_meta,
    date_basis,
    selected_start_date,
    selected_end_date,
    excel_bytes=None,
    csv_bytes=None,
    professional_excel_bytes=None,
):
    auth_user = get_auth_user()
    user_role = get_user_role(auth_user=auth_user)
    return ModuleDataContext(
        filtered_df=filtered_df,
        planner_source=planner_source,
        product_summary=product_summary,
        date_summary=date_summary,
        weekly_summary=weekly_summary,
        key_findings=key_findings,
        prev_meta=prev_meta,
        curr_meta=curr_meta,
        date_basis=date_basis,
        selected_start_date=selected_start_date,
        selected_end_date=selected_end_date,
        auth_user=auth_user,
        user_role=user_role,
        module_access="none",
        excel_bytes=excel_bytes,
        csv_bytes=csv_bytes,
        professional_excel_bytes=professional_excel_bytes,
        reference=build_reference_snapshot(weekly_summary, selected_end_date),
    )


def render_module_frame(
    active_module,
    filtered_df,
    planner_source,
    product_summary,
    date_summary,
    weekly_summary,
    key_findings,
    prev_meta,
    curr_meta,
    date_basis,
    selected_start_date,
    selected_end_date,
    excel_bytes=None,
    csv_bytes=None,
):
    auth_user = get_auth_user()
    if not can_access_module(active_module, auth_user=auth_user):
        allowed_modules = get_allowed_modules(auth_user=auth_user)
        fallback_module = allowed_modules[0] if allowed_modules else "dashboard"
        st.session_state["active_module"] = fallback_module
        active_module = fallback_module

    brand_context = detect_brand_context(prev_meta, curr_meta)
    render_compact_header(
        brand_context,
        prev_meta,
        curr_meta,
        date_basis,
        selected_start_date,
        selected_end_date,
    )

    report_metadata = [
        {"label": "Format", "value": describe_format_context(prev_meta, curr_meta)},
        {"label": "Numer PO", "value": curr_meta.get("po_number", "n/a")},
        {"label": "Planista", "value": curr_meta.get("planner_name", "n/a")},
        {"label": "E-mail", "value": curr_meta.get("planner_email", "n/a")},
        {
            "label": "Zakres analizy",
            "value": f"{selected_start_date:%Y-%m-%d} - {selected_end_date:%Y-%m-%d}",
        },
        {"label": "Modul", "value": MODULE_LABELS.get(active_module, "Dashboard")},
    ]
    render_report_metadata(report_metadata)

    ui = build_ui_helpers()
    module_data = build_module_context(
        filtered_df,
        planner_source,
        product_summary,
        date_summary,
        weekly_summary,
        key_findings,
        prev_meta,
        curr_meta,
        date_basis,
        selected_start_date,
        selected_end_date,
        excel_bytes=excel_bytes,
        csv_bytes=csv_bytes,
    )
    module_data.module_access = get_module_access_level(active_module, auth_user=auth_user)

    module_renderers = {
        "dashboard": render_dashboard_module,
        "planner": render_planner_module,
        "reports": render_reports_module,
        "details": render_details_module,
        "admin": render_admin_module,
    }
    module_renderer = module_renderers.get(active_module, render_dashboard_module)
    module_renderer(module_data, ui)


def render_analysis_main(
    filtered_df,
    planner_source,
    product_summary,
    date_summary,
    weekly_summary,
    key_findings,
    prev_meta,
    curr_meta,
    date_basis,
    selected_start_date,
    selected_end_date,
    excel_bytes=None,
    csv_bytes=None,
):
    brand_context = detect_brand_context(prev_meta, curr_meta)
    render_compact_header(
        brand_context,
        prev_meta,
        curr_meta,
        date_basis,
        selected_start_date,
        selected_end_date,
    )

    reference_week = get_last_completed_reference_week(selected_end_date)
    reference_row, previous_week_row = get_reference_week_rows(weekly_summary)
    reference_week_label = (
        reference_row["Week Label"] if reference_row is not None else reference_week.week_label
    )
    reference_range_label = (
        format_week_range(reference_row["Week Start"], reference_row["Week End"])
        if reference_row is not None
        else format_week_range(reference_week.week_start, reference_week.week_end)
    )
    reference_release_delta = (
        format_signed_int(reference_row["Delta"]) if reference_row is not None else "+0"
    )
    reference_release_pct = (
        format_percent_display(reference_row["Release Percent Label"])
        if reference_row is not None
        else "n/a"
    )
    reference_wow_delta = (
        format_signed_int(reference_row["WoW Delta"]) if reference_row is not None else "+0"
    )
    reference_wow_pct = (
        format_percent_display(reference_row["WoW Percent Label"])
        if reference_row is not None
        else "n/a"
    )
    reference_working_days = (
        int(reference_row["Working_Days_PL"]) if reference_row is not None else 0
    )
    reference_per_day = (
        "n/a"
        if reference_row is None or pd.isna(reference_row["Avg Current / Working Day"])
        else f"{float(reference_row['Avg Current / Working Day']):,.2f} / dzien"
    )
    previous_week_label = (
        previous_week_row["Week Label"] if previous_week_row is not None else "brak"
    )

    report_metadata = [
        {"label": "Format", "value": describe_format_context(prev_meta, curr_meta)},
        {"label": "Numer PO", "value": curr_meta.get("po_number", "n/a")},
        {"label": "Planista", "value": curr_meta.get("planner_name", "n/a")},
        {"label": "E-mail", "value": curr_meta.get("planner_email", "n/a")},
        {
            "label": "Zakres analizy",
            "value": f"{selected_start_date:%Y-%m-%d} — {selected_end_date:%Y-%m-%d}",
        },
    ]
    render_report_metadata(report_metadata)

    analytics_empty = filtered_df.empty
    if analytics_empty and planner_source.empty:
        st.warning(
            "Po zastosowaniu filtrów nie ma danych do pokazania. Poszerz zakres dat albo przywróć produkty w panelu po lewej stronie."
        )
        return

    if analytics_empty:
        st.warning(
            "Główne zakładki analityczne są puste dla bieżących filtrów, ale Planner nadal liczy aktualny demand według Ship Date i Quantity_Curr."
        )
    else:
        render_section_header(
            "KPI",
            "Najważniejsze wskaźniki",
            "Karty poniżej pokazują główne liczby do szybkiego odczytu bez przeskakiwania między zakładkami.",
        )
        render_kpi_cards(build_kpi_metrics(filtered_df, product_summary))

        render_section_header(
            "Alerts & Insights",
            "Priorytety do sprawdzenia",
            "Najważniejsze sygnały, które warto zweryfikować w pierwszej kolejności.",
        )
        render_alerts(build_alert_items(filtered_df, key_findings))

        render_section_header(
            "Reference Week",
            "Szybki odczyt tygodniowy",
            (
                f"Analiza tygodniowa odnosi się do {reference_week_label} ({reference_range_label}). "
                f"Data referencyjna: {selected_end_date:%Y-%m-%d}."
            ),
        )
        render_kpi_cards(
            [
                {
                    "label": "Wolumen tygodnia",
                    "value": f"{float(reference_row['Quantity_Curr']):,.0f}" if reference_row is not None else "0",
                    "copy": f"Bilans release: {reference_release_delta}",
                    "tone": "neutral",
                },
                {
                    "label": "Zmiana vs poprzedni release",
                    "value": reference_release_pct,
                    "copy": f"Poprzedni wolumen: {float(reference_row['Quantity_Prev']):,.0f}" if reference_row is not None else "Poprzedni wolumen: 0",
                    "tone": "neutral",
                },
                {
                    "label": "Zmiana WoW",
                    "value": reference_wow_delta,
                    "copy": f"{reference_wow_pct} względem {previous_week_label}",
                    "tone": "neutral",
                },
                {
                    "label": "Dni robocze PL",
                    "value": f"{reference_working_days}",
                    "copy": reference_per_day,
                    "tone": "neutral",
                },
            ]
        )

    dashboard_tab, weekly_tab, product_tab, planner_tab, matrix_tab, detail_tab = st.tabs(
        ["Dashboard", "Analiza tygodniowa", "Raport produktu", "Planner", "Macierz release'u", "Dane szczegółowe"]
    )

    with dashboard_tab:
        render_section_header(
            "Dashboard",
            f"Trend zmian według osi: {get_date_label(date_basis)}",
            "Widok główny zbiera najważniejsze wykresy, strukturę zmian oraz szybki podgląd produktów z największym ruchem.",
        )
        render_chart_table_switch(
            "dashboard_trend",
            build_quantity_chart(date_summary, get_date_label(date_basis)),
            date_summary,
            table_height=360,
        )

        trend_left, trend_right = st.columns([1.45, 1], gap="large")
        with trend_left:
            render_chart_table_switch(
                "dashboard_delta",
                build_delta_chart(date_summary, get_date_label(date_basis)),
                date_summary,
                table_height=320,
            )
        with trend_right:
            st.subheader("Struktura zmian")
            render_chart_table_switch(
                "dashboard_mix",
                build_change_mix_chart(filtered_df),
                build_change_mix_source(filtered_df),
                table_height=240,
            )

        increase_chart, increase_title = build_product_bar_chart(product_summary, "increase")
        decrease_chart, decrease_title = build_product_bar_chart(product_summary, "decrease")
        dashboard_left, dashboard_right = st.columns(2)

        with dashboard_left:
            st.subheader(increase_title)
            if increase_chart is None:
                st.info("Brak produktow ze wzrostem w aktualnym filtrowaniu.")
            else:
                render_chart_table_switch(
                    "dashboard_increase",
                    increase_chart,
                    build_product_bar_source(product_summary, "increase"),
                    table_height=340,
                )

        with dashboard_right:
            st.subheader(decrease_title)
            if decrease_chart is None:
                st.info("Brak produktow ze spadkiem w aktualnym filtrowaniu.")
            else:
                render_chart_table_switch(
                    "dashboard_decrease",
                    decrease_chart,
                    build_product_bar_source(product_summary, "decrease"),
                    table_height=340,
                )

        st.subheader("Najwazniejsze zmiany")
        highlight_table = (
            product_summary.assign(Abs_Delta=product_summary["Delta"].abs())
            .sort_values("Abs_Delta", ascending=False)
            .drop(columns=["Abs_Delta"])
            .head(10)
        )
        highlight_table["Quantity_Prev"] = highlight_table["Quantity_Prev"].map(
            lambda value: f"{value:,.0f}"
        )
        highlight_table["Quantity_Curr"] = highlight_table["Quantity_Curr"].map(
            lambda value: f"{value:,.0f}"
        )
        highlight_table["Delta"] = highlight_table["Delta"].map(format_signed_int)
        highlight_table = highlight_table.rename(
            columns={
                "Part Number": "Numer czesci",
                "Part Description": "Opis produktu",
                "Quantity_Prev": "Poprzednia ilosc",
                "Quantity_Curr": "Aktualna ilosc",
                "Delta": "Zmiana ilosci",
                "Alert_Count": "Liczba alertow",
                "Change Direction": "Kierunek zmiany",
            }
        )
        st.dataframe(highlight_table, use_container_width=True, height=360)

        st.subheader("Tygodnie ISO")
        weekly_chart = build_weekly_quantity_chart(weekly_summary)
        weekly_preview = prepare_weekly_display_table(weekly_summary).tail(8)
        render_chart_table_switch(
            "dashboard_weekly",
            weekly_chart,
            weekly_preview,
            chart_empty_message="Brak danych tygodniowych do wykresu.",
            table_height=320,
        )

    with weekly_tab:
        render_section_header(
            "Weekly View",
            "Analiza tygodniowa oparta na datach",
            "Ten widok agreguje dzienne dane do poziomu tygodni ISO i ułatwia porównanie release-over-release oraz week-over-week.",
        )
        weekly_partial = weekly_summary[
            weekly_summary["Is Partial Range"] | ~weekly_summary["Is Closed Week"]
        ]
        if not weekly_partial.empty:
            st.info(
                "W tabeli i wykresach tygodnie oznaczone jako 'Partial range' lub 'Open week' "
                "obejmuja niepelny zakres albo nie byly jeszcze zakonczone wzgledem daty referencyjnej."
            )

        weekly_qty_chart = build_weekly_quantity_chart(weekly_summary)
        render_chart_table_switch(
            "weekly_quantity",
            weekly_qty_chart,
            prepare_weekly_display_table(weekly_summary),
            chart_empty_message="Brak danych tygodniowych do wykresu.",
            table_height=360,
        )

        weekly_left, weekly_right = st.columns([1.3, 1], gap="large")
        with weekly_left:
            weekly_delta_chart = build_weekly_delta_chart(weekly_summary)
            render_chart_table_switch(
                "weekly_delta",
                weekly_delta_chart,
                prepare_weekly_display_table(weekly_summary),
                chart_empty_message="Brak danych tygodniowych do wykresu delta.",
                table_height=320,
            )
        with weekly_right:
            weekly_focus = pd.DataFrame(
                [
                    {
                        "Widok": "Referencyjny tydzien",
                        "Tydzien ISO": reference_week_label,
                        "Aktualny release": (
                            f"{float(reference_row['Quantity_Curr']):,.0f}"
                            if reference_row is not None
                            else "0"
                        ),
                        "Poprzedni release": (
                            f"{float(reference_row['Quantity_Prev']):,.0f}"
                            if reference_row is not None
                            else "0"
                        ),
                        "Delta release": reference_release_delta,
                        "Zmiana release %": reference_release_pct,
                        "Delta WoW": reference_wow_delta,
                        "Zmiana WoW %": reference_wow_pct,
                    },
                    {
                        "Widok": "Poprzedni tydzien",
                        "Tydzien ISO": previous_week_label,
                        "Aktualny release": (
                            f"{float(previous_week_row['Quantity_Curr']):,.0f}"
                            if previous_week_row is not None
                            else "0"
                        ),
                        "Poprzedni release": (
                            f"{float(previous_week_row['Quantity_Prev']):,.0f}"
                            if previous_week_row is not None
                            else "0"
                        ),
                        "Delta release": (
                            format_signed_int(previous_week_row["Delta"])
                            if previous_week_row is not None
                            else "+0"
                        ),
                        "Zmiana release %": (
                            format_percent_display(previous_week_row["Release Percent Label"])
                            if previous_week_row is not None
                            else "n/a"
                        ),
                        "Delta WoW": (
                            format_signed_int(previous_week_row["WoW Delta"])
                            if previous_week_row is not None
                            else "+0"
                        ),
                        "Zmiana WoW %": (
                            format_percent_display(previous_week_row["WoW Percent Label"])
                            if previous_week_row is not None
                            else "n/a"
                        ),
                    },
                ]
            )
            st.subheader("Porownanie tygodni")
            st.dataframe(weekly_focus, use_container_width=True, height=240)

        weekly_table = prepare_weekly_display_table(weekly_summary)
        st.subheader("Tabela tygodniowa")
        st.dataframe(weekly_table, use_container_width=True, height=420)

    with product_tab:
        if product_summary.empty:
            st.info("Brak danych produktowych dla aktywnych filtrów.")
        else:
            render_section_header(
                "Product Drilldown",
                "Analiza wybranego produktu",
                "Skup się na jednym materiale i prześledź jego ruch po dniach oraz tygodniach bez utraty kontekstu filtrowania.",
            )
            selected_product_label = st.selectbox(
                "Wybierz produkt",
                options=product_summary["Product Label"].tolist(),
            )
            product_detail = filtered_df[
                filtered_df["Product Label"] == selected_product_label
            ].sort_values(date_basis)
            product_date_summary = summarize_dates(product_detail, date_basis)

            product_metrics = st.columns(4)
            product_metrics[0].metric(
                "Poprzednia ilosc", f"{product_detail['Quantity_Prev'].sum():,.0f}"
            )
            product_metrics[1].metric(
                "Aktualna ilosc", f"{product_detail['Quantity_Curr'].sum():,.0f}"
            )
            product_metrics[2].metric(
                "Bilans zmian", f"{product_detail['Delta'].sum():+,.0f}"
            )
            product_metrics[3].metric("Liczba alertow", int(product_detail["Alert"].sum()))

            render_chart_table_switch(
                "product_quantity",
                build_quantity_chart(product_date_summary, get_date_label(date_basis)),
                product_date_summary,
                table_height=320,
            )
            render_chart_table_switch(
                "product_delta",
                build_delta_chart(product_date_summary, get_date_label(date_basis)),
                product_date_summary,
                table_height=320,
            )

            product_weekly_summary = build_weekly_summary(
                product_detail,
                date_basis,
                selected_start_date,
                selected_end_date,
                selected_end_date,
                THRESHOLD,
            )
            st.subheader("Tygodnie ISO dla produktu")
            product_weekly_chart = build_weekly_quantity_chart(product_weekly_summary)
            render_chart_table_switch(
                "product_weekly",
                product_weekly_chart,
                prepare_weekly_display_table(product_weekly_summary),
                chart_empty_message="Brak danych tygodniowych dla wybranego produktu.",
                table_height=280,
            )

            product_table = product_detail[available_detail_columns(product_detail)].copy()
            product_table["Ship Date"] = product_table["Ship Date"].dt.strftime("%Y-%m-%d")
            product_table["Receipt Date"] = product_table["Receipt Date"].dt.strftime("%Y-%m-%d")
            product_table["Change Direction"] = product_table["Change Direction"].map(
                get_change_label
            )
            product_table["Alert"] = product_table["Alert"].map(
                lambda value: "Tak" if value else "Nie"
            )
            product_table = product_table.rename(
                columns={
                    "Part Number": "Numer czesci",
                    "Part Description": "Opis produktu",
                    "Origin Doc": "Origin Doc",
                    "Item": "Pozycja",
                    "Ship To": "Ship-to",
                    "Customer Material": "Material klienta",
                    "Unrestricted Qty": "Ilosc unrestr.",
                    "Unloading Point": "Punkt rozladunku",
                    "Ship Date": "Data wysylki",
                    "Receipt Date": "Data odbioru",
                    "Unit of Measure": "JM",
                    "CumQty": "CumQty",
                    "Quantity_Prev": "Poprzednia ilosc",
                    "Quantity_Curr": "Aktualna ilosc",
                    "Delta": "Zmiana ilosci",
                    "Percent Change": "Zmiana %",
                    "Demand Status": "Status popytu",
                    "Change Direction": "Kierunek zmiany",
                    "Alert": "Alert",
                }
            )
            st.dataframe(product_table, use_container_width=True, height=360)

    with planner_tab:
        render_planner_tab(planner_source, curr_meta)

    with matrix_tab:
        render_section_header(
            "Release Matrix",
            "Macierz podobna do arkusza release'u",
            "Macierz zachowuje układ bliski pracy w Excelu, ale pozostaje spójna wizualnie z całym dashboardem.",
        )
        matrix_metric = st.segmented_control(
            "Metryka",
            options=["Current Quantity", "Previous Quantity", "Delta", "Percent Change"],
            selection_mode="single",
            default="Current Quantity",
            required=True,
            format_func=get_metric_label,
            width="stretch",
        )
        matrix_metric = matrix_metric or "Current Quantity"
        matrix = build_matrix(filtered_df, date_basis, matrix_metric)
        matrix_cells = matrix.shape[0] * max(matrix.shape[1], 1)

        if matrix.empty:
            st.info("Brak danych do macierzy.")
        elif matrix_cells <= MAX_MATRIX_STYLE_CELLS:
            st.dataframe(
                style_matrix(matrix, matrix_metric),
                use_container_width=True,
                height=520,
            )
        else:
            st.info(
                "Macierz jest zbyt duza do stylowania, dlatego pokazuje ja bez dodatkowego formatowania."
            )
            st.dataframe(matrix, use_container_width=True, height=520)

    with detail_tab:
        render_section_header(
            "Detailed Data",
            "Dane szczegolowe",
            "Pełny podgląd przefiltrowanych wierszy do szybkiej walidacji oraz eksportu do dalszej pracy operacyjnej.",
        )
        preview_limit = st.selectbox(
            "Liczba wierszy w podgladzie",
            options=[100, 250, 500, 1000],
            index=2,
        )
        detail_table = build_detail_export_table(filtered_df)

        if len(detail_table) > preview_limit:
            st.info(
                f"Pokazuje pierwsze {preview_limit} z {len(detail_table)} wierszy. "
                "Pelny raport jest dostepny do pobrania."
            )
        st.dataframe(
            detail_table.head(preview_limit),
            use_container_width=True,
            height=420,
        )

        if excel_bytes is None:
            current_matrix_for_export = build_matrix(filtered_df, date_basis, "Current Quantity")
            delta_matrix_for_export = build_matrix(filtered_df, date_basis, "Delta")
            excel_bytes = to_excel_bytes(
                filtered_df,
                weekly_summary,
                current_matrix_for_export,
                delta_matrix_for_export,
                prev_meta,
                curr_meta,
                product_summary,
                date_basis,
                selected_start_date,
                selected_end_date,
                key_findings,
            )
        if csv_bytes is None:
            csv_bytes = detail_table.to_csv(index=False).encode("utf-8")

        download_left, download_right = st.columns(2)
        with download_left:
            st.download_button(
                "Pobierz filtrowane dane CSV",
                data=csv_bytes,
                file_name="pjoter_development_release_change_filtered.csv",
                mime="text/csv",
            )
        with download_right:
            st.download_button(
                "Pobierz raport Excel",
                data=excel_bytes,
                file_name="pjoter_development_release_change_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


@st.cache_data(show_spinner=False)
def load_release(file_bytes, file_name):
    return load_release_file(file_bytes, file_name)


def compare_releases(prev_df, curr_df):
    return compare_release_frames(prev_df, curr_df, threshold=THRESHOLD)


def summarize_products(dataframe):
    product_summary = (
        dataframe.groupby(["Part Number", "Part Description", "Product Label"], as_index=False)
        .agg(
            Quantity_Prev=("Quantity_Prev", "sum"),
            Quantity_Curr=("Quantity_Curr", "sum"),
            Delta=("Delta", "sum"),
            Abs_Delta=("Abs Delta", "sum"),
            Alert_Count=("Alert", "sum"),
        )
        .sort_values("Delta", ascending=False)
    )
    product_summary["Change Direction"] = product_summary["Delta"].apply(
        lambda value: "Increase" if value > 0 else ("Decrease" if value < 0 else "No Change")
    )
    return product_summary.reset_index(drop=True)


def summarize_dates(dataframe, date_basis):
    if dataframe.empty:
        return pd.DataFrame(
            columns=["Analysis Date", "Quantity_Prev", "Quantity_Curr", "Delta", "Alerts"]
        )

    date_summary = (
        dataframe.groupby(date_basis, as_index=False)
        .agg(
            Quantity_Prev=("Quantity_Prev", "sum"),
            Quantity_Curr=("Quantity_Curr", "sum"),
            Delta=("Delta", "sum"),
            Alerts=("Alert", "sum"),
        )
        .sort_values(date_basis)
        .rename(columns={date_basis: "Analysis Date"})
    )
    return date_summary


def format_percent_display(value):
    return "Nowy tydzien" if value == "new" else value


def format_week_range(start_value, end_value):
    start_label = pd.Timestamp(start_value).strftime("%Y-%m-%d")
    end_label = pd.Timestamp(end_value).strftime("%Y-%m-%d")
    return f"{start_label} - {end_label}"


def get_reference_week_rows(weekly_summary):
    if weekly_summary.empty:
        return None, None

    reference_rows = weekly_summary[weekly_summary["Is Reference Week"]]
    if reference_rows.empty:
        closed_rows = weekly_summary[weekly_summary["Is Closed Week"]]
        if closed_rows.empty:
            return None, None
        reference_rows = closed_rows.tail(1)

    reference_row = reference_rows.iloc[0]
    previous_rows = weekly_summary[weekly_summary["Week Start"] < reference_row["Week Start"]]
    previous_row = previous_rows.tail(1).iloc[0] if not previous_rows.empty else None
    return reference_row, previous_row


def prepare_weekly_display_table(weekly_summary):
    if weekly_summary.empty:
        return pd.DataFrame()

    weekly_table = weekly_summary[
        [
            "Week Label",
            "Week Start",
            "Week End",
            "Week Status",
            "Working_Days_PL",
            "Products",
            "Quantity_Prev",
            "Quantity_Curr",
            "Delta",
            "Release Percent Label",
            "WoW Delta",
            "WoW Percent Label",
            "Avg Current / Working Day",
            "Any Weekly Alert",
        ]
    ].copy()
    weekly_table["Week Range"] = weekly_table.apply(
        lambda row: format_week_range(row["Week Start"], row["Week End"]),
        axis=1,
    )
    weekly_table["Release Delta"] = weekly_table["Delta"].map(format_signed_int)
    weekly_table["WoW Delta"] = weekly_table["WoW Delta"].map(format_signed_int)
    weekly_table["Previous Release"] = weekly_table["Quantity_Prev"].map(lambda value: f"{value:,.0f}")
    weekly_table["Current Release"] = weekly_table["Quantity_Curr"].map(lambda value: f"{value:,.0f}")
    weekly_table["Release Change %"] = weekly_table["Release Percent Label"].map(format_percent_display)
    weekly_table["WoW Change %"] = weekly_table["WoW Percent Label"].map(format_percent_display)
    weekly_table["Current / Working Day"] = weekly_table["Avg Current / Working Day"].map(
        lambda value: "n/a" if pd.isna(value) else f"{float(value):,.2f}"
    )
    weekly_table["Alert"] = weekly_table["Any Weekly Alert"].map(lambda value: "Tak" if value else "Nie")
    return weekly_table[
        [
            "Week Label",
            "Week Range",
            "Week Status",
            "Working_Days_PL",
            "Products",
            "Previous Release",
            "Current Release",
            "Release Delta",
            "Release Change %",
            "WoW Delta",
            "WoW Change %",
            "Current / Working Day",
            "Alert",
        ]
    ].rename(
        columns={
            "Week Label": "Tydzien ISO",
            "Week Range": "Zakres tygodnia",
            "Week Status": "Status",
            "Working_Days_PL": "Dni robocze PL",
            "Products": "Produkty",
        }
    )


def build_weekly_quantity_chart(weekly_summary):
    if weekly_summary.empty:
        return None

    chart_data = weekly_summary.copy()
    chart_data["Week Start"] = pd.to_datetime(chart_data["Week Start"])
    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=chart_data["Week Start"],
            y=chart_data["Quantity_Curr"],
            mode="lines",
            line={"color": "rgba(45,129,255,0.24)", "width": 0},
            fill="tozeroy",
            fillcolor="rgba(45,129,255,0.18)",
            name="Aktualny release area",
            hoverinfo="skip",
            showlegend=False,
        )
    )
    fig.add_trace(
        go.Scatter(
            x=chart_data["Week Start"],
            y=chart_data["Quantity_Prev"],
            mode="lines+markers",
            name="Poprzedni release",
            line={"color": "#8b949e", "width": 2.3},
            marker={"size": 6, "color": "#8b949e"},
            customdata=chart_data[["Week Label", "Quantity_Curr", "Delta", "Working_Days_PL", "Week Status"]].to_numpy(),
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "Poprzedni release: %{y:,.0f}<br>"
                "Aktualny release: %{customdata[1]:,.0f}<br>"
                "Delta: %{customdata[2]:+,.0f}<br>"
                "Dni robocze PL: %{customdata[3]}<br>"
                "Status: %{customdata[4]}<extra></extra>"
            ),
        )
    )
    fig.add_trace(
        go.Scatter(
            x=chart_data["Week Start"],
            y=chart_data["Quantity_Curr"],
            mode="lines+markers",
            name="Aktualny release",
            line={"color": "#2d81ff", "width": 3.2},
            marker={"size": 7, "color": "#f0f6fc", "line": {"color": "#2d81ff", "width": 2}},
            customdata=chart_data[["Week Label", "Quantity_Prev", "Delta", "Avg Current / Working Day", "Week Status"]].to_numpy(),
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "Aktualny release: %{y:,.0f}<br>"
                "Poprzedni release: %{customdata[1]:,.0f}<br>"
                "Delta: %{customdata[2]:+,.0f}<br>"
                "Na dzien roboczy: %{customdata[3]:,.2f}<br>"
                "Status: %{customdata[4]}<extra></extra>"
            ),
        )
    )
    fig.update_layout(
        height=360,
        hovermode="x unified",
        xaxis_title="Tydzien ISO",
        yaxis_title="Wolumen tygodniowy",
    )
    fig.update_xaxes(tickangle=-24)
    return fig


def build_weekly_delta_chart(weekly_summary):
    if weekly_summary.empty:
        return None

    chart_data = weekly_summary.copy()
    chart_data["Week Start"] = pd.to_datetime(chart_data["Week Start"])
    colors = [
        "#3fb950" if value > 0 else "#f85149" if value < 0 else "#8b949e"
        for value in chart_data["WoW Delta"]
    ]
    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=chart_data["Week Start"],
            y=chart_data["WoW Delta"],
            name="Delta WoW",
            marker={"color": colors, "line": {"color": "rgba(255,255,255,0.05)", "width": 1}},
            customdata=chart_data[["Week Label", "WoW Percent Label", "Working_Days_PL", "Week Status"]].to_numpy(),
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "Delta WoW: %{y:+,.0f}<br>"
                "Zmiana WoW %: %{customdata[1]}<br>"
                "Dni robocze PL: %{customdata[2]}<br>"
                "Status: %{customdata[3]}<extra></extra>"
            ),
        )
    )
    fig.add_trace(
        go.Scatter(
            x=chart_data["Week Start"],
            y=chart_data["Delta"],
            mode="lines+markers",
            name="Delta release",
            line={"color": "#00c4b4", "width": 2.1},
            marker={"size": 6, "color": "#00c4b4"},
            hovertemplate="Delta release: %{y:+,.0f}<extra></extra>",
        )
    )
    fig.update_layout(
        height=320,
        hovermode="x unified",
        xaxis_title="Tydzien ISO",
        yaxis_title="Delta tygodniowa",
    )
    fig.update_xaxes(tickangle=-24)
    return fig


def build_quantity_chart(date_summary, x_title):
    if date_summary.empty:
        return None

    chart_data = date_summary.sort_values("Analysis Date").copy()
    latest_point = chart_data.tail(1).copy()
    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=chart_data["Analysis Date"],
            y=chart_data["Quantity_Curr"],
            mode="lines",
            line={"color": "rgba(45,129,255,0.24)", "width": 0},
            fill="tozeroy",
            fillcolor="rgba(45,129,255,0.18)",
            name="Aktualna ilość area",
            hoverinfo="skip",
            showlegend=False,
        )
    )
    fig.add_trace(
        go.Scatter(
            x=chart_data["Analysis Date"],
            y=chart_data["Quantity_Prev"],
            mode="lines+markers",
            name="Poprzednia ilość",
            line={"color": "#8b949e", "width": 2.4},
            marker={"size": 6, "color": "#8b949e"},
            customdata=chart_data[["Quantity_Curr", "Delta"]].to_numpy(),
            hovertemplate=(
                "Data: %{x|%Y-%m-%d}<br>"
                "Poprzednia ilość: %{y:,.0f}<br>"
                "Aktualna ilość: %{customdata[0]:,.0f}<br>"
                "Bilans zmian: %{customdata[1]:+,.0f}<extra></extra>"
            ),
        )
    )
    fig.add_trace(
        go.Scatter(
            x=chart_data["Analysis Date"],
            y=chart_data["Quantity_Curr"],
            mode="lines+markers",
            name="Aktualna ilość",
            line={"color": "#2d81ff", "width": 3.4},
            marker={"size": 7, "color": "#f0f6fc", "line": {"color": "#2d81ff", "width": 2}},
            customdata=chart_data[["Quantity_Prev", "Delta"]].to_numpy(),
            hovertemplate=(
                "Data: %{x|%Y-%m-%d}<br>"
                "Aktualna ilość: %{y:,.0f}<br>"
                "Poprzednia ilość: %{customdata[0]:,.0f}<br>"
                "Bilans zmian: %{customdata[1]:+,.0f}<extra></extra>"
            ),
        )
    )
    if not latest_point.empty:
        latest_row = latest_point.iloc[0]
        fig.add_vline(
            x=latest_row["Analysis Date"],
            line_width=1,
            line_dash="dot",
            line_color="rgba(255,255,255,0.18)",
        )
        fig.add_annotation(
            x=latest_row["Analysis Date"],
            y=latest_row["Quantity_Curr"],
            text=f"Aktualnie {latest_row['Quantity_Curr']:,.0f}",
            showarrow=False,
            xanchor="left",
            yshift=-20,
            font={"color": "#f0f6fc", "size": 12},
        )
        fig.add_annotation(
            x=latest_row["Analysis Date"],
            y=latest_row["Quantity_Prev"],
            text=f"Poprzednio {latest_row['Quantity_Prev']:,.0f}",
            showarrow=False,
            xanchor="left",
            yshift=18,
            font={"color": "#8b949e", "size": 11},
        )
    fig.update_layout(
        height=420,
        hovermode="x unified",
        xaxis_title=x_title,
        yaxis_title="Ilość otwarta",
    )
    fig.update_xaxes(tickangle=-24)
    return fig


def build_delta_chart(date_summary, x_title):
    if date_summary.empty:
        return None

    chart_data = date_summary.sort_values("Analysis Date").copy()
    chart_data["Abs Delta"] = chart_data["Delta"].abs()
    label_source = chart_data.nlargest(min(6, len(chart_data)), "Abs Delta").copy()
    label_source["Delta Label"] = label_source["Delta"].map(lambda value: f"{value:+,.0f}")
    label_lookup = dict(zip(label_source["Analysis Date"], label_source["Delta Label"]))
    colors = [
        "#3fb950" if value > 0 else "#f85149" if value < 0 else "#8b949e"
        for value in chart_data["Delta"]
    ]
    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=chart_data["Analysis Date"],
            y=chart_data["Delta"],
            marker={"color": colors, "line": {"color": "rgba(255,255,255,0.05)", "width": 1}},
            name="Bilans zmian",
            customdata=chart_data[["Alerts"]].to_numpy() if "Alerts" in chart_data.columns else None,
            hovertemplate=(
                "Data: %{x|%Y-%m-%d}<br>"
                "Zmiana ilości: %{y:+,.0f}<br>"
                "Liczba alertów: %{customdata[0]}<extra></extra>"
                if "Alerts" in chart_data.columns
                else "Data: %{x|%Y-%m-%d}<br>Zmiana ilości: %{y:+,.0f}<extra></extra>"
            ),
            text=[label_lookup.get(value, "") for value in chart_data["Analysis Date"]],
            textposition="outside",
        )
    )
    fig.add_hline(y=0, line_width=1, line_color="rgba(255,255,255,0.14)")
    fig.update_layout(
        height=320,
        hovermode="x unified",
        xaxis_title=x_title,
        yaxis_title="Zmiana ilości",
    )
    fig.update_xaxes(tickangle=-24)
    return fig


def build_product_bar_source(product_summary, chart_type):
    if chart_type == "increase":
        return (
            product_summary[product_summary["Delta"] > 0]
            .nlargest(10, "Delta")[["Part Number", "Part Description", "Delta"]]
            .reset_index(drop=True)
        )

    return (
        product_summary[product_summary["Delta"] < 0]
        .nsmallest(10, "Delta")[["Part Number", "Part Description", "Delta"]]
        .reset_index(drop=True)
    )


def build_product_bar_chart(product_summary, chart_type):
    source = build_product_bar_source(product_summary, chart_type)
    if chart_type == "increase":
        color = "#3fb950"
        title = "Największe wzrosty"
    else:
        color = "#f85149"
        title = "Największe spadki"

    if source.empty:
        return None, title

    source["Display Label"] = source["Part Description"].map(
        lambda value: value if len(str(value)) <= 42 else f"{str(value)[:39]}..."
    )
    source["Delta Label"] = source["Delta"].map(lambda value: f"{value:+,.0f}")
    source = source.sort_values("Delta", ascending=True)
    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=source["Delta"],
            y=source["Display Label"],
            orientation="h",
            marker={"color": color, "line": {"color": "rgba(255,255,255,0.05)", "width": 1}},
            text=source["Delta Label"],
            textposition="outside",
            name=title,
            customdata=source[["Part Number", "Part Description"]].to_numpy(),
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "%{customdata[1]}<br>"
                "Zmiana ilości: %{x:+,.0f}<extra></extra>"
            ),
        )
    )
    fig.update_layout(
        height=max(340, len(source) * 34),
        xaxis_title="Zmiana ilości",
        yaxis_title=None,
        showlegend=False,
    )
    return fig, title


def build_product_waterfall_chart(product_summary):
    if product_summary.empty:
        return None

    source = (
        product_summary.assign(Abs_Delta=product_summary["Delta"].abs())
        .sort_values("Abs_Delta", ascending=False)
        .head(8)
        .copy()
    )
    if source.empty:
        return None

    source["Label"] = source["Part Number"].astype(str).str[:14]
    fig = go.Figure(
        go.Waterfall(
            x=source["Label"],
            y=source["Delta"],
            measure=["relative"] * len(source),
            connector={"line": {"color": "rgba(255,255,255,0.18)"}},
            increasing={"marker": {"color": "#3fb950"}},
            decreasing={"marker": {"color": "#f85149"}},
            totals={"marker": {"color": "#2d81ff"}},
            customdata=source[["Part Number", "Part Description"]].to_numpy(),
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "%{customdata[1]}<br>"
                "Delta: %{y:+,.0f}<extra></extra>"
            ),
        )
    )
    fig.update_layout(
        height=360,
        xaxis_title="Produkty",
        yaxis_title="Delta wolumenu",
        showlegend=False,
    )
    return fig


def build_change_mix_source(dataframe):
    mix = (
        dataframe.groupby("Change Direction", as_index=False)
        .agg(Rows=("Change Direction", "size"), Total_Delta=("Delta", "sum"))
    )
    mix["Direction Label"] = mix["Change Direction"].map(get_change_label)
    mix["Share"] = mix["Rows"] / max(int(mix["Rows"].sum()), 1)
    return mix


def build_change_mix_chart(dataframe):
    mix = build_change_mix_source(dataframe)
    if mix.empty:
        return None

    order = ["Wzrost", "Spadek", "Bez zmian"]
    color_map = {"Wzrost": "#3fb950", "Spadek": "#f85149", "Bez zmian": "#8b949e"}
    mix["SortOrder"] = mix["Direction Label"].map({label: index for index, label in enumerate(order)})
    mix = mix.sort_values("SortOrder", ascending=False)
    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=mix["Rows"],
            y=mix["Direction Label"],
            orientation="h",
            marker={"color": [color_map.get(label, "#8b949e") for label in mix["Direction Label"]]},
            text=mix["Share"].map(lambda value: f"{value:.1%}"),
            textposition="outside",
            customdata=mix[["Total_Delta"]].to_numpy(),
            hovertemplate=(
                "Kierunek: %{y}<br>"
                "Liczba pozycji: %{x}<br>"
                "Bilans zmian: %{customdata[0]:+,.0f}<extra></extra>"
            ),
            showlegend=False,
        )
    )
    fig.update_layout(height=240, xaxis_title="Liczba pozycji", yaxis_title=None)
    return fig


def build_key_findings(dataframe, product_summary, date_summary, date_basis):
    findings = []
    if dataframe.empty or product_summary.empty:
        return findings

    ranked = product_summary.assign(Abs_Delta=product_summary["Delta"].abs()).sort_values(
        "Abs_Delta", ascending=False
    )
    largest_move = ranked.iloc[0]
    findings.append(
        {
            "label": "Największa zmiana",
            "title": largest_move["Part Description"],
            "copy": (
                f"Najsilniejszy ruch w badanym oknie: {format_signed_int(largest_move['Delta'])} szt."
            ),
        }
    )

    negative = product_summary[product_summary["Delta"] < 0]
    if not negative.empty:
        largest_drop = negative.nsmallest(1, "Delta").iloc[0]
        findings.append(
            {
                "label": "Największy spadek",
                "title": largest_drop["Part Description"],
                "copy": (
                    f"Ta pozycja notuje najmocniejszy spadek: {format_signed_int(largest_drop['Delta'])} szt."
                ),
            }
        )

    alert_summary = (
        dataframe.groupby(["Part Description", "Part Number"], as_index=False)
        .agg(Alert_Count=("Alert", "sum"), Delta=("Delta", "sum"))
        .sort_values(["Alert_Count", "Delta"], ascending=[False, False])
    )
    if not alert_summary.empty and alert_summary.iloc[0]["Alert_Count"] > 0:
        top_alert = alert_summary.iloc[0]
        findings.append(
            {
                "label": "Najwięcej alertów",
                "title": top_alert["Part Description"],
                "copy": (
                    f"Produkt pojawia się w {int(top_alert['Alert_Count'])} alertach i wymaga szybkiej weryfikacji."
                ),
            }
        )

    if not date_summary.empty:
        peak_row = date_summary.sort_values(["Alerts", "Delta"], ascending=[False, False]).iloc[0]
        day_slice = dataframe[dataframe[date_basis] == peak_row["Analysis Date"]]
        if not day_slice.empty:
            day_product = (
                day_slice.groupby(["Part Description", "Part Number"], as_index=False)["Delta"]
                .sum()
                .assign(Abs_Delta=lambda df: df["Delta"].abs())
                .sort_values("Abs_Delta", ascending=False)
                .iloc[0]
            )
            findings.append(
                {
                    "label": "Kluczowy dzień",
                    "title": day_product["Part Description"],
                    "copy": (
                        f"Dnia {format_date(peak_row['Analysis Date'])} ta pozycja wygenerowała zmianę {format_signed_int(day_product['Delta'])} szt."
                    ),
                }
            )

    return findings[:4]


def build_matrix(dataframe, date_basis, metric_name):
    pivot_prev = pd.pivot_table(
        dataframe,
        index="Product Label",
        columns=date_basis,
        values="Quantity_Prev",
        aggfunc="sum",
        fill_value=0,
    )
    pivot_curr = pd.pivot_table(
        dataframe,
        index="Product Label",
        columns=date_basis,
        values="Quantity_Curr",
        aggfunc="sum",
        fill_value=0,
    )

    if metric_name == "Current Quantity":
        matrix = pivot_curr
    elif metric_name == "Previous Quantity":
        matrix = pivot_prev
    elif metric_name == "Delta":
        matrix = pivot_curr - pivot_prev
    else:
        matrix = ((pivot_curr - pivot_prev) / pivot_prev.replace(0, pd.NA)) * 100
        matrix = matrix.fillna(0)

    matrix = matrix.sort_index(axis=1)
    matrix.columns = [pd.Timestamp(column).strftime("%Y-%m-%d") for column in matrix.columns]
    return matrix


def blend_hex(start_hex, end_hex, ratio):
    ratio = max(0.0, min(1.0, float(ratio)))
    start = tuple(int(start_hex[index : index + 2], 16) for index in (1, 3, 5))
    end = tuple(int(end_hex[index : index + 2], 16) for index in (1, 3, 5))
    blended = tuple(
        round(start[channel] + (end[channel] - start[channel]) * ratio)
        for channel in range(3)
    )
    return "#{:02x}{:02x}{:02x}".format(*blended)


def style_value(value, metric_name, max_value, max_abs):
    base_style = "text-align: center;"

    if pd.isna(value):
        return base_style

    if metric_name == "Current Quantity":
        ratio = 0 if max_value <= 0 else value / max_value
        background = blend_hex("#f8fafc", "#2563eb", ratio)
        text_color = "#ffffff" if ratio > 0.6 else "#0f172a"
    elif metric_name == "Previous Quantity":
        ratio = 0 if max_value <= 0 else value / max_value
        background = blend_hex("#f8fafc", "#64748b", ratio)
        text_color = "#ffffff" if ratio > 0.6 else "#0f172a"
    else:
        ratio = 0 if max_abs <= 0 else abs(value) / max_abs
        if value > 0:
            background = blend_hex("#f0fdf4", "#16a34a", ratio)
            text_color = "#ffffff" if ratio > 0.6 else "#14532d"
        elif value < 0:
            background = blend_hex("#fef2f2", "#dc2626", ratio)
            text_color = "#ffffff" if ratio > 0.6 else "#7f1d1d"
        else:
            background = "#f8fafc"
            text_color = "#334155"

    return f"{base_style} background-color: {background}; color: {text_color};"


def style_matrix(matrix, metric_name):
    max_value = float(matrix.max().max()) if not matrix.empty else 0
    max_abs = float(matrix.abs().max().max()) if not matrix.empty else 0
    style_frame = matrix.map(
        lambda value: style_value(value, metric_name, max_value, max_abs)
    )

    styled = matrix.style.apply(lambda _: style_frame, axis=None)

    if metric_name in ["Current Quantity", "Previous Quantity", "Delta"]:
        return styled.format("{:,.0f}")

    return styled.format("{:+,.1f}%")


def style_excel_header(worksheet, row_number):
    fill = PatternFill(fill_type="solid", fgColor="0F172A")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for cell in worksheet[row_number]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border


def autosize_worksheet(worksheet, min_width=12, max_width=42):
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, min_width), max_width
        )


def insert_logo(worksheet, cell):
    if not logo_available():
        return

    logo = OpenpyxlImage(str(LOGO_PATH))
    logo.width = 160
    logo.height = 48
    worksheet.add_image(logo, cell)


def decorate_delta_column(worksheet, header_row=1):
    headers = {cell.value: cell.column for cell in worksheet[header_row]}
    delta_column = headers.get("Delta")
    percent_column = headers.get("Percent Change")
    if delta_column is None and percent_column is None:
        return

    green_fill = PatternFill(fill_type="solid", fgColor="DCFCE7")
    red_fill = PatternFill(fill_type="solid", fgColor="FEE2E2")
    blue_fill = PatternFill(fill_type="solid", fgColor="DBEAFE")

    for row in range(header_row + 1, worksheet.max_row + 1):
        if delta_column is not None:
            delta_cell = worksheet.cell(row=row, column=delta_column)
            if isinstance(delta_cell.value, (int, float)):
                delta_cell.fill = green_fill if delta_cell.value > 0 else red_fill if delta_cell.value < 0 else blue_fill
                delta_cell.font = Font(color="000000", bold=False)
        if percent_column is not None:
            percent_cell = worksheet.cell(row=row, column=percent_column)
            if isinstance(percent_cell.value, (int, float)):
                percent_cell.number_format = '0.0"%"'
                percent_cell.font = Font(color="000000", bold=False)


def excel_fill_color(value, metric_name, max_value, max_abs):
    if metric_name == "Current Quantity":
        ratio = 0 if max_value <= 0 else float(value) / max_value
        return blend_hex("#eff6ff", "#93c5fd", ratio)
    if metric_name == "Previous Quantity":
        ratio = 0 if max_value <= 0 else float(value) / max_value
        return blend_hex("#f8fafc", "#cbd5e1", ratio)
    ratio = 0 if max_abs <= 0 else abs(float(value)) / max_abs
    if value > 0:
        return blend_hex("#f0fdf4", "#86efac", ratio)
    if value < 0:
        return blend_hex("#fef2f2", "#fca5a5", ratio)
    return "#e2e8f0"


def ensure_numeric_cells_black(worksheet, start_row=1):
    for row in worksheet.iter_rows(min_row=start_row, max_row=worksheet.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.font = Font(color="000000", bold=bool(cell.font.bold))


def apply_polish_calendar_highlights(worksheet, date_columns, header_row=1):
    headers = {cell.value: cell.column for cell in worksheet[header_row]}
    saturday_fill = PatternFill(fill_type="solid", fgColor="DBEAFE")
    sunday_fill = PatternFill(fill_type="solid", fgColor="FEE2E2")
    holiday_fill = PatternFill(fill_type="solid", fgColor="FEF3C7")

    for column_name in date_columns:
        column_index = headers.get(column_name)
        if column_index is None:
            continue
        for row in range(header_row + 1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=column_index)
            if cell.value in (None, ""):
                continue
            try:
                day_info = classify_polish_day(cell.value)
            except Exception:
                continue
            if day_info["Is Holiday"]:
                cell.fill = holiday_fill
            elif day_info["Day Type"] == "Saturday":
                cell.fill = saturday_fill
            elif day_info["Day Type"] == "Sunday":
                cell.fill = sunday_fill
            cell.font = Font(color="000000", bold=bool(cell.font.bold))


def style_matrix_sheet(worksheet, metric_name, header_row=1, start_col=2):
    label_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    label_font = Font(color="0F172A", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    values = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        for col in range(start_col, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row, column=col).value
            if isinstance(cell_value, (int, float)):
                values.append(float(cell_value))

    max_value = max(values) if values else 0
    max_abs = max((abs(value) for value in values), default=0)

    for row in range(header_row + 1, worksheet.max_row + 1):
        label_cell = worksheet.cell(row=row, column=1)
        label_cell.fill = label_fill
        label_cell.font = label_font
        label_cell.border = thin_border
        for col in range(start_col, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(cell.value, (int, float)):
                bg = excel_fill_color(cell.value, metric_name, max_value, max_abs).replace("#", "")
                cell.fill = PatternFill(fill_type="solid", fgColor=bg)
                cell.font = Font(color="000000", bold=False)
                if metric_name == "Percent Change":
                    cell.number_format = '0.0"%"'
                else:
                    cell.number_format = '#,##0'


def highlight_calendar_rows(worksheet, header_row=1):
    date_column = None
    for cell in worksheet[header_row]:
        if cell.value == "Date":
            date_column = cell.column
            break
    if date_column is None:
        return

    saturday_fill = PatternFill(fill_type="solid", fgColor="DBEAFE")
    sunday_fill = PatternFill(fill_type="solid", fgColor="FEE2E2")
    holiday_fill = PatternFill(fill_type="solid", fgColor="FEF3C7")

    for row in range(header_row + 1, worksheet.max_row + 1):
        date_value = worksheet.cell(row=row, column=date_column).value
        if not date_value:
            continue
        try:
            day_info = classify_polish_day(date_value)
        except Exception:
            continue

        row_fill = None
        if day_info["Is Holiday"]:
            row_fill = holiday_fill
        elif day_info["Day Type"] == "Saturday":
            row_fill = saturday_fill
        elif day_info["Day Type"] == "Sunday":
            row_fill = sunday_fill

        if row_fill is None:
            continue

        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row, column=col).fill = row_fill


def highlight_weekly_rows(worksheet, header_row=1):
    headers = {cell.value: cell.column for cell in worksheet[header_row]}
    status_column = headers.get("Week Status")
    alert_column = headers.get("Any Weekly Alert")
    reference_column = headers.get("Is Reference Week")
    if status_column is None and alert_column is None and reference_column is None:
        return

    partial_fill = PatternFill(fill_type="solid", fgColor="FEF3C7")
    open_fill = PatternFill(fill_type="solid", fgColor="FCE7F3")
    reference_fill = PatternFill(fill_type="solid", fgColor="DBEAFE")
    alert_fill = PatternFill(fill_type="solid", fgColor="FDECEC")

    for row in range(header_row + 1, worksheet.max_row + 1):
        status_value = worksheet.cell(row=row, column=status_column).value if status_column else None
        is_alert = worksheet.cell(row=row, column=alert_column).value if alert_column else False
        is_reference = worksheet.cell(row=row, column=reference_column).value if reference_column else False

        row_fill = None
        if bool(is_reference):
            row_fill = reference_fill
        elif bool(is_alert):
            row_fill = alert_fill
        elif status_value == "Partial range":
            row_fill = partial_fill
        elif status_value == "Open week":
            row_fill = open_fill

        if row_fill is None:
            continue

        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row, column=col).fill = row_fill


def build_weekly_by_part_report(detail_df, date_basis):
    columns = [
        "Part Number",
        "Part Description",
        "Week Label",
        "Week Start",
        "Week End",
        "Previous Release Qty",
        "Current Release Qty",
        "Release Delta",
        "Release Change %",
        "Previous Week Qty",
    ]
    if detail_df is None or detail_df.empty or date_basis not in detail_df.columns:
        return pd.DataFrame(columns=columns)

    report = detail_df.copy()
    report["Analysis Date"] = pd.to_datetime(report[date_basis], errors="coerce")
    report = report[report["Analysis Date"].notna()].copy()
    if report.empty:
        return pd.DataFrame(columns=columns)

    iso = report["Analysis Date"].dt.isocalendar()
    report["ISO Year"] = iso.year.astype(int)
    report["ISO Week"] = iso.week.astype(int)
    report["Week Label"] = report["ISO Year"].astype(str) + "-W" + report["ISO Week"].astype(str).str.zfill(2)
    report["Week Start"] = (
        report["Analysis Date"] - pd.to_timedelta(report["Analysis Date"].dt.weekday, unit="D")
    ).dt.normalize()
    report["Week End"] = report["Week Start"] + pd.Timedelta(days=6)

    weekly = (
        report.groupby(
            ["Part Number", "Part Description", "Week Label", "Week Start", "Week End"],
            as_index=False,
        )
        .agg(
            **{
                "Previous Release Qty": ("Quantity_Prev", "sum"),
                "Current Release Qty": ("Quantity_Curr", "sum"),
                "Release Delta": ("Delta", "sum"),
            }
        )
        .sort_values(["Part Number", "Week Start", "Part Description"])
        .reset_index(drop=True)
    )
    weekly["Previous Week Qty"] = (
        weekly.groupby("Part Number")["Current Release Qty"].shift(1).fillna(0.0)
    )

    def _format_release_change(row):
        previous_qty = float(row["Previous Release Qty"])
        current_qty = float(row["Current Release Qty"])
        if previous_qty == 0:
            return "New" if current_qty != 0 else "0.0%"
        percent_value = ((current_qty - previous_qty) / previous_qty) * 100
        return f"{percent_value:+.1f}%"

    weekly["Release Change %"] = weekly.apply(_format_release_change, axis=1)
    return weekly[columns]


def build_weekly_by_part_chart_source(weekly_by_part_df):
    columns = [
        "Week Label",
        "Week Start",
        "Previous Release Qty",
        "Current Release Qty",
        "Release Delta",
        "Parts",
    ]
    if weekly_by_part_df is None or weekly_by_part_df.empty:
        return pd.DataFrame(columns=columns)

    source = (
        weekly_by_part_df.groupby(["Week Label", "Week Start"], as_index=False)
        .agg(
            **{
                "Previous Release Qty": ("Previous Release Qty", "sum"),
                "Current Release Qty": ("Current Release Qty", "sum"),
                "Release Delta": ("Release Delta", "sum"),
                "Parts": ("Part Number", "nunique"),
            }
        )
        .sort_values("Week Start")
        .reset_index(drop=True)
    )
    return source[columns]


def build_qty_matrix_report(weekly_by_part_df):
    if weekly_by_part_df is None or weekly_by_part_df.empty:
        return (
            pd.DataFrame(columns=["Part Number", "Part Description"]),
            pd.DataFrame(
                columns=[
                    "Week Label",
                    "Week Start",
                    "Week End",
                    "Parts",
                    "Previous Release Qty",
                    "Current Release Qty",
                    "Release Delta",
                ]
            ),
        )

    week_order = (
        weekly_by_part_df[["Week Label", "Week Start", "Week End"]]
        .drop_duplicates()
        .sort_values("Week Start")
        .reset_index(drop=True)
    )
    ordered_weeks = week_order["Week Label"].tolist()

    matrix = (
        weekly_by_part_df.pivot_table(
            index=["Part Number", "Part Description"],
            columns="Week Label",
            values="Current Release Qty",
            aggfunc="sum",
            fill_value=0,
        )
        .reindex(columns=ordered_weeks, fill_value=0)
        .reset_index()
    )

    weekly_totals = (
        weekly_by_part_df.groupby(["Week Label", "Week Start", "Week End"], as_index=False)
        .agg(
            **{
                "Parts": ("Part Number", "nunique"),
                "Previous Release Qty": ("Previous Release Qty", "sum"),
                "Current Release Qty": ("Current Release Qty", "sum"),
                "Release Delta": ("Release Delta", "sum"),
            }
        )
        .sort_values("Week Start")
        .reset_index(drop=True)
    )
    return matrix, weekly_totals


def write_parameter_section(worksheet, start_row, title, items, start_col=1):
    worksheet.cell(row=start_row, column=start_col, value=title)
    worksheet.cell(row=start_row, column=start_col).font = Font(size=13, bold=True, color="0F172A")
    for offset, (label, value) in enumerate(items, start=1):
        label_cell = worksheet.cell(row=start_row + offset, column=start_col, value=label)
        value_cell = worksheet.cell(row=start_row + offset, column=start_col + 1, value=value)
        label_cell.font = Font(bold=True, color="334155")
        label_cell.fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
        value_cell.font = Font(color="0F172A")
        label_cell.alignment = value_cell.alignment = Alignment(horizontal="left", vertical="center")


def write_dataframe_block(worksheet, dataframe, start_row, start_col=1):
    dataframe = dataframe.copy()
    if dataframe is None or dataframe.empty:
        return

    for col_offset, column_name in enumerate(dataframe.columns, start=0):
        worksheet.cell(row=start_row, column=start_col + col_offset, value=column_name)

    for row_offset, row in enumerate(dataframe.itertuples(index=False), start=1):
        for col_offset, value in enumerate(row, start=0):
            worksheet.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)


def style_table_region(worksheet, header_row, start_row=None, end_row=None):
    start_row = start_row or (header_row + 1)
    end_row = end_row or worksheet.max_row
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row != header_row:
                cell.alignment = Alignment(vertical="center", horizontal="left")


def apply_number_formats(worksheet, header_row, format_map):
    headers = {cell.value: cell.column for cell in worksheet[header_row]}
    for header_name, number_format in format_map.items():
        column_index = headers.get(header_name)
        if column_index is None:
            continue
        for row in range(header_row + 1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=column_index)
            if isinstance(cell.value, (int, float)):
                cell.number_format = number_format


def style_multi_label_matrix_sheet(worksheet, metric_name, header_row=1, start_col=3, label_columns=(1, 2)):
    label_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    label_font = Font(color="0F172A", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    values = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        for col in range(start_col, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row, column=col).value
            if isinstance(cell_value, (int, float)):
                values.append(float(cell_value))

    max_value = max(values) if values else 0
    max_abs = max((abs(value) for value in values), default=0)

    for row in range(header_row + 1, worksheet.max_row + 1):
        for label_col in label_columns:
            label_cell = worksheet.cell(row=row, column=label_col)
            label_cell.fill = label_fill
            label_cell.font = label_font
            label_cell.border = thin_border
        for col in range(start_col, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(cell.value, (int, float)):
                bg = excel_fill_color(cell.value, metric_name, max_value, max_abs).replace("#", "")
                cell.fill = PatternFill(fill_type="solid", fgColor=bg)
                cell.font = Font(color="000000", bold=False)
                cell.number_format = '#,##0'


def add_weekly_report_chart(worksheet, chart_source_df, start_row, start_col=12):
    if chart_source_df is None or chart_source_df.empty:
        return

    helper_col = start_col
    headers = ["Week Label", "Previous Release Qty", "Current Release Qty", "Release Delta"]
    for offset, header in enumerate(headers):
        worksheet.cell(row=2, column=helper_col + offset, value=header)
    for index, row in enumerate(chart_source_df.itertuples(index=False), start=3):
        worksheet.cell(row=index, column=helper_col, value=row[0])
        worksheet.cell(row=index, column=helper_col + 1, value=float(row[2]))
        worksheet.cell(row=index, column=helper_col + 2, value=float(row[3]))
        worksheet.cell(row=index, column=helper_col + 3, value=float(row[4]))

    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "Weekly Release Trend"
    bar_chart.y_axis.title = "Qty"
    bar_chart.x_axis.title = "Week"
    bar_chart.height = 8.5
    bar_chart.width = 18
    bar_chart.gapWidth = 55
    data = Reference(
        worksheet,
        min_col=helper_col + 1,
        max_col=helper_col + 2,
        min_row=2,
        max_row=2 + len(chart_source_df),
    )
    categories = Reference(
        worksheet,
        min_col=helper_col,
        min_row=3,
        max_row=2 + len(chart_source_df),
    )
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)

    line_chart = LineChart()
    line_chart.y_axis.title = "Delta"
    line_chart.y_axis.axId = 200
    line_chart.height = 8.5
    line_chart.width = 18
    delta_data = Reference(
        worksheet,
        min_col=helper_col + 3,
        max_col=helper_col + 3,
        min_row=2,
        max_row=2 + len(chart_source_df),
    )
    line_chart.add_data(delta_data, titles_from_data=True)
    line_chart.set_categories(categories)
    line_chart.y_axis.crosses = "max"

    bar_chart += line_chart
    worksheet.add_chart(bar_chart, f"{get_column_letter(start_col)}{start_row}")

    for col in range(helper_col, helper_col + len(headers)):
        worksheet.column_dimensions[get_column_letter(col)].hidden = True


def write_weekly_by_part_sheet(
    worksheet,
    weekly_by_part_df,
    chart_source_df,
    prev_meta,
    curr_meta,
    date_basis,
    selected_start_date,
    selected_end_date,
):
    insert_logo(worksheet, "J1")
    worksheet.merge_cells("A1:H1")
    worksheet["A1"] = "Weekly by Part Report"
    worksheet["A1"].font = Font(size=16, bold=True, color="0F172A")
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

    parameter_items = [
        ("Date Basis", get_date_label(date_basis)),
        ("Selected Period", f"{selected_start_date:%Y-%m-%d} to {selected_end_date:%Y-%m-%d}"),
        ("PO Number", curr_meta.get("po_number", "n/a")),
        ("Previous Release", format_release_summary(prev_meta)),
        ("Current Release", format_release_summary(curr_meta)),
        ("Planner", curr_meta.get("planner_name", "n/a")),
    ]
    write_parameter_section(worksheet, 3, "Parameters", parameter_items, start_col=1)

    worksheet["A11"] = "Weekly by Part"
    worksheet["A11"].font = Font(size=13, bold=True, color="0F172A")
    table_start_row = 12
    weekly_export = weekly_by_part_df.copy()
    if not weekly_export.empty:
        weekly_export["Week Start"] = pd.to_datetime(weekly_export["Week Start"]).dt.strftime("%Y-%m-%d")
        weekly_export["Week End"] = pd.to_datetime(weekly_export["Week End"]).dt.strftime("%Y-%m-%d")
    write_dataframe_block(worksheet, weekly_export, table_start_row, start_col=1)
    style_excel_header(worksheet, table_start_row)
    style_table_region(worksheet, table_start_row, start_row=table_start_row + 1)
    decorate_delta_column(worksheet, header_row=table_start_row)
    apply_number_formats(
        worksheet,
        table_start_row,
        {
            "Previous Release Qty": '#,##0',
            "Current Release Qty": '#,##0',
            "Release Delta": '+#,##0;-#,##0;0',
            "Previous Week Qty": '#,##0',
        },
    )
    autosize_worksheet(worksheet)
    ensure_numeric_cells_black(worksheet, start_row=table_start_row + 1)
    worksheet.freeze_panes = f"A{table_start_row + 1}"
    add_weekly_report_chart(worksheet, chart_source_df, start_row=3, start_col=10)


def write_qty_matrix_sheet(
    worksheet,
    qty_matrix_df,
    weekly_totals_df,
    curr_meta,
    date_basis,
    selected_start_date,
    selected_end_date,
):
    insert_logo(worksheet, "I1")
    worksheet.merge_cells("A1:G1")
    worksheet["A1"] = "Qty Matrix"
    worksheet["A1"].font = Font(size=16, bold=True, color="0F172A")
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

    parameter_items = [
        ("Date Basis", get_date_label(date_basis)),
        ("Selected Period", f"{selected_start_date:%Y-%m-%d} to {selected_end_date:%Y-%m-%d}"),
        ("PO Number", curr_meta.get("po_number", "n/a")),
        ("Planner", curr_meta.get("planner_name", "n/a")),
    ]
    write_parameter_section(worksheet, 3, "Parameters", parameter_items, start_col=1)

    worksheet["A9"] = "Weekly Aggregates"
    worksheet["A9"].font = Font(size=13, bold=True, color="0F172A")
    helper_start_row = 10
    helper_export = weekly_totals_df.copy()
    if not helper_export.empty:
        helper_export["Week Start"] = pd.to_datetime(helper_export["Week Start"]).dt.strftime("%Y-%m-%d")
        helper_export["Week End"] = pd.to_datetime(helper_export["Week End"]).dt.strftime("%Y-%m-%d")
    write_dataframe_block(worksheet, helper_export, helper_start_row, start_col=1)
    style_excel_header(worksheet, helper_start_row)
    style_table_region(worksheet, helper_start_row, start_row=helper_start_row + 1, end_row=helper_start_row + len(helper_export))
    apply_number_formats(
        worksheet,
        helper_start_row,
        {
            "Previous Release Qty": '#,##0',
            "Current Release Qty": '#,##0',
            "Release Delta": '+#,##0;-#,##0;0',
        },
    )
    ensure_numeric_cells_black(worksheet, start_row=helper_start_row + 1)

    matrix_start_row = helper_start_row + len(helper_export) + 4
    worksheet.cell(row=matrix_start_row - 1, column=1, value="Weekly Matrix")
    worksheet.cell(row=matrix_start_row - 1, column=1).font = Font(size=13, bold=True, color="0F172A")
    write_dataframe_block(worksheet, qty_matrix_df, matrix_start_row, start_col=1)
    style_excel_header(worksheet, matrix_start_row)
    style_multi_label_matrix_sheet(worksheet, "Current Quantity", header_row=matrix_start_row, start_col=3, label_columns=(1, 2))
    autosize_worksheet(worksheet)
    ensure_numeric_cells_black(worksheet, start_row=matrix_start_row + 1)
    worksheet.freeze_panes = f"C{matrix_start_row + 1}"


def to_professional_weekly_report_bytes(
    detail_df,
    prev_meta,
    curr_meta,
    date_basis,
    selected_start_date,
    selected_end_date,
):
    weekly_by_part_df = build_weekly_by_part_report(detail_df, date_basis)
    chart_source_df = build_weekly_by_part_chart_source(weekly_by_part_df)
    qty_matrix_df, weekly_totals_df = build_qty_matrix_report(weekly_by_part_df)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="Weekly by Part", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Qty Matrix", index=False)

        weekly_sheet = writer.book["Weekly by Part"]
        write_weekly_by_part_sheet(
            weekly_sheet,
            weekly_by_part_df,
            chart_source_df,
            prev_meta,
            curr_meta,
            date_basis,
            selected_start_date,
            selected_end_date,
        )

        qty_matrix_sheet = writer.book["Qty Matrix"]
        write_qty_matrix_sheet(
            qty_matrix_sheet,
            qty_matrix_df,
            weekly_totals_df,
            curr_meta,
            date_basis,
            selected_start_date,
            selected_end_date,
        )

    return output.getvalue()


def write_summary_sheet(
    worksheet,
    prev_meta,
    curr_meta,
    detail_df,
    product_summary,
    weekly_summary,
    date_basis,
    selected_start_date,
    selected_end_date,
    key_findings,
):
    insert_logo(worksheet, "G1")
    worksheet.merge_cells("A1:F1")
    worksheet["A1"] = BRAND_NAME
    worksheet["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    worksheet["A1"].fill = PatternFill(fill_type="solid", fgColor="0F172A")
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    worksheet.merge_cells("A2:F2")
    worksheet["A2"] = "Release Change Executive Summary"
    worksheet["A2"].font = Font(size=16, bold=True, color="0F172A")
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center")

    worksheet["A4"] = "PO Number"
    worksheet["B4"] = curr_meta["po_number"]
    worksheet["A5"] = "Previous Release"
    worksheet["B5"] = format_release_summary(prev_meta)
    worksheet["A6"] = "Current Release"
    worksheet["B6"] = format_release_summary(curr_meta)
    worksheet["A7"] = "Date Basis"
    worksheet["B7"] = date_basis
    worksheet["A8"] = "Selected Period"
    worksheet["B8"] = f"{selected_start_date:%Y-%m-%d} to {selected_end_date:%Y-%m-%d}"
    worksheet["A9"] = "Planner"
    worksheet["B9"] = curr_meta["planner_name"]
    worksheet["A10"] = "Planner Email"
    worksheet["B10"] = curr_meta["planner_email"]

    total_prev = detail_df["Quantity_Prev"].sum()
    total_curr = detail_df["Quantity_Curr"].sum()
    total_delta = detail_df["Delta"].sum()
    alert_count = int(detail_df["Alert"].sum())
    products_changed = int((product_summary["Delta"] != 0).sum())

    worksheet["D4"] = "Previous Qty"
    worksheet["E4"] = total_prev
    worksheet["D5"] = "Current Qty"
    worksheet["E5"] = total_curr
    worksheet["D6"] = "Net Delta"
    worksheet["E6"] = total_delta
    worksheet["D7"] = "Alert Rows"
    worksheet["E7"] = alert_count
    worksheet["D8"] = "Products Changed"
    worksheet["E8"] = products_changed

    reference_week = get_last_completed_reference_week(selected_end_date)
    reference_row, previous_row = get_reference_week_rows(weekly_summary)
    worksheet["D10"] = "Reference Week"
    worksheet["E10"] = reference_row["Week Label"] if reference_row is not None else reference_week.week_label
    worksheet["D11"] = "Reference Current Qty"
    worksheet["E11"] = float(reference_row["Quantity_Curr"]) if reference_row is not None else 0
    worksheet["D12"] = "Reference Release %"
    worksheet["E12"] = (
        format_percent_display(reference_row["Release Percent Label"])
        if reference_row is not None
        else "n/a"
    )
    worksheet["D13"] = "Reference WoW %"
    worksheet["E13"] = (
        format_percent_display(reference_row["WoW Percent Label"])
        if reference_row is not None
        else "n/a"
    )
    worksheet["D14"] = "Working Days PL"
    worksheet["E14"] = int(reference_row["Working_Days_PL"]) if reference_row is not None else 0
    worksheet["D15"] = "Previous Closed Week"
    worksheet["E15"] = previous_row["Week Label"] if previous_row is not None else "n/a"

    worksheet["A13"] = "Key Findings"
    worksheet["A13"].font = Font(size=13, bold=True, color="0F172A")
    for idx, finding in enumerate(key_findings[:4], start=14):
        worksheet[f"A{idx}"] = finding["label"]
        worksheet[f"A{idx}"].font = Font(bold=True, color="2563EB")
        worksheet[f"B{idx}"] = finding["title"]
        worksheet[f"C{idx}"] = finding["copy"]

    worksheet["A20"] = "Top Product Changes"
    worksheet["A20"].font = Font(size=13, bold=True, color="0F172A")
    worksheet["A21"] = "Part Number"
    worksheet["B21"] = "Part Description"
    worksheet["C21"] = "Previous Qty"
    worksheet["D21"] = "Current Qty"
    worksheet["E21"] = "Delta"
    worksheet["F21"] = "Alert Count"
    style_excel_header(worksheet, 21)

    top_rows = (
        product_summary.assign(Abs_Delta=product_summary["Delta"].abs())
        .sort_values("Abs_Delta", ascending=False)
        .drop(columns=["Abs_Delta", "Product Label", "Change Direction"])
        .head(10)
    )
    start_row = 22
    for offset, (_, row) in enumerate(top_rows.iterrows()):
        excel_row = start_row + offset
        worksheet.cell(excel_row, 1, row["Part Number"])
        worksheet.cell(excel_row, 2, row["Part Description"])
        worksheet.cell(excel_row, 3, row["Quantity_Prev"])
        worksheet.cell(excel_row, 4, row["Quantity_Curr"])
        worksheet.cell(excel_row, 5, row["Delta"])
        worksheet.cell(excel_row, 6, int(row["Alert_Count"]))

    decorate_delta_column(worksheet, header_row=21)
    autosize_worksheet(worksheet)


def build_weekly_comparison_export(weekly_summary):
    if weekly_summary is None or weekly_summary.empty:
        return pd.DataFrame(
            columns=[
                "Week Label",
                "Week Start",
                "Week End",
                "Working Days PL",
                "Polish Holidays",
                "Weekend Days",
                "Previous Release Qty",
                "Current Release Qty",
                "Release Delta",
                "Release Change %",
                "Previous Week Qty",
                "WoW Delta",
                "WoW Change %",
                "Trend",
                "Holiday Week",
                "Week Status",
            ]
        )

    export = weekly_summary[
        [
            "Week Label",
            "Week Start",
            "Week End",
            "Working_Days_PL",
            "Holidays_PL",
            "Weekend_Days",
            "Quantity_Prev",
            "Quantity_Curr",
            "Delta",
            "Release Percent Label",
            "Previous Week Current Qty",
            "WoW Delta",
            "WoW Percent Label",
            "Week Status",
        ]
    ].copy()
    export["Trend"] = export["WoW Delta"].map(
        lambda value: "Wzrost" if float(value or 0) > 0 else "Spadek" if float(value or 0) < 0 else "Stabilnie"
    )
    export["Holiday Week"] = export["Holidays_PL"].map(lambda value: "Tak" if int(value or 0) > 0 else "")
    export["Week Start"] = pd.to_datetime(export["Week Start"], errors="coerce").dt.strftime("%Y-%m-%d")
    export["Week End"] = pd.to_datetime(export["Week End"], errors="coerce").dt.strftime("%Y-%m-%d")
    export["Release Percent Label"] = export["Release Percent Label"].map(format_percent_display)
    export["WoW Percent Label"] = export["WoW Percent Label"].map(format_percent_display)
    export = export.rename(
        columns={
            "Working_Days_PL": "Working Days PL",
            "Holidays_PL": "Polish Holidays",
            "Weekend_Days": "Weekend Days",
            "Quantity_Prev": "Previous Release Qty",
            "Quantity_Curr": "Current Release Qty",
            "Delta": "Release Delta",
            "Release Percent Label": "Release Change %",
            "Previous Week Current Qty": "Previous Week Qty",
            "WoW Percent Label": "WoW Change %",
        }
    )
    return export


def build_report_matrix_export(matrix_df):
    if matrix_df is None or matrix_df.empty:
        return pd.DataFrame(columns=["Part Number", "Part Description"])

    export = matrix_df.reset_index()
    first_column = export.columns[0]
    export = export.rename(columns={first_column: "Product Label"})
    label_parts = export["Product Label"].astype(str).str.split(" | ", n=1, regex=False, expand=True)
    part_numbers = label_parts[0].fillna(export["Product Label"])
    part_descriptions = label_parts[1].fillna("") if 1 in label_parts.columns else ""
    export.insert(0, "Part Number", part_numbers)
    export.insert(1, "Part Description", part_descriptions)
    export = export.drop(columns=["Product Label"])
    return export


def build_matrix_totals_export(matrix_export_df, metric_label):
    if matrix_export_df is None or matrix_export_df.empty:
        return pd.DataFrame(columns=["Date", metric_label])

    value_columns = [
        column for column in matrix_export_df.columns if column not in {"Part Number", "Part Description"}
    ]
    totals = [
        {
            "Date": column,
            metric_label: float(pd.to_numeric(matrix_export_df[column], errors="coerce").fillna(0).sum()),
        }
        for column in value_columns
    ]
    return pd.DataFrame(totals)


def build_calendar_operational_export(selected_start_date, selected_end_date):
    calendar_export = build_calendar_frame(selected_start_date, selected_end_date).copy()
    if calendar_export.empty:
        return pd.DataFrame()

    calendar_export["Date"] = pd.to_datetime(calendar_export["Date"]).dt.strftime("%Y-%m-%d")
    calendar_export["Week Start"] = pd.to_datetime(calendar_export["Week Start"]).dt.strftime("%Y-%m-%d")
    calendar_export["Week End"] = pd.to_datetime(calendar_export["Week End"]).dt.strftime("%Y-%m-%d")
    return calendar_export[
        [
            "Date",
            "Day Name",
            "Day Type",
            "Holiday Name",
            "ISO Week Label",
            "Week Start",
            "Week End",
            "Is Working Day",
            "Is Weekend",
            "Is Holiday",
        ]
    ]


def build_calendar_weekly_export(calendar_export):
    if calendar_export is None or calendar_export.empty:
        return pd.DataFrame(
            columns=["ISO Week Label", "Working Days PL", "Saturdays", "Sundays", "Polish Holidays"]
        )

    weekly = (
        calendar_export.groupby(["ISO Week Label", "Week Start", "Week End"], as_index=False)
        .agg(
            **{
                "Working Days PL": ("Is Working Day", "sum"),
                "Saturdays": ("Day Type", lambda values: int((pd.Series(values) == "Saturday").sum())),
                "Sundays": ("Day Type", lambda values: int((pd.Series(values) == "Sunday").sum())),
                "Polish Holidays": ("Is Holiday", "sum"),
            }
        )
        .sort_values(["Week Start", "Week End"])
        .reset_index(drop=True)
    )
    return weekly


def add_weekly_comparison_chart(worksheet, header_row=1, anchor_cell="M2"):
    headers = {cell.value: cell.column for cell in worksheet[header_row]}
    week_column = headers.get("Week Label")
    previous_column = headers.get("Previous Release Qty")
    current_column = headers.get("Current Release Qty")
    if week_column is None or previous_column is None or current_column is None or worksheet.max_row <= header_row:
        return

    chart = LineChart()
    chart.title = "Weekly quantity comparison"
    chart.y_axis.title = "Qty"
    chart.x_axis.title = "Week"
    chart.height = 8.5
    chart.width = 18
    data = Reference(
        worksheet,
        min_col=previous_column,
        max_col=current_column,
        min_row=header_row,
        max_row=worksheet.max_row,
    )
    categories = Reference(
        worksheet,
        min_col=week_column,
        min_row=header_row + 1,
        max_row=worksheet.max_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    worksheet.add_chart(chart, anchor_cell)


def add_totals_chart(worksheet, totals_df, *, value_label, title, anchor_cell="L2"):
    if totals_df is None or totals_df.empty:
        return

    helper_col = worksheet.max_column + 3
    worksheet.cell(row=2, column=helper_col, value="Date")
    worksheet.cell(row=2, column=helper_col + 1, value=value_label)
    for index, row in enumerate(totals_df.itertuples(index=False), start=3):
        worksheet.cell(row=index, column=helper_col, value=row[0])
        worksheet.cell(row=index, column=helper_col + 1, value=float(row[1]))

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = value_label
    chart.x_axis.title = "Date"
    chart.height = 8.5
    chart.width = 18
    data = Reference(
        worksheet,
        min_col=helper_col + 1,
        max_col=helper_col + 1,
        min_row=2,
        max_row=2 + len(totals_df),
    )
    categories = Reference(
        worksheet,
        min_col=helper_col,
        min_row=3,
        max_row=2 + len(totals_df),
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    worksheet.add_chart(chart, anchor_cell)

    worksheet.column_dimensions[get_column_letter(helper_col)].hidden = True
    worksheet.column_dimensions[get_column_letter(helper_col + 1)].hidden = True


def add_calendar_summary_chart(worksheet, calendar_weekly_df, anchor_cell="L2"):
    if calendar_weekly_df is None or calendar_weekly_df.empty:
        return

    helper_col = worksheet.max_column + 3
    headers = ["ISO Week Label", "Working Days PL", "Saturdays", "Polish Holidays"]
    for offset, header in enumerate(headers):
        worksheet.cell(row=2, column=helper_col + offset, value=header)
    for index, row in enumerate(calendar_weekly_df.itertuples(index=False), start=3):
        worksheet.cell(row=index, column=helper_col, value=row[0])
        worksheet.cell(row=index, column=helper_col + 1, value=int(row[3]))
        worksheet.cell(row=index, column=helper_col + 2, value=int(row[4]))
        worksheet.cell(row=index, column=helper_col + 3, value=int(row[6]))

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Operational calendar summary"
    chart.y_axis.title = "Days"
    chart.x_axis.title = "ISO Week"
    chart.height = 8.5
    chart.width = 18
    data = Reference(
        worksheet,
        min_col=helper_col + 1,
        max_col=helper_col + 3,
        min_row=2,
        max_row=2 + len(calendar_weekly_df),
    )
    categories = Reference(
        worksheet,
        min_col=helper_col,
        min_row=3,
        max_row=2 + len(calendar_weekly_df),
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    worksheet.add_chart(chart, anchor_cell)

    for offset in range(len(headers)):
        worksheet.column_dimensions[get_column_letter(helper_col + offset)].hidden = True


def to_excel_bytes(
    detail_df,
    weekly_summary,
    current_matrix_df,
    delta_matrix_df,
    prev_meta,
    curr_meta,
    product_summary,
    date_basis,
    selected_start_date,
    selected_end_date,
    key_findings,
):
    output = io.BytesIO()
    weekly_export = build_weekly_comparison_export(weekly_summary)
    calendar_export = build_calendar_operational_export(selected_start_date, selected_end_date)
    calendar_weekly_export = build_calendar_weekly_export(calendar_export)
    current_matrix_export = build_report_matrix_export(current_matrix_df)
    delta_matrix_export = build_report_matrix_export(delta_matrix_df)
    current_totals_export = build_matrix_totals_export(current_matrix_export, "Current Release Qty")
    delta_totals_export = build_matrix_totals_export(delta_matrix_export, "Release Delta")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        weekly_export.to_excel(writer, sheet_name="Weekly Comparison", index=False)
        calendar_export.to_excel(writer, sheet_name="Calendar PL", index=False)
        current_matrix_export.to_excel(writer, sheet_name="Current Matrix", index=False)
        delta_matrix_export.to_excel(writer, sheet_name="Delta Heatmap", index=False)

        weekly_sheet = writer.book["Weekly Comparison"]
        style_excel_header(weekly_sheet, 1)
        style_table_region(weekly_sheet, 1)
        apply_number_formats(
            weekly_sheet,
            1,
            {
                "Previous Release Qty": '#,##0',
                "Current Release Qty": '#,##0',
                "Release Delta": '#,##0',
                "Previous Week Qty": '#,##0',
                "WoW Delta": '#,##0',
            },
        )
        weekly_sheet.freeze_panes = "A2"
        weekly_sheet.auto_filter.ref = weekly_sheet.dimensions
        autosize_worksheet(weekly_sheet)
        ensure_numeric_cells_black(weekly_sheet, start_row=2)
        add_weekly_comparison_chart(weekly_sheet)

        calendar_sheet = writer.book["Calendar PL"]
        style_excel_header(calendar_sheet, 1)
        style_table_region(calendar_sheet, 1)
        highlight_calendar_rows(calendar_sheet, header_row=1)
        apply_polish_calendar_highlights(calendar_sheet, ["Date"], header_row=1)
        calendar_sheet.freeze_panes = "A2"
        calendar_sheet.auto_filter.ref = calendar_sheet.dimensions
        autosize_worksheet(calendar_sheet)
        ensure_numeric_cells_black(calendar_sheet, start_row=2)
        add_calendar_summary_chart(calendar_sheet, calendar_weekly_export)

        current_matrix_sheet = writer.book["Current Matrix"]
        style_excel_header(current_matrix_sheet, 1)
        current_matrix_sheet.freeze_panes = "C2"
        current_matrix_sheet.auto_filter.ref = current_matrix_sheet.dimensions
        autosize_worksheet(current_matrix_sheet)
        style_multi_label_matrix_sheet(current_matrix_sheet, "Current Quantity", header_row=1, start_col=3, label_columns=(1, 2))
        ensure_numeric_cells_black(current_matrix_sheet, start_row=2)
        add_totals_chart(
            current_matrix_sheet,
            current_totals_export,
            value_label="Current Release Qty",
            title="Current quantity by date",
        )

        delta_heatmap_sheet = writer.book["Delta Heatmap"]
        style_excel_header(delta_heatmap_sheet, 1)
        delta_heatmap_sheet.freeze_panes = "C2"
        delta_heatmap_sheet.auto_filter.ref = delta_heatmap_sheet.dimensions
        autosize_worksheet(delta_heatmap_sheet)
        style_multi_label_matrix_sheet(delta_heatmap_sheet, "Delta", header_row=1, start_col=3, label_columns=(1, 2))
        ensure_numeric_cells_black(delta_heatmap_sheet, start_row=2)
        add_totals_chart(
            delta_heatmap_sheet,
            delta_totals_export,
            value_label="Release Delta",
            title="Delta by date",
        )

    return output.getvalue()


def render_workspace_actions():
    auth_user = get_auth_user()
    info_col, clear_col, logout_col = st.columns([0.72, 0.14, 0.14], gap="small")
    with info_col:
        display_name = auth_user.get("display_name", "User")
        role_name = auth_user.get("role", "Viewer")
        st.caption(f"Aktywna sesja: {display_name} | {role_name}")
    with clear_col:
        clear_disabled = not workspace_has_uploads()
        if st.button("Wyczysc pliki", key="workspace_clear_files", use_container_width=True, disabled=clear_disabled):
            clear_workspace_uploads()
            st.rerun()
    with logout_col:
        if st.button("Wyloguj", key="workspace_logout_button", use_container_width=True):
            logout_user()
            st.rerun()


def render_workspace_upload_panel():
    ui_shell.render_panel_intro(
        "Analiza plikow",
        "Upload i status plikow",
        "To jedyne miejsce, w ktorym dodajesz lub podmieniasz pliki. Po zaladowaniu pozostaja dostepne w pozostalych widokach.",
    )
    upload_left, upload_right = st.columns(2, gap="large")
    with upload_left:
        render_upload_card(
            "Poprzedni",
            "Baseline release",
            "Plik referencyjny do porownania aktualnego stanu planu i wysylek.",
        )
        previous_upload = st.file_uploader(
            "Upload Previous Release",
            type=["xlsx"],
            key=get_upload_widget_key("previous"),
            label_visibility="visible",
        )
        if previous_upload is not None:
            store_uploaded_release("previous", previous_upload)
        if get_stored_upload("previous") is not None:
            st.caption(f"Zaladowany plik: {get_stored_upload('previous')['name']}")
            if st.button("Usun poprzedni plik", key="clear_previous_upload", use_container_width=True):
                clear_uploaded_release("previous")
                st.rerun()

    with upload_right:
        render_upload_card(
            "Aktualny",
            "Current release",
            "Plik, z ktorego aplikacja liczy zmiany, alerty i aktualny wolumen.",
        )
        current_upload = st.file_uploader(
            "Upload Current Release",
            type=["xlsx"],
            key=get_upload_widget_key("current"),
            label_visibility="visible",
        )
        if current_upload is not None:
            store_uploaded_release("current", current_upload)
        if get_stored_upload("current") is not None:
            st.caption(f"Zaladowany plik: {get_stored_upload('current')['name']}")
            if st.button("Usun aktualny plik", key="clear_current_upload", use_container_width=True):
                clear_uploaded_release("current")
                st.rerun()

    return get_stored_upload("previous"), get_stored_upload("current")


def analyze_uploaded_releases():
    previous_release = get_stored_upload("previous")
    current_release = get_stored_upload("current")
    if previous_release is None or current_release is None:
        clear_analysis_state()
        return None

    signature = current_upload_pair_signature()
    cached_bundle = st.session_state.get("analysis_results")
    if cached_bundle is not None and st.session_state.get("analysis_signature") == signature:
        return cached_bundle

    prev_df, prev_meta = load_release(previous_release["bytes"], previous_release["name"])
    curr_df, curr_meta = load_release(current_release["bytes"], current_release["name"])
    if prev_df is None or curr_df is None:
        raise ValueError("Parser nie zwrocil danych dla jednego z plikow.")
    if prev_df.empty or curr_df.empty:
        raise ValueError("Jeden z zaladowanych plikow nie zawiera danych do porownania.")

    result = compare_releases(prev_df, curr_df)
    if result is None:
        raise ValueError("Analiza nie zwrocila wyniku.")

    bundle = {
        "prev_meta": prev_meta,
        "curr_meta": curr_meta,
        "result": result,
        "brand_context": detect_brand_context(prev_meta, curr_meta),
    }
    st.session_state["file_type"] = {
        "previous": prev_meta.get("file_type"),
        "current": curr_meta.get("file_type"),
    }
    st.session_state["parsed_data"] = {
        "previous": prev_meta,
        "current": curr_meta,
    }
    st.session_state["transformed_data"] = {
        "previous": prev_df,
        "current": curr_df,
    }
    st.session_state["analysis_results"] = bundle
    st.session_state["analysis_signature"] = signature
    st.session_state["error_message"] = ""
    st.session_state["success_status"] = "Dane zostaly rozpoznane, sparsowane i przeanalizowane."
    set_pipeline_debug(
        "analysis ready "
        f"(prev_rows={len(prev_df)}, curr_rows={len(curr_df)}, result_rows={len(result)})"
    )
    return bundle


def render_global_filter_drawer(result):
    with st.expander("Filtry", expanded=st.session_state.get("filters_expanded", False)):
        ui_shell.render_panel_intro(
            "Filtry",
            "Zakres i kontekst analizy",
            "Filtry sa ukryte do momentu otwarcia panelu. Zmieniaja wspolny kontekst dla dashboardu, analizy plikow, plannera i wykresow.",
        )
        _, close_col = st.columns([0.78, 0.22], gap="small")
        with close_col:
            if st.button("Zwin panel", key="close_filter_drawer", use_container_width=True):
                close_filters_panel()
                st.rerun()
        return render_filter_controls(result)


def build_default_filter_state():
    return {
        "date_basis": DATE_OPTIONS[0],
        "selected_start_date": None,
        "selected_end_date": None,
        "week_from": None,
        "week_to": None,
        "selected_products": [],
        "search_term": "",
        "selected_change_directions": ["Increase", "Decrease", "No Change"],
        "only_alerts": False,
        "full_product_summary": pd.DataFrame(),
    }


def render_view_shell(active_view, logo_markup):
    view_copy = {
        "dashboard": (
            "Dashboard",
            "Osobny widok dla KPI, alertow i kluczowych danych podsumowujacych.",
        ),
        "files": (
            "Analiza plikow",
            "Workspace uploadu, statusu plikow, plannera, eksportu oraz danych szczegolowych.",
        ),
        "charts": (
            "Wykresy",
            "Osobny widok dla raportow i wizualizacji, bez przeciazania strony glownej.",
        ),
    }
    title, copy = view_copy.get(active_view, ("Aplikacja", "Workspace analityczny."))
    ui_shell.render_workspace_shell(logo_markup, APP_TITLE, title, copy)
    action_cols = st.columns([0.2, 0.2, 0.6], gap="small")
    with action_cols[0]:
        if st.button("Strona glowna", key=f"view_home_{active_view}", use_container_width=True):
            set_active_view("home")
            st.rerun()
    with action_cols[1]:
        filter_label = "Ukryj filtry" if st.session_state.get("filters_expanded", False) else "Pokaz filtry"
        if st.button(filter_label, key=f"view_filters_{active_view}", use_container_width=True):
            if st.session_state.get("filters_expanded", False):
                close_filters_panel()
            else:
                open_filters_panel()
            st.rerun()


def format_workspace_date_range(filter_state):
    if not isinstance(filter_state, dict):
        return "Zakres nieustawiony"

    start_date = filter_state.get("selected_start_date")
    end_date = filter_state.get("selected_end_date")
    if start_date is None or end_date is None:
        return "Zakres nieustawiony"

    try:
        start_label = pd.Timestamp(start_date).strftime("%Y-%m-%d")
        end_label = pd.Timestamp(end_date).strftime("%Y-%m-%d")
    except Exception:
        return "Zakres nieustawiony"
    return f"{start_label} - {end_label}"


def format_workspace_week_range(filter_state):
    if not isinstance(filter_state, dict):
        return "Tygodnie nieustawione"
    week_from = filter_state.get("week_from")
    week_to = filter_state.get("week_to")
    if not week_from or not week_to:
        return "Tygodnie nieustawione"
    return f"{week_from} - {week_to}"


def build_workspace_context_cards(prev_meta, curr_meta, filter_state, filtered_df):
    items = [
        {"label": "Format", "value": describe_format_context(prev_meta, curr_meta)},
        {"label": "Numer PO", "value": curr_meta.get("po_number", "n/a")},
        {"label": "Planista", "value": curr_meta.get("planner_name", "n/a")},
        {
            "label": "Zakres",
            "value": format_workspace_date_range(filter_state),
        },
        {
            "label": "Tygodnie",
            "value": format_workspace_week_range(filter_state),
        },
        {"label": "Wiersze po filtrach", "value": f"{len(filtered_df):,}"},
    ]
    ui_shell.render_context_cards(items)


def render_empty_analysis_prompt(title, copy):
    ui_shell.render_panel_intro("Workspace", title, copy)
    if st.button("Przejdz do Analiza plikow", key=f"empty_prompt_{title}", use_container_width=False):
        set_active_view("files")
        st.rerun()


def render_module_content(module_name, module_data, ui):
    module_renderers = {
        "dashboard": render_dashboard_module,
        "planner": render_planner_module,
        "reports": render_reports_module,
        "details": render_details_module,
        "admin": render_admin_module,
    }
    module_renderer = module_renderers.get(module_name, render_dashboard_module)
    module_renderer(module_data, ui)


def render_file_analysis_workspace(
    module_data,
    ui,
    filtered_df,
    product_summary,
    prev_meta,
    curr_meta,
    filter_state,
    excel_bytes,
    csv_bytes,
    professional_excel_bytes,
):
    previous_release = get_stored_upload("previous")
    current_release = get_stored_upload("current")
    render_workspace_upload_panel()
    render_file_slot_cards(
        prev_file=None if prev_meta else previous_release,
        current_file=None if curr_meta else current_release,
        prev_meta=prev_meta,
        curr_meta=curr_meta,
    )

    available_sections = ["overview", "planner", "details"]
    if can_access_module("admin", auth_user=get_auth_user()):
        available_sections.append("admin")
    if st.session_state.get("file_view") not in available_sections:
        st.session_state["file_view"] = available_sections[0]

    selected_section = st.segmented_control(
        "Sekcja analizy plikow",
        options=available_sections,
        selection_mode="single",
        default=st.session_state.get("file_view", available_sections[0]),
        required=True,
        key="file_view",
        format_func=lambda value: FILE_VIEW_OPTIONS.get(value, value),
        width="stretch",
    )
    selected_section = selected_section or available_sections[0]

    if not workspace_is_ready():
        st.info("Dodaj dwa pliki Excel, aby uruchomic porownanie release'ow, planner oraz eksport.")
        return

    if selected_section == "overview":
        ui_shell.render_panel_intro(
            "Workspace",
            "Status analizy i wyniki",
            "Ten widok porzadkuje upload, status parsera, komunikaty robocze oraz glowne akcje eksportu.",
        )
        overview_metrics = [
            {
                "label": "Pliki zaladowane",
                "value": "2 / 2",
                "copy": describe_format_context(prev_meta, curr_meta),
                "tone": "positive",
            },
            {
                "label": "Widoczne wiersze",
                "value": f"{len(filtered_df):,}",
                "copy": "Wynik po aktywnych filtrach.",
                "tone": "neutral",
            },
            {
                "label": "Produkty",
                "value": f"{product_summary['Part Number'].nunique():,}",
                "copy": "Unikalne materialy w aktualnym zakresie.",
                "tone": "neutral",
            },
            {
                "label": "Eksport",
                "value": "CSV + 2x Excel",
                "copy": "Standardowy raport oraz nowy Weekly by Part sa gotowe do pobrania.",
                "tone": "neutral",
            },
        ]
        ui.render_kpi_cards(overview_metrics)
        build_workspace_context_cards(prev_meta, curr_meta, filter_state, filtered_df)
        if filtered_df.empty:
            st.warning("Po aktywnych filtrach nie ma danych do pokazania w wynikach analizy plikow.")
        else:
            st.success("Analiza plikow jest gotowa. Mozesz przejsc do plannera, wykresow lub pobrac eksport.")
            preview_table = build_detail_export_table(filtered_df).head(40)
            st.dataframe(preview_table, use_container_width=True, height=360)
        render_extended_export_actions(csv_bytes, excel_bytes, professional_excel_bytes)
        return

    if selected_section == "planner":
        render_module_content("planner", module_data, ui)
        return

    if selected_section == "details":
        render_module_content("details", module_data, ui)
        return

    render_module_content("admin", module_data, ui)


def render_sidebar_upload_controls():
    render_section_header(
        "Workspace",
        "Pliki wejściowe",
        "Dodaj poprzedni i aktualny release. Logika uploadu i przechowywania plikow pozostaje bez zmian.",
    )

    render_upload_card(
        "Poprzedni",
        "Baseline release",
        "Plik referencyjny do porownania aktualnego stanu planu i wysylek.",
    )
    previous_upload = st.file_uploader(
        "Upload Previous Release",
        type=["xlsx"],
        key=get_upload_widget_key("previous"),
        label_visibility="visible",
    )
    if previous_upload is not None:
        store_uploaded_release("previous", previous_upload)
    stored_previous = get_stored_upload("previous")
    if stored_previous is not None:
        st.caption(f"Zaladowany plik: {stored_previous['name']}")
        if st.button("Usun poprzedni plik", key="sidebar_clear_previous_upload", use_container_width=True):
            clear_uploaded_release("previous")
            st.rerun()

    render_upload_card(
        "Aktualny",
        "Current release",
        "Plik, z ktorego aplikacja liczy zmiany, alerty i aktualny wolumen.",
    )
    current_upload = st.file_uploader(
        "Upload Current Release",
        type=["xlsx"],
        key=get_upload_widget_key("current"),
        label_visibility="visible",
    )
    if current_upload is not None:
        store_uploaded_release("current", current_upload)
    stored_current = get_stored_upload("current")
    if stored_current is not None:
        st.caption(f"Zaladowany plik: {stored_current['name']}")
        if st.button("Usun aktualny plik", key="sidebar_clear_current_upload", use_container_width=True):
            clear_uploaded_release("current")
            st.rerun()

    return stored_previous, stored_current


def render_sidebar_filters(analysis_bundle=None):
    render_sidebar_user(st)
    previous_release = get_stored_upload("previous")
    current_release = get_stored_upload("current")

    prev_meta = analysis_bundle["prev_meta"] if analysis_bundle else None
    curr_meta = analysis_bundle["curr_meta"] if analysis_bundle else None
    brand_context = (
        analysis_bundle["brand_context"]
        if analysis_bundle
        else detect_brand_context(
            {"file_name": previous_release.get("name")} if previous_release else None,
            {"file_name": current_release.get("name")} if current_release else None,
        )
    )

    render_filter_panel_shell(
        kicker="Filters",
        title="Filtry i status analizy",
        copy="Filtry i kalendarz pozostaja stale widoczne w lewej kolumnie podczas pracy z analizą.",
    )
    render_side_panel_brand(brand_context)

    if analysis_bundle is None:
        render_sidebar_upload_controls()
        render_file_slot_cards(
            prev_file=previous_release,
            current_file=current_release,
        )
        st.info("Dodaj oba pliki, aby aktywowac filtry i porownanie release'ow.")
        return build_default_filter_state(), brand_context

    st.caption(brand_context.get("format_copy", ""))
    if st.session_state.get("success_status"):
        st.success(st.session_state["success_status"])
    filter_state = render_filters_panel(analysis_bundle["result"])
    st.markdown('<hr class="side-panel-divider" />', unsafe_allow_html=True)
    render_section_header(
        "Workspace",
        "Pliki i status",
        "Upload i status pozostaja dostepne ponizej filtrow bez zaslaniania kalendarza.",
    )
    render_sidebar_upload_controls()
    render_file_slot_cards(
        prev_file=None if prev_meta else previous_release,
        current_file=None if curr_meta else current_release,
        prev_meta=prev_meta,
        curr_meta=curr_meta,
    )
    if uploads_ready() and st.button("Przelicz analize", key="sidebar_recompute_analysis", use_container_width=True):
        trigger_analysis_refresh()
        st.rerun()
    return filter_state, brand_context


def render_sidebar_preload_state():
    previous_release = get_stored_upload("previous")
    current_release = get_stored_upload("current")
    brand_context = detect_brand_context(
        {"file_name": previous_release.get("name")} if previous_release else None,
        {"file_name": current_release.get("name")} if current_release else None,
    )

    with st.sidebar:
        render_sidebar_user(st)
        render_filter_panel_shell(
            kicker="Workspace",
            title="Status analizy",
            copy="Upload plikow jest widoczny od razu w glownej sekcji startowej. Filtry aktywuja sie po dodaniu obu release'ow.",
        )
        render_side_panel_brand(brand_context)
        render_file_slot_cards(
            prev_file=previous_release,
            current_file=current_release,
        )
        if uploads_ready() and st.button("Uruchom analize", key="sidebar_run_analysis", use_container_width=True):
            trigger_analysis_refresh()
            st.rerun()
        st.info("Dodaj oba pliki, aby aktywowac Dashboard i Reports.")

    return brand_context


def render_dashboard_view(module_data, ui):
    module_data.module_access = get_module_access_level("dashboard", auth_user=get_auth_user())
    render_module_content("dashboard", module_data, ui)


def render_reports_view(module_data, ui):
    module_data.module_access = get_module_access_level("reports", auth_user=get_auth_user())
    render_module_content("reports", module_data, ui)
    render_section_header(
        "Export",
        "Eksport analityczny",
        "Pelny eksport filtrowanych danych oraz raportow Excel pozostaje dostepny w sekcji Reports.",
    )
    render_extended_export_actions(
        module_data.csv_bytes or b"",
        module_data.excel_bytes or b"",
        module_data.professional_excel_bytes or b"",
    )


init_auth_state()
init_ui_state()

if not st.session_state["authenticated"]:
    render_login_screen()
    st.stop()

if st.session_state.get("active_view") not in PRIMARY_VIEW_KEYS:
    st.session_state["active_view"] = "dashboard"

sync_uploaded_files_from_widgets()

analysis_bundle = None
analysis_error = None
if workspace_is_ready():
    try:
        analysis_bundle = analyze_uploaded_releases()
    except Exception as exc:
        analysis_error = exc
        st.session_state["analysis_results"] = None
        st.session_state["error_message"] = str(exc)
        st.session_state["success_status"] = ""
        set_pipeline_debug(f"analysis failed: {exc}")

logo_markup = ui_shell.build_logo_markup(logo_data_uri())

if analysis_error is not None:
    st.error(f"Blad wczytywania plikow: {analysis_error}")

if analysis_bundle is None:
    sidebar_brand_context = render_sidebar_preload_state()
    render_app_header(
        sidebar_brand_context,
        APP_TITLE,
        "Enterprise dashboard do porownywania release'ow. Dashboard i Reports aktywuja sie po zaladowaniu dwoch plikow Excel.",
        meta_items=[
            "Tylko Dashboard i Reports",
            "Filtry pozostaja po lewej stronie",
            "Planner i pozostala logika zachowane w repo",
        ],
        file_caption="Oczekiwanie na komplet plikow wejsciowych",
    )
    ui_shell.render_panel_intro(
        "Workspace",
        "Dashboard oczekuje na dane",
        "Dodaj poprzedni i aktualny release. Upload jest widoczny bezposrednio w glownej sekcji startowej.",
    )
    render_preload_state(logo_markup)
    st.stop()

filters_col, content_col = st.columns([0.30, 0.70], gap="large")

with filters_col:
    filter_state, sidebar_brand_context = render_sidebar_filters(analysis_bundle)

prev_meta = analysis_bundle["prev_meta"]
curr_meta = analysis_bundle["curr_meta"]
result = analysis_bundle["result"]
brand_context = analysis_bundle["brand_context"]

date_basis = filter_state["date_basis"]
selected_start_date = filter_state["selected_start_date"]
selected_end_date = filter_state["selected_end_date"]
selected_products = filter_state["selected_products"]
search_term = filter_state["search_term"]
selected_change_directions = filter_state["selected_change_directions"]
only_alerts = filter_state["only_alerts"]

filtered_df = apply_analysis_filters(result, filter_state)
st.session_state["filtered_data"] = filtered_df.copy()

planner_source = build_planner_scope_source(result, filter_state)

product_summary = summarize_products(filtered_df)
date_summary = summarize_dates(filtered_df, date_basis)
weekly_summary = build_weekly_summary(
    filtered_df,
    date_basis,
    selected_start_date,
    selected_end_date,
    selected_end_date,
    THRESHOLD,
)
key_findings = build_key_findings(
    filtered_df, product_summary, date_summary, date_basis
)

detail_export_table = build_detail_export_table(filtered_df)
csv_bytes = detail_export_table.to_csv(index=False).encode("utf-8")
current_matrix_for_export = build_matrix(filtered_df, date_basis, "Current Quantity")
delta_matrix_for_export = build_matrix(filtered_df, date_basis, "Delta")
excel_bytes = to_excel_bytes(
    filtered_df,
    weekly_summary,
    current_matrix_for_export,
    delta_matrix_for_export,
    prev_meta,
    curr_meta,
    product_summary,
    date_basis,
    selected_start_date,
    selected_end_date,
    key_findings,
)
professional_excel_bytes = to_professional_weekly_report_bytes(
    filtered_df,
    prev_meta,
    curr_meta,
    date_basis,
    selected_start_date,
    selected_end_date,
)

ui = build_ui_helpers()
module_data = build_module_context(
    filtered_df,
    planner_source,
    product_summary,
    date_summary,
    weekly_summary,
    key_findings,
    prev_meta,
    curr_meta,
    date_basis,
    selected_start_date,
    selected_end_date,
    excel_bytes=excel_bytes,
    csv_bytes=csv_bytes,
    professional_excel_bytes=professional_excel_bytes,
)

with content_col:
    render_app_header(
        brand_context,
        APP_TITLE,
        "Premium dark workspace z dwoma glownymi modulami: Dashboard i Reports. Filtry oraz upload pozostaja po lewej stronie.",
        meta_items=[
            describe_format_context(prev_meta, curr_meta),
            f"PO {curr_meta.get('po_number', 'n/a')}",
            f"Zakres {format_workspace_date_range(filter_state)}",
            f"Tygodnie {format_workspace_week_range(filter_state)}",
            f"Wiersze po filtrach: {len(filtered_df):,}",
        ],
        file_caption=curr_meta.get("file_name", ""),
    )
    build_workspace_context_cards(prev_meta, curr_meta, filter_state, filtered_df)

    selected_view = st.segmented_control(
        "Main view",
        options=list(MAIN_VIEW_OPTIONS),
        selection_mode="single",
        default=st.session_state.get("active_view", "dashboard"),
        required=True,
        key="active_view",
        format_func=lambda value: MODULE_LABELS.get(value, value.title()),
        width="stretch",
    )
    selected_view = selected_view or "dashboard"

    if selected_view == "reports":
        render_reports_view(module_data, ui)
    else:
        render_dashboard_view(module_data, ui)

st.stop()

app_sidebar, app_main = st.columns([0.27, 0.73], gap="large")

with app_sidebar:
    render_sidebar_user(st)

with app_main:
    prev_file, current_file = render_preload_state()

if prev_file is None or current_file is None:
    with app_sidebar:
        render_welcome_side_panel(prev_file, current_file)
    st.stop()

try:
    prev_df, prev_meta = load_release(prev_file.getvalue(), prev_file.name)
    curr_df, curr_meta = load_release(current_file.getvalue(), current_file.name)
    result = compare_releases(prev_df, curr_df)
except Exception as exc:
    with app_sidebar:
        render_welcome_side_panel(prev_file, current_file)
    with app_main:
        render_welcome_state(prev_file, current_file)
        st.error(f"Błąd wczytywania plików: {exc}")
    st.stop()

brand_context = detect_brand_context(prev_meta, curr_meta)
with app_sidebar:
    filter_state = render_analysis_side_panel(
        result,
        brand_context,
        prev_meta=prev_meta,
        curr_meta=curr_meta,
    )
    active_module = render_module_navigation(auth_user=get_auth_user())

date_basis = filter_state["date_basis"]
selected_start_date = filter_state["selected_start_date"]
selected_end_date = filter_state["selected_end_date"]
selected_products = filter_state["selected_products"]
search_term = filter_state["search_term"]
selected_change_directions = filter_state["selected_change_directions"]
only_alerts = filter_state["only_alerts"]

filtered_df = result.copy()
filtered_df = filtered_df[
    filtered_df[date_basis].dt.date.between(
        selected_start_date, selected_end_date
    )
]

if selected_products:
    filtered_df = filtered_df[filtered_df["Product Label"].isin(selected_products)]
else:
    filtered_df = filtered_df.iloc[0:0]

if search_term.strip():
    query = search_term.strip().lower()
    filtered_df = filtered_df[
        filtered_df["Part Number"].str.lower().str.contains(query, na=False)
        | filtered_df["Part Description"].str.lower().str.contains(query, na=False)
    ]

planner_source = build_planner_scope_source(
    result,
    selected_start_date,
    selected_end_date,
    selected_products,
    search_term,
)

filtered_df = filtered_df[
    filtered_df["Change Direction"].isin(selected_change_directions)
]

if only_alerts:
    filtered_df = filtered_df[filtered_df["Alert"]]

product_summary = summarize_products(filtered_df)
date_summary = summarize_dates(filtered_df, date_basis)
weekly_summary = build_weekly_summary(
    filtered_df,
    date_basis,
    selected_start_date,
    selected_end_date,
    selected_end_date,
    THRESHOLD,
)
key_findings = build_key_findings(
    filtered_df, product_summary, date_summary, date_basis
)

detail_export_table = build_detail_export_table(filtered_df)
csv_bytes = detail_export_table.to_csv(index=False).encode("utf-8")
current_matrix_for_export = build_matrix(filtered_df, date_basis, "Current Quantity")
delta_matrix_for_export = build_matrix(filtered_df, date_basis, "Delta")
excel_bytes = to_excel_bytes(
    filtered_df,
    weekly_summary,
    current_matrix_for_export,
    delta_matrix_for_export,
    prev_meta,
    curr_meta,
    product_summary,
    date_basis,
    selected_start_date,
    selected_end_date,
    key_findings,
)

with app_sidebar:
    render_export_actions(csv_bytes, excel_bytes)

with app_main:
    render_module_frame(
        active_module,
        filtered_df,
        planner_source,
        product_summary,
        date_summary,
        weekly_summary,
        key_findings,
        prev_meta,
        curr_meta,
        date_basis,
        selected_start_date,
        selected_end_date,
        excel_bytes=excel_bytes,
        csv_bytes=csv_bytes,
    )

st.stop()

app_sidebar, app_main = st.columns([0.28, 0.72], gap="large")

with app_main:
    upload_left, upload_right = st.columns(2, gap="large")
    with upload_left:
        render_upload_card(
            "Krok 1",
            "Poprzedni release / poprzedni plan",
            "Dodaj bazowy plik Excel, do ktorego bedzie porownywany aktualny stan zamowien i wysylek.",
        )
        prev_file = st.file_uploader(
            "Upload Previous Release",
            type=["xlsx"],
            key="previous_release_upload",
            label_visibility="visible",
        )
    with upload_right:
        render_upload_card(
            "Krok 2",
            "Aktualny release / aktualny plan",
            "Dodaj nowy plik Excel, aby dashboard automatycznie policzyl delty, alerty i zmiany procentowe.",
        )
        current_file = st.file_uploader(
            "Upload Current Release",
            type=["xlsx"],
            key="current_release_upload",
            label_visibility="visible",
        )

if prev_file is None and current_file is None:
    with app_sidebar:
        render_welcome_side_panel(prev_file, current_file)
    with app_main:
        quick_cols = st.columns(3, gap="large")
        with quick_cols[0]:
            render_quick_card(
                "Czytelny dashboard porownawczy",
                "Aplikacja zestawia poprzedni i aktualny release, od razu pokazujac bilans zmian, alerty oraz produkty z najwiekszym ruchem.",
            )
        with quick_cols[1]:
            render_quick_card(
                "Macierz podobna do Excela",
                "Otrzymujesz widok tabelaryczny z datami, zmianami ilosci i filtrowaniem po produkcie, kierunku ruchu oraz zakresie dat.",
            )
        with quick_cols[2]:
            render_quick_card(
                "Raport gotowy do wyslania",
                "Po analizie pobierzesz CSV oraz biznesowy raport Excel z podsumowaniem KPI i kluczowymi zmianami.",
            )
        st.info(
            "Zacznij od dodania dwoch plikow Excel. Po zaladowaniu obu release'ow dashboard uruchomi pelna analize porownawcza."
        )
    st.stop()

if prev_file is None or current_file is None:
    with app_sidebar:
        render_welcome_side_panel(prev_file, current_file)
    with app_main:
        missing_label = "poprzedni" if prev_file is None else "aktualny"
        loaded_label = "aktualny" if prev_file is None else "poprzedni"
        st.info(
            f"Plik {loaded_label} jest juz dodany. Dodaj jeszcze plik {missing_label}, aby uruchomic analize i wygenerowac dashboard."
        )
    st.stop()

try:
    prev_df, prev_meta = load_release(prev_file.getvalue(), prev_file.name)
    curr_df, curr_meta = load_release(current_file.getvalue(), current_file.name)
    result = compare_releases(prev_df, curr_df)
except Exception as exc:
    with app_sidebar:
        render_welcome_side_panel(prev_file, current_file)
    with app_main:
        st.error(f"Blad: {exc}")
    st.stop()

brand_context = detect_brand_context(prev_meta, curr_meta)
with app_sidebar:
    filter_state = render_analysis_side_panel(result, brand_context)
    active_module = render_module_navigation(auth_user=get_auth_user())

date_basis = filter_state["date_basis"]
selected_start_date = filter_state["selected_start_date"]
selected_end_date = filter_state["selected_end_date"]
selected_products = filter_state["selected_products"]
search_term = filter_state["search_term"]
selected_change_directions = filter_state["selected_change_directions"]
only_alerts = filter_state["only_alerts"]

filtered_df = result.copy()
filtered_df = filtered_df[
    filtered_df[date_basis].dt.date.between(
        selected_start_date, selected_end_date
    )
]

if selected_products:
    filtered_df = filtered_df[filtered_df["Product Label"].isin(selected_products)]
else:
    filtered_df = filtered_df.iloc[0:0]

if search_term.strip():
    query = search_term.strip().lower()
    filtered_df = filtered_df[
        filtered_df["Part Number"].str.lower().str.contains(query, na=False)
        | filtered_df["Part Description"].str.lower().str.contains(query, na=False)
    ]

planner_source = build_planner_scope_source(
    result,
    selected_start_date,
    selected_end_date,
    selected_products,
    search_term,
)

filtered_df = filtered_df[
    filtered_df["Change Direction"].isin(selected_change_directions)
]

if only_alerts:
    filtered_df = filtered_df[filtered_df["Alert"]]

product_summary = summarize_products(filtered_df)
date_summary = summarize_dates(filtered_df, date_basis)
weekly_summary = build_weekly_summary(
    filtered_df,
    date_basis,
    selected_start_date,
    selected_end_date,
    selected_end_date,
    THRESHOLD,
)
key_findings = build_key_findings(
    filtered_df, product_summary, date_summary, date_basis
)

with app_main:
    render_module_frame(
        active_module,
        filtered_df,
        planner_source,
        product_summary,
        date_summary,
        weekly_summary,
        key_findings,
        prev_meta,
        curr_meta,
        date_basis,
        selected_start_date,
        selected_end_date,
    )

st.stop()

render_sidebar_user(st)

upload_left, upload_right = st.columns(2, gap="large")
with upload_left:
    render_upload_card(
        "Krok 1",
        "Poprzedni release / poprzedni plan",
        "Dodaj bazowy plik Excel, do którego będzie porównywany aktualny stan zamówień i wysyłek.",
    )
    prev_file = st.file_uploader(
        "Upload Previous Release",
        type=["xlsx"],
        key="previous_release_upload",
        label_visibility="visible",
    )
with upload_right:
    render_upload_card(
        "Krok 2",
        "Aktualny release / aktualny plan",
        "Dodaj nowy plik Excel, aby dashboard automatycznie policzył delty, alerty i zmiany procentowe.",
    )
    current_file = st.file_uploader(
        "Upload Current Release",
        type=["xlsx"],
        key="current_release_upload",
        label_visibility="visible",
    )

if prev_file is None and current_file is None:
    quick_cols = st.columns(3, gap="large")
    with quick_cols[0]:
        render_quick_card(
            "Czytelny dashboard porównawczy",
            "Aplikacja zestawia poprzedni i aktualny release, od razu pokazując bilans zmian, alerty oraz produkty z największym ruchem.",
        )
    with quick_cols[1]:
        render_quick_card(
            "Macierz podobna do Excela",
            "Otrzymujesz widok tabelaryczny z datami, zmianami ilości i filtrowaniem po produkcie, kierunku ruchu oraz zakresie dat.",
        )
    with quick_cols[2]:
        render_quick_card(
            "Raport gotowy do wysłania",
            "Po analizie pobierzesz CSV oraz biznesowy raport Excel z podsumowaniem KPI i kluczowymi zmianami.",
        )
    st.info("Zacznij od dodania dwóch plików Excel. Po załadowaniu obu release'ów dashboard uruchomi pełną analizę porównawczą.")
elif prev_file is None or current_file is None:
    missing_label = "poprzedni" if prev_file is None else "aktualny"
    loaded_label = "aktualny" if prev_file is None else "poprzedni"
    st.info(
        f"Plik {loaded_label} jest już dodany. Dodaj jeszcze plik {missing_label}, aby uruchomić analizę i wygenerować dashboard."
    )
else:
    try:
        prev_df, prev_meta = load_release(prev_file.getvalue(), prev_file.name)
        curr_df, curr_meta = load_release(current_file.getvalue(), current_file.name)
        result = compare_releases(prev_df, curr_df)
    except Exception as exc:
        st.error(f"Błąd: {exc}")
    else:
        filter_col, _ = st.columns([0.32, 0.68], gap="large")
        with filter_col:
            filter_state = render_filter_controls(result)

        date_basis = filter_state["date_basis"]
        selected_start_date = filter_state["selected_start_date"]
        selected_end_date = filter_state["selected_end_date"]
        full_product_summary = filter_state["full_product_summary"]
        all_products = full_product_summary["Product Label"].tolist()
        selected_products = filter_state["selected_products"]
        search_term = filter_state["search_term"]
        selected_change_directions = filter_state["selected_change_directions"]
        only_alerts = filter_state["only_alerts"]

        filtered_df = result.copy()
        filtered_df = filtered_df[
            filtered_df[date_basis].dt.date.between(
                selected_start_date, selected_end_date
            )
        ]

        if selected_products:
            filtered_df = filtered_df[filtered_df["Product Label"].isin(selected_products)]
        else:
            filtered_df = filtered_df.iloc[0:0]

        if search_term.strip():
            query = search_term.strip().lower()
            filtered_df = filtered_df[
                filtered_df["Part Number"].str.lower().str.contains(query, na=False)
                | filtered_df["Part Description"].str.lower().str.contains(query, na=False)
            ]

        filtered_df = filtered_df[
            filtered_df["Change Direction"].isin(selected_change_directions)
        ]

        if only_alerts:
            filtered_df = filtered_df[filtered_df["Alert"]]

        product_summary = summarize_products(filtered_df)
        date_summary = summarize_dates(filtered_df, date_basis)
        weekly_summary = build_weekly_summary(
            filtered_df,
            date_basis,
            selected_start_date,
            selected_end_date,
            selected_end_date,
            THRESHOLD,
        )
        key_findings = build_key_findings(
            filtered_df, product_summary, date_summary, date_basis
        )

        st.success("Analiza porównawcza jest gotowa.")

        hero_left, hero_right = st.columns([1.8, 1], gap="large")
        with hero_left:
            hero_logo_html = (
                f'<img class="hero-logo" src="{logo_data_uri()}" alt="{BRAND_NAME} logo" />'
                if logo_available()
                else f'<div class="brand-badge">{BRAND_NAME}</div>'
            )
            st.markdown(
                f"""
                <div class="hero-card">
                    {hero_logo_html}
                    <div class="hero-kicker">Release Intelligence</div>
                    <div class="hero-title">Raport zmian dla PO {curr_meta['po_number']}</div>
                    <p class="hero-copy">
                        Porównaj wersje release'ów, śledź ruch dzień po dniu i szybko
                        wychwyć produkty, które zwiększyły lub zmniejszyły wolumen w wybranym oknie.
                    </p>
                </div>
                """,
                unsafe_allow_html=True,
            )
        with hero_right:
            st.markdown(
                f"""
                <div class="hero-card">
                    <div class="hero-kicker">Aktywne okno analizy</div>
                    <div class="hero-title">{selected_start_date.strftime('%Y-%m-%d')}</div>
                    <p class="hero-copy">
                        do {selected_end_date.strftime('%Y-%m-%d')} na osi <strong>{get_date_label(date_basis)}</strong>
                    </p>
                    <div class="hero-stat-grid">
                        <div class="hero-stat">
                            <div class="hero-stat-label">Poprzedni release</div>
                            <div class="hero-stat-value">{format_release_label(prev_meta)}</div>
                        </div>
                        <div class="hero-stat">
                            <div class="hero-stat-label">Aktualny release</div>
                            <div class="hero-stat-value">{format_release_label(curr_meta)}</div>
                        </div>
                        <div class="hero-stat">
                            <div class="hero-stat-label">Poprzedni plik</div>
                            <div class="hero-stat-value">{prev_meta['file_name']}</div>
                        </div>
                        <div class="hero-stat">
                            <div class="hero-stat-label">Aktualny plik</div>
                            <div class="hero-stat-value">{curr_meta['file_name']}</div>
                        </div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        header_left, header_right = st.columns([1.5, 1], gap="large")
        with header_left:
            render_meta_card(
                "Kontekst release'u",
                [
                    f"<strong>Numer PO:</strong> {curr_meta['po_number']}",
                    (
                        f"<strong>Poprzedni release:</strong> {format_release_summary(prev_meta)}"
                    ),
                    (
                        f"<strong>Aktualny release:</strong> {format_release_summary(curr_meta)}"
                    ),
                ],
            )
        with header_right:
            render_meta_card(
                "Planista",
                [
                    f"<strong>Planista:</strong> {curr_meta['planner_name']}",
                    f"<strong>Email:</strong> {curr_meta['planner_email']}",
                    (
                        f"<strong>Produkty w zakresie:</strong> "
                        f"{product_summary['Part Number'].nunique()}"
                    ),
                ],
            )

        if filtered_df.empty:
            st.warning(
                "Po zastosowaniu filtrów nie ma danych do pokazania. "
                "Poszerz zakres dat albo przywróć produkty w filtrach bocznych."
            )
        else:
            total_prev = filtered_df["Quantity_Prev"].sum()
            total_curr = filtered_df["Quantity_Curr"].sum()
            total_delta = filtered_df["Delta"].sum()
            alert_count = int(filtered_df["Alert"].sum())
            products_changed = int((product_summary["Delta"] != 0).sum())

            render_status_pills(total_delta, alert_count, products_changed)
            metric_cols = st.columns(5)
            metric_cols[0].metric("Poprzednia ilość", f"{total_prev:,.0f}")
            metric_cols[1].metric(
                "Aktualna ilość",
                f"{total_curr:,.0f}",
                delta=f"{total_curr - total_prev:+,.0f}",
            )
            metric_cols[2].metric("Bilans zmian", f"{total_delta:+,.0f}")
            metric_cols[3].metric(
                "Liczba alertów",
                f"{alert_count:,}",
                delta=f"{(alert_count / len(filtered_df)):.1%}",
                delta_color="inverse",
            )
            metric_cols[4].metric("Zmienne produkty", f"{products_changed:,}")

            reference_week = get_last_completed_reference_week(selected_end_date)
            reference_row, previous_week_row = get_reference_week_rows(weekly_summary)
            reference_week_label = (
                reference_row["Week Label"] if reference_row is not None else reference_week.week_label
            )
            reference_range_label = (
                format_week_range(reference_row["Week Start"], reference_row["Week End"])
                if reference_row is not None
                else format_week_range(reference_week.week_start, reference_week.week_end)
            )
            reference_release_delta = (
                format_signed_int(reference_row["Delta"]) if reference_row is not None else "+0"
            )
            reference_release_pct = (
                format_percent_display(reference_row["Release Percent Label"])
                if reference_row is not None
                else "n/a"
            )
            reference_wow_delta = (
                format_signed_int(reference_row["WoW Delta"]) if reference_row is not None else "+0"
            )
            reference_wow_pct = (
                format_percent_display(reference_row["WoW Percent Label"])
                if reference_row is not None
                else "n/a"
            )
            reference_working_days = (
                int(reference_row["Working_Days_PL"]) if reference_row is not None else 0
            )
            reference_per_day = (
                "n/a"
                if reference_row is None or pd.isna(reference_row["Avg Current / Working Day"])
                else f"{float(reference_row['Avg Current / Working Day']):,.2f} / dzien"
            )
            previous_week_label = (
                previous_week_row["Week Label"] if previous_week_row is not None else "brak"
            )

            st.caption(
                f"Analiza tygodniowa odnosi sie do {reference_week_label} ({reference_range_label}). "
                f"Data referencyjna: {selected_end_date:%Y-%m-%d}. "
                + (
                    "Poniewaz data koncowa wypada w trakcie tygodnia, jako referencje przyjeto ostatni pelny zakonczony tydzien ISO."
                    if selected_end_date.weekday() != 6
                    else "Poniewaz data koncowa wypada w niedziele, ten tydzien zostal uznany za pelny zakonczony tydzien ISO."
                )
            )

            weekly_metric_cols = st.columns(5)
            weekly_metric_cols[0].metric(
                "Referencyjny tydzien ISO",
                reference_week_label,
                delta=reference_range_label,
            )
            weekly_metric_cols[1].metric(
                "Aktualny wolumen tygodnia",
                f"{float(reference_row['Quantity_Curr']):,.0f}" if reference_row is not None else "0",
                delta=reference_release_delta,
            )
            weekly_metric_cols[2].metric(
                "Zmiana vs poprzedni release",
                reference_release_pct,
                delta=f"prev {float(reference_row['Quantity_Prev']):,.0f}" if reference_row is not None else "prev 0",
            )
            weekly_metric_cols[3].metric(
                "Zmiana WoW",
                reference_wow_delta,
                delta=f"{reference_wow_pct} vs {previous_week_label}",
            )
            weekly_metric_cols[4].metric(
                "Dni robocze PL",
                f"{reference_working_days}",
                delta=reference_per_day,
            )

            st.markdown(
                """
                <div class="section-banner">
                    <div class="section-kicker">Executive Summary</div>
                    <div class="section-copy">
                        Najważniejsze sygnały, które warto sprawdzić w pierwszej kolejności.
                        Duży nagłówek pokazuje nazwę produktu, a krótki opis pod spodem wyjaśnia znaczenie zmiany.
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.subheader("Kluczowe wnioski")
            finding_cols = st.columns(max(1, min(len(key_findings), 4)), gap="large")
            for idx, finding in enumerate(key_findings):
                with finding_cols[idx]:
                    render_finding_card(
                        finding["label"], finding["title"], finding["copy"]
                    )

            dashboard_tab, weekly_tab, product_tab, matrix_tab, detail_tab = st.tabs(
                ["Dashboard", "Analiza tygodniowa", "Raport produktu", "Macierz release'u", "Dane szczegółowe"]
            )

            with dashboard_tab:
                st.subheader(f"Trend zmian według osi: {get_date_label(date_basis)}")
                render_chart_table_switch(
                    "dashboard_trend",
                    build_quantity_chart(date_summary, get_date_label(date_basis)),
                    date_summary,
                    table_height=360,
                )

                trend_left, trend_right = st.columns([1.45, 1], gap="large")
                with trend_left:
                    render_chart_table_switch(
                        "dashboard_delta",
                        build_delta_chart(date_summary, get_date_label(date_basis)),
                        date_summary,
                        table_height=320,
                    )
                with trend_right:
                    st.subheader("Struktura zmian")
                    render_chart_table_switch(
                        "dashboard_mix",
                        build_change_mix_chart(filtered_df),
                        build_change_mix_source(filtered_df),
                        table_height=240,
                    )

                increase_chart, increase_title = build_product_bar_chart(
                    product_summary, "increase"
                )
                decrease_chart, decrease_title = build_product_bar_chart(
                    product_summary, "decrease"
                )
                dashboard_left, dashboard_right = st.columns(2)

                with dashboard_left:
                    st.subheader(increase_title)
                    if increase_chart is None:
                        st.info("Brak produktów ze wzrostem w aktualnym filtrowaniu.")
                    else:
                        render_chart_table_switch(
                            "dashboard_increase",
                            increase_chart,
                            build_product_bar_source(product_summary, "increase"),
                            table_height=340,
                        )

                with dashboard_right:
                    st.subheader(decrease_title)
                    if decrease_chart is None:
                        st.info("Brak produktów ze spadkiem w aktualnym filtrowaniu.")
                    else:
                        render_chart_table_switch(
                            "dashboard_decrease",
                            decrease_chart,
                            build_product_bar_source(product_summary, "decrease"),
                            table_height=340,
                        )

                st.subheader("Najważniejsze zmiany")
                highlight_table = (
                    product_summary.assign(Abs_Delta=product_summary["Delta"].abs())
                    .sort_values("Abs_Delta", ascending=False)
                    .drop(columns=["Abs_Delta"])
                    .head(10)
                )
                highlight_table["Quantity_Prev"] = highlight_table["Quantity_Prev"].map(
                    lambda value: f"{value:,.0f}"
                )
                highlight_table["Quantity_Curr"] = highlight_table["Quantity_Curr"].map(
                    lambda value: f"{value:,.0f}"
                )
                highlight_table["Delta"] = highlight_table["Delta"].map(format_signed_int)
                highlight_table = highlight_table.rename(
                    columns={
                        "Part Number": "Numer części",
                        "Part Description": "Opis produktu",
                        "Quantity_Prev": "Poprzednia ilość",
                        "Quantity_Curr": "Aktualna ilość",
                        "Delta": "Zmiana ilości",
                        "Alert_Count": "Liczba alertów",
                        "Change Direction": "Kierunek zmiany",
                    }
                )
                st.dataframe(highlight_table, use_container_width=True, height=360)

                st.subheader("Tygodnie ISO")
                weekly_chart = build_weekly_quantity_chart(weekly_summary)
                weekly_preview = prepare_weekly_display_table(weekly_summary).tail(8)
                render_chart_table_switch(
                    "dashboard_weekly",
                    weekly_chart,
                    weekly_preview,
                    chart_empty_message="Brak danych tygodniowych do wykresu.",
                    table_height=320,
                )

            with weekly_tab:
                st.subheader("Analiza tygodniowa oparta na datach")
                weekly_partial = weekly_summary[
                    weekly_summary["Is Partial Range"] | ~weekly_summary["Is Closed Week"]
                ]
                if not weekly_partial.empty:
                    st.info(
                        "W tabeli i wykresach tygodnie oznaczone jako 'Partial range' lub 'Open week' "
                        "obejmują niepełny zakres albo nie były jeszcze zakończone względem daty referencyjnej."
                    )

                weekly_qty_chart = build_weekly_quantity_chart(weekly_summary)
                render_chart_table_switch(
                    "weekly_quantity",
                    weekly_qty_chart,
                    prepare_weekly_display_table(weekly_summary),
                    chart_empty_message="Brak danych tygodniowych do wykresu.",
                    table_height=360,
                )

                weekly_left, weekly_right = st.columns([1.3, 1], gap="large")
                with weekly_left:
                    weekly_delta_chart = build_weekly_delta_chart(weekly_summary)
                    render_chart_table_switch(
                        "weekly_delta",
                        weekly_delta_chart,
                        prepare_weekly_display_table(weekly_summary),
                        chart_empty_message="Brak danych tygodniowych do wykresu delta.",
                        table_height=320,
                    )
                with weekly_right:
                    weekly_focus = pd.DataFrame(
                        [
                            {
                                "Widok": "Referencyjny tydzien",
                                "Tydzien ISO": reference_week_label,
                                "Aktualny release": (
                                    f"{float(reference_row['Quantity_Curr']):,.0f}"
                                    if reference_row is not None
                                    else "0"
                                ),
                                "Poprzedni release": (
                                    f"{float(reference_row['Quantity_Prev']):,.0f}"
                                    if reference_row is not None
                                    else "0"
                                ),
                                "Delta release": reference_release_delta,
                                "Zmiana release %": reference_release_pct,
                                "Delta WoW": reference_wow_delta,
                                "Zmiana WoW %": reference_wow_pct,
                            },
                            {
                                "Widok": "Poprzedni tydzien",
                                "Tydzien ISO": previous_week_label,
                                "Aktualny release": (
                                    f"{float(previous_week_row['Quantity_Curr']):,.0f}"
                                    if previous_week_row is not None
                                    else "0"
                                ),
                                "Poprzedni release": (
                                    f"{float(previous_week_row['Quantity_Prev']):,.0f}"
                                    if previous_week_row is not None
                                    else "0"
                                ),
                                "Delta release": (
                                    format_signed_int(previous_week_row["Delta"])
                                    if previous_week_row is not None
                                    else "+0"
                                ),
                                "Zmiana release %": (
                                    format_percent_display(previous_week_row["Release Percent Label"])
                                    if previous_week_row is not None
                                    else "n/a"
                                ),
                                "Delta WoW": (
                                    format_signed_int(previous_week_row["WoW Delta"])
                                    if previous_week_row is not None
                                    else "+0"
                                ),
                                "Zmiana WoW %": (
                                    format_percent_display(previous_week_row["WoW Percent Label"])
                                    if previous_week_row is not None
                                    else "n/a"
                                ),
                            },
                        ]
                    )
                    st.subheader("Porównanie tygodni")
                    st.dataframe(weekly_focus, use_container_width=True, height=240)

                weekly_table = prepare_weekly_display_table(weekly_summary)
                st.subheader("Tabela tygodniowa")
                st.dataframe(weekly_table, use_container_width=True, height=420)

            with product_tab:
                st.subheader("Analiza wybranego produktu")
                selected_product_label = st.selectbox(
                    "Wybierz produkt",
                    options=product_summary["Product Label"].tolist(),
                )
                product_detail = filtered_df[
                    filtered_df["Product Label"] == selected_product_label
                ].sort_values(date_basis)
                product_date_summary = summarize_dates(product_detail, date_basis)

                product_metrics = st.columns(4)
                product_metrics[0].metric(
                    "Poprzednia ilość", f"{product_detail['Quantity_Prev'].sum():,.0f}"
                )
                product_metrics[1].metric(
                    "Aktualna ilość", f"{product_detail['Quantity_Curr'].sum():,.0f}"
                )
                product_metrics[2].metric(
                    "Bilans zmian", f"{product_detail['Delta'].sum():+,.0f}"
                )
                product_metrics[3].metric(
                    "Liczba alertów", int(product_detail["Alert"].sum())
                )

                render_chart_table_switch(
                    "product_quantity",
                    build_quantity_chart(product_date_summary, get_date_label(date_basis)),
                    product_date_summary,
                    table_height=320,
                )
                render_chart_table_switch(
                    "product_delta",
                    build_delta_chart(product_date_summary, get_date_label(date_basis)),
                    product_date_summary,
                    table_height=320,
                )

                product_weekly_summary = build_weekly_summary(
                    product_detail,
                    date_basis,
                    selected_start_date,
                    selected_end_date,
                    selected_end_date,
                    THRESHOLD,
                )
                st.subheader("Tygodnie ISO dla produktu")
                product_weekly_chart = build_weekly_quantity_chart(product_weekly_summary)
                render_chart_table_switch(
                    "product_weekly",
                    product_weekly_chart,
                    prepare_weekly_display_table(product_weekly_summary),
                    chart_empty_message="Brak danych tygodniowych dla wybranego produktu.",
                    table_height=280,
                )

                product_table = product_detail[available_detail_columns(product_detail)].copy()
                product_table["Ship Date"] = product_table["Ship Date"].dt.strftime("%Y-%m-%d")
                product_table["Receipt Date"] = product_table["Receipt Date"].dt.strftime(
                    "%Y-%m-%d"
                )
                product_table["Change Direction"] = product_table["Change Direction"].map(
                    get_change_label
                )
                product_table["Alert"] = product_table["Alert"].map(
                    lambda value: "Tak" if value else "Nie"
                )
                product_table = product_table.rename(
                    columns={
                        "Part Number": "Numer części",
                        "Part Description": "Opis produktu",
                        "Origin Doc": "Origin Doc",
                        "Item": "Pozycja",
                        "Ship To": "Ship-to",
                        "Customer Material": "Materiał klienta",
                        "Unrestricted Qty": "Ilość unrestr.",
                        "Unloading Point": "Punkt rozładunku",
                        "Ship Date": "Data wysyłki",
                        "Receipt Date": "Data odbioru",
                        "Unit of Measure": "JM",
                        "CumQty": "CumQty",
                        "Quantity_Prev": "Poprzednia ilość",
                        "Quantity_Curr": "Aktualna ilość",
                        "Delta": "Zmiana ilości",
                        "Percent Change": "Zmiana %",
                        "Demand Status": "Status popytu",
                        "Change Direction": "Kierunek zmiany",
                        "Alert": "Alert",
                    }
                )
                st.dataframe(product_table, use_container_width=True, height=360)

            with matrix_tab:
                st.subheader("Macierz podobna do arkusza release'u")
                matrix_metric = st.radio(
                    "Metryka",
                    options=["Current Quantity", "Previous Quantity", "Delta", "Percent Change"],
                    horizontal=True,
                    format_func=get_metric_label,
                )
                matrix = build_matrix(filtered_df, date_basis, matrix_metric)
                matrix_cells = matrix.shape[0] * max(matrix.shape[1], 1)

                if matrix.empty:
                    st.info("Brak danych do macierzy.")
                elif matrix_cells <= MAX_MATRIX_STYLE_CELLS:
                    st.dataframe(
                        style_matrix(matrix, matrix_metric),
                        use_container_width=True,
                        height=520,
                    )
                else:
                    st.info(
                        "Macierz jest zbyt duza do stylowania, dlatego pokazuje ja "
                        "bez dodatkowego formatowania."
                    )
                    st.dataframe(matrix, use_container_width=True, height=520)

            with detail_tab:
                st.subheader("Dane szczegółowe")
                preview_limit = st.selectbox(
                    "Liczba wierszy w podglądzie",
                    options=[100, 250, 500, 1000],
                    index=2,
                )
                detail_table = filtered_df[available_detail_columns(filtered_df)].copy()
                detail_table["Ship Date"] = detail_table["Ship Date"].dt.strftime("%Y-%m-%d")
                detail_table["Receipt Date"] = detail_table["Receipt Date"].dt.strftime(
                    "%Y-%m-%d"
                )
                detail_table["Change Direction"] = detail_table["Change Direction"].map(
                    get_change_label
                )
                detail_table["Alert"] = detail_table["Alert"].map(
                    lambda value: "Tak" if value else "Nie"
                )
                detail_table = detail_table.rename(
                    columns={
                        "PO Number": "Numer PO",
                        "Origin Doc": "Origin Doc",
                        "Item": "Pozycja",
                        "Ship To": "Ship-to",
                        "Part Number": "Numer części",
                        "Part Description": "Opis produktu",
                        "Customer Material": "Materiał klienta",
                        "Unrestricted Qty": "Ilość unrestr.",
                        "Unloading Point": "Punkt rozładunku",
                        "Ship Date": "Data wysyłki",
                        "Receipt Date": "Data odbioru",
                        "Unit of Measure": "JM",
                        "CumQty": "CumQty",
                        "Quantity_Prev": "Poprzednia ilość",
                        "Quantity_Curr": "Aktualna ilość",
                        "Delta": "Zmiana ilości",
                        "Percent Change": "Zmiana %",
                        "Demand Status": "Status popytu",
                        "Change Direction": "Kierunek zmiany",
                        "Alert": "Alert",
                    }
                )

                if len(detail_table) > preview_limit:
                    st.info(
                        f"Pokazuje pierwsze {preview_limit} z {len(detail_table)} wierszy. "
                        "Pełny raport jest dostępny do pobrania."
                    )
                st.dataframe(
                    detail_table.head(preview_limit),
                    use_container_width=True,
                    height=420,
                )

                current_matrix_for_export = build_matrix(
                    filtered_df, date_basis, "Current Quantity"
                )
                delta_matrix_for_export = build_matrix(filtered_df, date_basis, "Delta")
                excel_bytes = to_excel_bytes(
                    filtered_df,
                    weekly_summary,
                    current_matrix_for_export,
                    delta_matrix_for_export,
                    prev_meta,
                    curr_meta,
                    product_summary,
                    date_basis,
                    selected_start_date,
                    selected_end_date,
                    key_findings,
                )
                csv_bytes = detail_table.to_csv(index=False).encode("utf-8")

                download_left, download_right = st.columns(2)
                with download_left:
                    st.download_button(
                        "Pobierz filtrowane dane CSV",
                        data=csv_bytes,
                        file_name="pjoter_development_release_change_filtered.csv",
                        mime="text/csv",
                    )
                with download_right:
                    st.download_button(
                        "Pobierz raport Excel",
                        data=excel_bytes,
                        file_name="pjoter_development_release_change_report.xlsx",
                        mime=(
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        ),
                    )
