from __future__ import annotations

import ast
import io
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from types import SimpleNamespace


def load_chart_export_functions() -> dict[str, object]:
    source_path = Path(__file__).resolve().parents[1] / "streamlit_app.py"
    source = source_path.read_text(encoding="utf-8")
    tree = ast.parse(source, filename=str(source_path))
    required_names = {
        "normalize_excel_export_table",
        "get_chart_export_dataset",
        "build_matrix_chart_export_table",
        "prepare_chart_export_plot_table",
        "add_excel_chart_export_data",
        "build_excel_native_chart",
        "build_chart_export_metadata",
        "write_chart_export_metadata",
        "build_excel_chart_workbook",
        "style_excel_header",
        "autosize_worksheet",
        "ensure_numeric_cells_black",
        "style_table_region",
    }
    selected_nodes = [
        node for node in tree.body if isinstance(node, ast.FunctionDef) and node.name in required_names
    ]
    module = ast.Module(body=selected_nodes, type_ignores=[])
    namespace: dict[str, object] = {
        "io": io,
        "pd": pd,
        "Alignment": Alignment,
        "BarChart": BarChart,
        "Border": Border,
        "Font": Font,
        "LineChart": LineChart,
        "PatternFill": PatternFill,
        "Reference": Reference,
        "Side": Side,
        "st": SimpleNamespace(cache_data=lambda **kwargs: (lambda func: func)),
        "get_column_letter": get_column_letter,
    }
    exec(compile(module, filename=str(source_path), mode="exec"), namespace)
    return namespace


class ChartExcelExportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.export_functions = load_chart_export_functions()

    def test_build_excel_chart_workbook_creates_native_line_chart(self) -> None:
        date_summary = pd.DataFrame(
            [
                {
                    "Analysis Date": pd.Timestamp("2026-04-21"),
                    "Quantity_Prev": 80.0,
                    "Quantity_Curr": 100.0,
                    "Delta": 20.0,
                },
                {
                    "Analysis Date": pd.Timestamp("2026-04-22"),
                    "Quantity_Prev": 90.0,
                    "Quantity_Curr": 110.0,
                    "Delta": 20.0,
                },
            ]
        )
        build_excel_chart_workbook = self.export_functions["build_excel_chart_workbook"]
        workbook_bytes = build_excel_chart_workbook(
            "dashboard_trend",
            date_summary,
            {
                "title": "Dashboard - Release trend",
                "dataset": date_summary,
                "chart_type": "line",
                "category_column": "Analysis Date",
                "series_columns": ["Quantity_Prev", "Quantity_Curr"],
                "x_axis_title": "Receipt Date",
                "y_axis_title": "Open quantity",
                "filter_summary": "Data: 2026-04-21 - 2026-04-22",
                "filtered_record_count": 2,
                "generated_at": "2026-04-23 10:00:00 CEST",
            },
        )

        workbook = load_workbook(io.BytesIO(workbook_bytes))
        self.assertIn("Data", workbook.sheetnames)
        self.assertIn("Chart", workbook.sheetnames)
        chart_sheet = workbook["Chart"]
        self.assertEqual(chart_sheet["A1"].value, "Dashboard - Release trend")
        self.assertGreaterEqual(len(chart_sheet._charts), 1)

    def test_build_matrix_chart_export_table_sums_date_columns(self) -> None:
        matrix_df = pd.DataFrame(
            [
                {"Part Number": "MAT-001", "Part Description": "Mercury", "2026-04-21": 10.0, "2026-04-22": 15.0},
                {"Part Number": "MAT-002", "Part Description": "Venus", "2026-04-21": -5.0, "2026-04-22": 20.0},
            ]
        )
        build_matrix_chart_export_table = self.export_functions["build_matrix_chart_export_table"]
        chart_df = build_matrix_chart_export_table(matrix_df, "Release Delta")

        self.assertEqual(list(chart_df.columns), ["Date", "Release Delta"])
        self.assertEqual(chart_df.iloc[0]["Release Delta"], 5.0)
        self.assertEqual(chart_df.iloc[1]["Release Delta"], 35.0)


if __name__ == "__main__":
    unittest.main()
