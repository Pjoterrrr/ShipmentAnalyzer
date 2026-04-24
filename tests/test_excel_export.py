from __future__ import annotations

import ast
import io
import unittest
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from analytics_calendar import (
    build_calendar_frame,
    build_weekly_summary,
    classify_polish_day,
    get_last_completed_reference_week,
)


def load_export_functions() -> dict[str, object]:
    source_path = Path(__file__).resolve().parents[1] / "streamlit_app.py"
    source = source_path.read_text(encoding="utf-8")
    tree = ast.parse(source, filename=str(source_path))
    required_names = {
        "blend_hex",
        "style_value",
        "format_date",
        "format_signed_int",
        "format_percent_display",
        "get_date_label",
        "format_release_label",
        "format_release_summary",
        "get_reference_week_rows",
        "logo_available",
        "insert_logo",
        "style_excel_header",
        "autosize_worksheet",
        "decorate_delta_column",
        "excel_fill_color",
        "ensure_numeric_cells_black",
        "build_weekly_comparison_export",
        "classify_weekly_change",
        "format_weekly_change_label",
        "build_weekly_by_part_report",
        "build_weekly_by_part_chart_source",
        "build_qty_matrix_report",
        "build_weekly_delta_map_report",
        "build_weekly_delta_matrix_report",
        "build_report_matrix_export",
        "build_matrix_totals_export",
        "build_calendar_operational_export",
        "build_calendar_weekly_export",
        "style_table_region",
        "apply_number_formats",
        "decorate_trend_columns",
        "style_multi_label_matrix_sheet",
        "write_parameter_section",
        "write_dataframe_block",
        "write_weekly_by_part_sheet",
        "write_qty_matrix_sheet",
        "write_weekly_delta_map_sheet",
        "write_weekly_delta_matrix_sheet",
        "to_professional_weekly_report_bytes",
        "apply_polish_calendar_highlights",
        "style_matrix_sheet",
        "highlight_calendar_rows",
        "highlight_weekly_rows",
        "write_summary_sheet",
        "to_excel_bytes",
    }
    selected_nodes = [
        node for node in tree.body if isinstance(node, ast.FunctionDef) and node.name in required_names
    ]
    module = ast.Module(body=selected_nodes, type_ignores=[])
    namespace: dict[str, object] = {
        "io": io,
        "pd": pd,
        "OpenpyxlImage": OpenpyxlImage,
        "Alignment": Alignment,
        "Border": Border,
        "Font": Font,
        "PatternFill": PatternFill,
        "Side": Side,
        "get_column_letter": get_column_letter,
        "build_calendar_frame": build_calendar_frame,
        "classify_polish_day": classify_polish_day,
        "get_last_completed_reference_week": get_last_completed_reference_week,
        "BRAND_NAME": "Pjoter Development",
        "DATE_LABELS": {"Receipt Date": "Receipt Date", "Ship Date": "Ship Date"},
        "LOGO_PATH": Path("__missing_logo__.png"),
    }
    exec(compile(module, filename=str(source_path), mode="exec"), namespace)
    return namespace


def rgb_suffix(style_value) -> str:
    if style_value is None:
        return ""
    return str(style_value).upper()[-6:]


class ExcelExportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.export_functions = load_export_functions()

    def test_excel_export_contains_filtered_report_sheets_without_charts(self) -> None:
        detail_df = pd.DataFrame(
            [
                {
                    "PO Number": "PO-1",
                    "Part Number": "A1",
                    "Part Description": "Produkt A",
                    "Ship Date": pd.Timestamp("2026-05-01"),
                    "Receipt Date": pd.Timestamp("2026-05-04"),
                    "Unit of Measure": "HL",
                    "Quantity_Prev": 80.0,
                    "Quantity_Curr": 100.0,
                    "Delta": 20.0,
                    "Percent Change": 25.0,
                    "Change Direction": "Increase",
                    "Alert": True,
                    "Product Label": "A1 | Produkt A",
                },
                {
                    "PO Number": "PO-1",
                    "Part Number": "A1",
                    "Part Description": "Produkt A",
                    "Ship Date": pd.Timestamp("2026-05-02"),
                    "Receipt Date": pd.Timestamp("2026-05-05"),
                    "Unit of Measure": "HL",
                    "Quantity_Prev": 90.0,
                    "Quantity_Curr": 120.0,
                    "Delta": 30.0,
                    "Percent Change": 33.33,
                    "Change Direction": "Increase",
                    "Alert": True,
                    "Product Label": "A1 | Produkt A",
                },
            ]
        )
        weekly_summary = build_weekly_summary(
            detail_df,
            "Receipt Date",
            date(2026, 4, 27),
            date(2026, 5, 10),
            date(2026, 5, 10),
            15,
        )
        current_matrix_df = pd.DataFrame({"2026-05-04": [100.0], "2026-05-05": [120.0]}, index=["A1 | Produkt A"])
        delta_matrix_df = pd.DataFrame({"2026-05-04": [20.0], "2026-05-05": [30.0]}, index=["A1 | Produkt A"])
        prev_meta = {
            "po_number": "PO-1",
            "release_version": "15",
            "release_date": pd.Timestamp("2026-04-20"),
            "planner_name": "Planner",
            "planner_email": "planner@example.com",
        }
        curr_meta = {
            "po_number": "PO-1",
            "release_version": "16",
            "release_date": pd.Timestamp("2026-04-27"),
            "planner_name": "Planner",
            "planner_email": "planner@example.com",
        }
        product_summary = pd.DataFrame(
            [
                {
                    "Part Number": "A1",
                    "Part Description": "Produkt A",
                    "Product Label": "A1 | Produkt A",
                    "Quantity_Prev": 170.0,
                    "Quantity_Curr": 220.0,
                    "Delta": 50.0,
                    "Abs_Delta": 50.0,
                    "Alert_Count": 2,
                    "Change Direction": "Increase",
                }
            ]
        )

        to_excel_bytes = self.export_functions["to_excel_bytes"]
        excel_bytes = to_excel_bytes(
            detail_df,
            weekly_summary,
            current_matrix_df,
            delta_matrix_df,
            prev_meta,
            curr_meta,
            product_summary,
            "Receipt Date",
            date(2026, 4, 27),
            date(2026, 5, 10),
            [{"label": "Test", "title": "Produkt A", "copy": "Zmiana tygodniowa"}],
        )

        workbook = load_workbook(io.BytesIO(excel_bytes))
        self.assertIn("Weekly Comparison", workbook.sheetnames)
        self.assertIn("Weekly by Part", workbook.sheetnames)
        self.assertIn("Weekly Delta Map", workbook.sheetnames)
        self.assertIn("Weekly Delta Matrix", workbook.sheetnames)
        self.assertIn("Calendar PL", workbook.sheetnames)
        self.assertIn("Current Matrix", workbook.sheetnames)
        self.assertIn("Delta Heatmap", workbook.sheetnames)
        for worksheet in workbook.worksheets:
            self.assertEqual(len(worksheet._charts), 0, worksheet.title)

        calendar_sheet = workbook["Calendar PL"]
        calendar_rows = {
            calendar_sheet.cell(row=row, column=1).value: row
            for row in range(2, calendar_sheet.max_row + 1)
        }
        may_first_row = calendar_rows["2026-05-01"]
        may_third_row = calendar_rows["2026-05-03"]
        self.assertEqual(rgb_suffix(calendar_sheet.cell(row=may_first_row, column=1).fill.fgColor.rgb), "FEF3C7")
        self.assertEqual(rgb_suffix(calendar_sheet.cell(row=may_third_row, column=1).fill.fgColor.rgb), "FEF3C7")

        weekly_sheet = workbook["Weekly Comparison"]
        weekly_header = {cell.value: cell.column for cell in weekly_sheet[1]}
        current_release_cell = weekly_sheet.cell(row=2, column=weekly_header["Current Release Qty"])
        self.assertEqual(rgb_suffix(current_release_cell.font.color.rgb), "000000")

        weekly_delta_sheet = workbook["Weekly Delta Map"]
        weekly_delta_header = {cell.value: cell.column for cell in weekly_delta_sheet[1]}
        self.assertIn("Weekly Delta", weekly_delta_header)
        self.assertIn("Weekly Change %", weekly_delta_header)
        self.assertIn("Trend / Status", weekly_delta_header)

    def test_professional_weekly_report_contains_weekly_delta_sheets_without_charts(self) -> None:
        detail_df = pd.DataFrame(
            [
                {
                    "Part Number": "A1",
                    "Part Description": "Produkt A",
                    "Receipt Date": pd.Timestamp("2026-04-20"),
                    "Ship Date": pd.Timestamp("2026-04-17"),
                    "Quantity_Prev": 80.0,
                    "Quantity_Curr": 100.0,
                    "Delta": 20.0,
                    "Product Label": "A1 | Produkt A",
                },
                {
                    "Part Number": "A1",
                    "Part Description": "Produkt A",
                    "Receipt Date": pd.Timestamp("2026-04-27"),
                    "Ship Date": pd.Timestamp("2026-04-24"),
                    "Quantity_Prev": 120.0,
                    "Quantity_Curr": 140.0,
                    "Delta": 20.0,
                    "Product Label": "A1 | Produkt A",
                },
                {
                    "Part Number": "B2",
                    "Part Description": "Produkt B",
                    "Receipt Date": pd.Timestamp("2026-04-27"),
                    "Ship Date": pd.Timestamp("2026-04-24"),
                    "Quantity_Prev": 0.0,
                    "Quantity_Curr": 30.0,
                    "Delta": 30.0,
                    "Product Label": "B2 | Produkt B",
                },
            ]
        )
        prev_meta = {
            "po_number": "PO-1",
            "release_version": "15",
            "release_date": pd.Timestamp("2026-04-13"),
            "planner_name": "Planner",
            "planner_email": "planner@example.com",
        }
        curr_meta = {
            "po_number": "PO-1",
            "release_version": "16",
            "release_date": pd.Timestamp("2026-04-20"),
            "planner_name": "Planner",
            "planner_email": "planner@example.com",
        }
        to_professional_weekly_report_bytes = self.export_functions["to_professional_weekly_report_bytes"]
        excel_bytes = to_professional_weekly_report_bytes(
            detail_df,
            prev_meta,
            curr_meta,
            "Receipt Date",
            date(2026, 4, 20),
            date(2026, 5, 3),
        )
        workbook = load_workbook(io.BytesIO(excel_bytes))

        self.assertIn("Weekly by Part", workbook.sheetnames)
        self.assertIn("Weekly Qty Matrix", workbook.sheetnames)
        self.assertIn("Weekly Delta Map", workbook.sheetnames)
        self.assertIn("Weekly Delta Matrix", workbook.sheetnames)
        for worksheet in workbook.worksheets:
            self.assertEqual(len(worksheet._charts), 0, worksheet.title)

        delta_sheet = workbook["Weekly Delta Map"]
        headers = {cell.value: cell.column for cell in delta_sheet[11]}
        self.assertIn("Weekly Delta", headers)
        self.assertIn("Weekly Change %", headers)
        self.assertIn("Trend / Status", headers)


if __name__ == "__main__":
    unittest.main()
